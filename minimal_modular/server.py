"""
PhD Research Extraction Tool - Backend API Server
Flask-based REST API with SQLite persistence
Spawns minimal_modular extract.py as subprocess for runs
"""

import os
import json
import uuid
import time
import threading
import subprocess
import hashlib
import secrets
import shutil
import re
from datetime import datetime, timezone, timedelta
from functools import wraps
from flask import Flask, request, jsonify, Response, send_file, stream_with_context, g
from flask_cors import CORS
from werkzeug.utils import secure_filename
import sqlite3

# JWT secret key - in production, use environment variable
JWT_SECRET = os.environ.get("JWT_SECRET", "cretextract-dev-secret-key-change-in-production")
JWT_EXPIRY_HOURS = 24

# Import from minimal_modular
import sys
sys.path.insert(0, os.path.dirname(__file__))
from cache_utils import get_cache_stats, clear_cache, CACHE_DIR
from config import (
    LLM_PROVIDER, OPENAI_API_KEY, GEMINI_API_KEY, ANTHROPIC_API_KEY, 
    DEEPSEEK_API_KEY, DATALAB_API_KEY
)


# ============================================================================
# App Configuration
# ============================================================================

app = Flask(__name__)
CORS(app, supports_credentials=True, origins="*")

# Config
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), "uploads")
EXPORTS_FOLDER = os.path.join(os.path.dirname(__file__), "exports")
DB_PATH = os.path.join(os.path.dirname(__file__), "app.db")
IPC_DIR = os.path.join(os.path.dirname(__file__), "ipc")
EXTRACT_SCRIPT = os.path.join(os.path.dirname(__file__), "extract.py")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(EXPORTS_FOLDER, exist_ok=True)
os.makedirs(IPC_DIR, exist_ok=True)

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

# Global log buffer for SSE
log_buffer = []
log_buffer_lock = threading.Lock()

# Active run processes
active_processes = {}  # run_id -> subprocess.Popen

# ============================================================================
# Database Setup
# ============================================================================

def get_db():
    """Get thread-local database connection with timeout for concurrency."""
    conn = sqlite3.connect(DB_PATH, timeout=30.0)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=30000")
    return conn

def init_db():
    """Initialize database tables."""
    conn = get_db()
    cur = conn.cursor()
    
    # Users table for authentication
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id TEXT PRIMARY KEY,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            created_at TEXT,
            is_active INTEGER DEFAULT 1
        )
    """)
    
    # Runs table - stores file IDs (not paths) for frontend safety
    # source_type: 'pdf' (ZIP upload), 'links' (manual URLs), 'deep_research' (Gemini search)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS runs (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            source_type TEXT DEFAULT 'pdf',
            status TEXT DEFAULT 'waiting',
            start_date TEXT,
            sources_count INTEGER DEFAULT 0,
            data_entries_count INTEGER DEFAULT 0,
            llm_provider TEXT,
            pdfs_dir TEXT,
            excel_path TEXT,
            output_dir TEXT,
            prompt TEXT,
            search_methods TEXT,
            search_queries TEXT,
            links TEXT,
            table_file_url TEXT,
            per_link_prompt TEXT,
            schema_file_id TEXT,
            zip_file_id TEXT,
            deep_research_query TEXT,
            deep_research_result TEXT,
            deep_research_interaction_id TEXT,
            user_id TEXT,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    """)

    # Migration: rename legacy runs count column to sources_count (idempotent)
    try:
        cur.execute("PRAGMA table_info(runs)")
        cols = {r[1] for r in cur.fetchall()}
        legacy_count_col = "articles" + "_count"
        if legacy_count_col in cols and "sources_count" not in cols:
            cur.execute(f"ALTER TABLE runs RENAME COLUMN {legacy_count_col} TO sources_count")
    except Exception:
        pass

    # Migration: rename legacy sources table (articles -> sources) (idempotent)
    try:
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='sources'")
        has_sources = cur.fetchone() is not None
        if not has_sources:
            cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='articles'")
            has_articles = cur.fetchone() is not None
            if has_articles:
                cur.execute("ALTER TABLE articles RENAME TO sources")
    except Exception:
        pass

    # Meta Sources table - provenance/method that aggregated/produced sources for a run
    cur.execute("""
        CREATE TABLE IF NOT EXISTS meta_sources (
            id TEXT PRIMARY KEY,
            run_id TEXT,
            method TEXT,
            name TEXT,
            query TEXT,
            config_json TEXT,
            created_at TEXT,
            updated_at TEXT,
            user_id TEXT,
            FOREIGN KEY (run_id) REFERENCES runs(id),
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    """)

    # Sources table - stores document sources (link/pdf) with HTML content in SQL
    cur.execute("""
        CREATE TABLE IF NOT EXISTS sources (
            id TEXT PRIMARY KEY,
            run_id TEXT,
            crawl_job_id TEXT,
            url TEXT,
            domain TEXT,
            title TEXT,
            html_content TEXT,
            pdf_file_id TEXT,
            source_type TEXT,
            status TEXT DEFAULT 'PENDING',
            error TEXT,
            meta_source_id TEXT,
            content_type TEXT DEFAULT 'html',
            created_at TEXT,
            updated_at TEXT,
            FOREIGN KEY (run_id) REFERENCES runs(id),
            FOREIGN KEY (crawl_job_id) REFERENCES crawl_jobs(id)
        )
    """)

    # Idempotent schema upgrades for existing sources tables
    for stmt in [
        "ALTER TABLE sources ADD COLUMN source_type TEXT",
        "ALTER TABLE sources ADD COLUMN status TEXT DEFAULT 'PENDING'",
        "ALTER TABLE sources ADD COLUMN error TEXT",
        "ALTER TABLE sources ADD COLUMN meta_source_id TEXT",
        "ALTER TABLE sources ADD COLUMN updated_at TEXT",
        "ALTER TABLE sources ADD COLUMN row_count INTEGER",
        "ALTER TABLE sources ADD COLUMN row_count_logic TEXT",
        "ALTER TABLE sources ADD COLUMN rejection_reason TEXT",
    ]:
        try:
            cur.execute(stmt)
        except Exception:
            pass

    # Idempotent schema upgrades for runs (meta_source_id)
    try:
        cur.execute("ALTER TABLE runs ADD COLUMN meta_source_id TEXT")
    except Exception:
        pass

    # Idempotent schema upgrades for runs (prompt file IDs and validation)
    for stmt in [
        "ALTER TABLE runs ADD COLUMN extraction_prompt_file_id TEXT",
        "ALTER TABLE runs ADD COLUMN validation_prompt_file_id TEXT",
        "ALTER TABLE runs ADD COLUMN validation_enabled INTEGER DEFAULT 0",
        "ALTER TABLE runs ADD COLUMN validation_max_retries INTEGER DEFAULT 3",
        "ALTER TABLE runs ADD COLUMN validation_pass_rate REAL",
        "ALTER TABLE runs ADD COLUMN validation_accepted_count INTEGER",
        "ALTER TABLE runs ADD COLUMN validation_rejected_count INTEGER",
        "ALTER TABLE runs ADD COLUMN cache_flags TEXT",
    ]:
        try:
            cur.execute(stmt)
        except Exception:
            pass
    
    # Exports table
    cur.execute("""
        CREATE TABLE IF NOT EXISTS exports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            run_id TEXT,
            created_at TEXT,
            file_path TEXT
        )
    """)
    
    # Domains table
    cur.execute("""
        CREATE TABLE IF NOT EXISTS domains (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE,
            visited_count INTEGER DEFAULT 0,
            success_count INTEGER DEFAULT 0,
            failed_count INTEGER DEFAULT 0
        )
    """)
    
    # Deep Research runs table - Gemini Deep Research API integration
    cur.execute("""
        CREATE TABLE IF NOT EXISTS deep_research_runs (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            status TEXT DEFAULT 'pending',
            interaction_id TEXT,
            query TEXT NOT NULL,
            search_config TEXT,
            result_text TEXT,
            extracted_links TEXT,
            logs TEXT,
            created_at TEXT,
            started_at TEXT,
            completed_at TEXT,
            error TEXT,
            user_id TEXT,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    """)
    
    # Config table - per-user configuration with enhanced metadata
    # user_id NULL = global default, user_id set = user override
    cur.execute("""
        CREATE TABLE IF NOT EXISTS config (
            key TEXT NOT NULL,
            user_id TEXT,
            value TEXT,
            value_type TEXT DEFAULT 'string',
            input_type TEXT DEFAULT 'text',
            allowed_values TEXT,
            default_value TEXT,
            category TEXT DEFAULT 'general',
            description TEXT DEFAULT '',
            sensitive INTEGER DEFAULT 0,
            required INTEGER DEFAULT 0,
            display_order INTEGER DEFAULT 0,
            last_modified TEXT,
            PRIMARY KEY (key, user_id)
        )
    """)
    
    # Cache providers table
    cur.execute("""
        CREATE TABLE IF NOT EXISTS cache_providers (
            id TEXT PRIMARY KEY,
            name TEXT,
            type TEXT DEFAULT 'LLM',
            entries_count INTEGER DEFAULT 0,
            total_size_bytes INTEGER DEFAULT 0,
            hit_rate REAL DEFAULT 0.0,
            last_accessed TEXT
        )
    """)
    
    # Cache entries table
    cur.execute("""
        CREATE TABLE IF NOT EXISTS cache_entries (
            id TEXT PRIMARY KEY,
            provider_id TEXT,
            key TEXT,
            size_bytes INTEGER DEFAULT 0,
            hit_count INTEGER DEFAULT 0,
            created_date TEXT,
            last_accessed TEXT,
            status TEXT DEFAULT 'ACTIVE',
            FOREIGN KEY (provider_id) REFERENCES cache_providers(id)
        )
    """)
    
    # Logs table
    cur.execute("""
        CREATE TABLE IF NOT EXISTS logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at TEXT,
            level TEXT DEFAULT 'INFO',
            message TEXT,
            run_id TEXT
        )
    """)
    
    # Files table - maps file IDs to paths (paths never exposed to frontend)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS files (
            id TEXT PRIMARY KEY,
            filename TEXT NOT NULL,
            original_name TEXT,
            mime_type TEXT,
            size_bytes INTEGER DEFAULT 0,
            file_type TEXT,
            run_id TEXT,
            created_at TEXT,
            FOREIGN KEY (run_id) REFERENCES runs(id)
        )
    """)
    
    # Crawl jobs table - jobs for Chrome extension to process
    cur.execute("""
        CREATE TABLE IF NOT EXISTS crawl_jobs (
            id TEXT PRIMARY KEY,
            deep_research_id TEXT,
            run_id TEXT,
            user_id TEXT NOT NULL,
            url TEXT NOT NULL,
            title TEXT,
            status TEXT DEFAULT 'PENDING',
            html TEXT,
            pdf_path TEXT,
            error TEXT,
            attempts INTEGER DEFAULT 0,
            created_at TEXT,
            claimed_at TEXT,
            claim_expires_at TEXT,
            completed_at TEXT,
            FOREIGN KEY (deep_research_id) REFERENCES deep_research_runs(id) ON DELETE CASCADE,
            FOREIGN KEY (run_id) REFERENCES runs(id) ON DELETE CASCADE,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    """)
    
    # Domain scripts table - per-domain crawl scripts
    cur.execute("""
        CREATE TABLE IF NOT EXISTS domain_scripts (
            id TEXT PRIMARY KEY,
            domain TEXT NOT NULL,
            user_id TEXT,
            script TEXT,
            condition TEXT,
            wait_before_ms INTEGER DEFAULT 0,
            wait_after_ms INTEGER DEFAULT 0,
            created_at TEXT,
            updated_at TEXT,
            UNIQUE(domain, user_id)
        )
    """)
    
    # Create indexes for crawl_jobs
    cur.execute("CREATE INDEX IF NOT EXISTS idx_crawl_jobs_user_status ON crawl_jobs(user_id, status)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_crawl_jobs_deep_research ON crawl_jobs(deep_research_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_crawl_jobs_run ON crawl_jobs(run_id)")
    
    # Migrate: Add new columns if they don't exist
    try:
        cur.execute("ALTER TABLE runs ADD COLUMN schema_file_id TEXT")
    except:
        pass
    try:
        cur.execute("ALTER TABLE runs ADD COLUMN zip_file_id TEXT")
    except:
        pass
    try:
        cur.execute("ALTER TABLE runs ADD COLUMN enable_row_counting INTEGER DEFAULT 0")
    except:
        pass
    try:
        cur.execute("ALTER TABLE runs ADD COLUMN user_id TEXT")
    except:
        pass
    
    # Migrate logs table for extension logging
    try:
        cur.execute("ALTER TABLE logs ADD COLUMN source TEXT DEFAULT 'server'")
    except:
        pass
    try:
        cur.execute("ALTER TABLE logs ADD COLUMN context TEXT")
    except:
        pass
    
    # Migrate sources table (physical table name: sources) for crawled content
    try:
        cur.execute("ALTER TABLE sources ADD COLUMN deep_research_id TEXT")
    except:
        pass
    try:
        cur.execute("ALTER TABLE sources ADD COLUMN crawl_job_id TEXT")
    except:
        pass
    try:
        cur.execute("ALTER TABLE sources ADD COLUMN title TEXT")
    except:
        pass
    try:
        cur.execute("ALTER TABLE sources ADD COLUMN html_content TEXT")
    except:
        pass
    try:
        cur.execute("ALTER TABLE sources ADD COLUMN pdf_file_id TEXT")
    except:
        pass
    try:
        cur.execute("ALTER TABLE sources ADD COLUMN content_type TEXT DEFAULT 'html'")
    except:
        pass
    try:
        cur.execute("ALTER TABLE sources ADD COLUMN created_at TEXT")
    except:
        pass
    
    # Migrate runs table for unified source types
    try:
        cur.execute("ALTER TABLE runs ADD COLUMN source_type TEXT DEFAULT 'pdf'")
    except:
        pass
    try:
        cur.execute("ALTER TABLE runs ADD COLUMN deep_research_query TEXT")
    except:
        pass
    try:
        cur.execute("ALTER TABLE runs ADD COLUMN deep_research_result TEXT")
    except:
        pass
    try:
        cur.execute("ALTER TABLE runs ADD COLUMN deep_research_interaction_id TEXT")
    except:
        pass
    
    conn.commit()
    conn.close()


def process_pdf_sources_for_run(run_id: str, user_id: str = None):
    """Convert PENDING/PROCESSING PDF sources to READY by generating html_content."""
    try:
        from pdf_converter import convert_pdf_to_text
    except Exception:
        convert_pdf_to_text = None

    if convert_pdf_to_text is None:
        return

    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            SELECT id, pdf_file_id, title
            FROM sources
            WHERE run_id = ? AND source_type = 'pdf' AND pdf_file_id IS NOT NULL AND (status IS NULL OR status IN ('PENDING', 'PROCESSING'))
            """,
            (run_id,),
        )
        rows = cur.fetchall()
        for r in rows:
            source_id = r["id"]
            pdf_file_id = r["pdf_file_id"]
            title = r["title"] or "PDF"
            pdf_path = get_file_internal_path(pdf_file_id)
            if not pdf_path or not os.path.exists(pdf_path):
                continue

            now = datetime.now(timezone.utc).isoformat()
            try:
                cur.execute(
                    "UPDATE sources SET status = 'PROCESSING', updated_at = ? WHERE id = ?",
                    (now, source_id),
                )
                conn.commit()
            except Exception:
                pass

            try:
                pdf_text = convert_pdf_to_text(pdf_path, use_cache=True)
                if not pdf_text or len(str(pdf_text).strip()) == 0:
                    raise ValueError("Empty PDF text")

                html_content = f"""<!DOCTYPE html>
<html>
<head><title>{title}</title></head>
<body>
<source>
<h1>{title}</h1>
<div class=\"pdf-content\">\n{pdf_text}\n</div>
</source>
</body>
</html>"""

                now = datetime.now(timezone.utc).isoformat()
                cur.execute(
                    """
                    UPDATE sources
                    SET html_content = ?, content_type = 'pdf', status = 'READY', error = NULL, updated_at = ?
                    WHERE id = ?
                    """,
                    (html_content, now, source_id),
                )
                conn.commit()
            except Exception as e:
                now = datetime.now(timezone.utc).isoformat()
                try:
                    cur.execute(
                        "UPDATE sources SET status = 'FAILED', error = ?, updated_at = ? WHERE id = ?",
                        (str(e)[:500], now, source_id),
                    )
                    conn.commit()
                except Exception:
                    pass
    finally:
        conn.close()


def create_meta_source(run_id: str, method: str, user_id: str = None, name: str = None, query: str = None, config: dict = None) -> str:
    meta_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO meta_sources (id, run_id, method, name, query, config_json, created_at, updated_at, user_id)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            meta_id,
            run_id,
            method,
            name or method,
            query,
            json.dumps(config or {}),
            now,
            now,
            user_id,
        ),
    )
    conn.commit()
    conn.close()
    return meta_id


def ensure_source_row(
    *,
    source_id: str,
    run_id: str,
    source_type: str,
    status: str,
    url: str = None,
    title: str = None,
    crawl_job_id: str = None,
    pdf_file_id: str = None,
    meta_source_id: str = None,
    cur=None,
):
    """Insert a source row if it doesn't exist. Optionally reuse an existing cursor."""
    now = datetime.now(timezone.utc).isoformat()
    own_conn = False
    if cur is None:
        conn = get_db()
        cur = conn.cursor()
        own_conn = True
    else:
        conn = None
    cur.execute("SELECT id FROM sources WHERE id = ?", (source_id,))
    exists = cur.fetchone()
    if exists:
        if own_conn:
            conn.close()
        return
    cur.execute(
        """
        INSERT INTO sources (id, run_id, crawl_job_id, url, domain, title, html_content, pdf_file_id, source_type, status, error, meta_source_id, content_type, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            source_id,
            run_id,
            crawl_job_id,
            url,
            "",
            title,
            "",
            pdf_file_id,
            source_type,
            status,
            None,
            meta_source_id,
            "pdf" if source_type == "pdf" else "html",
            now,
            now,
        ),
    )
    if own_conn:
        conn.commit()
        conn.close()


# ============================================================================
# Authentication Utilities
# ============================================================================

def hash_password(password: str) -> str:
    """Hash password using SHA-256 with salt."""
    salt = secrets.token_hex(16)
    hash_obj = hashlib.sha256((salt + password).encode())
    return f"{salt}:{hash_obj.hexdigest()}"

def verify_password(password: str, password_hash: str) -> bool:
    """Verify password against stored hash."""
    try:
        salt, stored_hash = password_hash.split(":")
        hash_obj = hashlib.sha256((salt + password).encode())
        return hash_obj.hexdigest() == stored_hash
    except:
        return False

def create_token(user_id: str, email: str) -> str:
    """Create a simple JWT-like token (base64 encoded JSON with signature)."""
    import base64
    import hmac
    
    payload = {
        "user_id": user_id,
        "email": email,
        "exp": (datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRY_HOURS)).isoformat()
    }
    payload_json = json.dumps(payload)
    payload_b64 = base64.urlsafe_b64encode(payload_json.encode()).decode()
    
    signature = hmac.new(JWT_SECRET.encode(), payload_b64.encode(), hashlib.sha256).hexdigest()
    
    return f"{payload_b64}.{signature}"

def verify_token(token: str) -> dict | None:
    """Verify token and return payload if valid."""
    import base64
    import hmac
    
    try:
        parts = token.split(".")
        if len(parts) != 2:
            return None
        
        payload_b64, signature = parts
        
        # Verify signature
        expected_sig = hmac.new(JWT_SECRET.encode(), payload_b64.encode(), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(signature, expected_sig):
            return None
        
        # Decode payload
        payload_json = base64.urlsafe_b64decode(payload_b64.encode()).decode()
        payload = json.loads(payload_json)
        
        # Check expiry
        exp = datetime.fromisoformat(payload["exp"])
        if datetime.now(timezone.utc) > exp:
            return None
        
        return payload
    except:
        return None

def get_current_user():
    """Get current user from request Authorization header."""
    auth_header = request.headers.get("Authorization", "")
    if not auth_header.startswith("Bearer "):
        return None
    
    token = auth_header[7:]  # Remove "Bearer " prefix
    payload = verify_token(token)
    if not payload:
        return None
    
    return {"id": payload["user_id"], "email": payload["email"]}

def require_auth(f):
    """Decorator to require authentication for a route."""
    @wraps(f)
    def decorated(*args, **kwargs):
        user = get_current_user()
        if not user:
            return jsonify({"error": "Authentication required"}), 401
        g.current_user = user
        return f(*args, **kwargs)
    return decorated

def optional_auth(f):
    """Decorator to optionally get user if authenticated."""
    @wraps(f)
    def decorated(*args, **kwargs):
        g.current_user = get_current_user()
        return f(*args, **kwargs)
    return decorated

# Initialize DB on startup
init_db()

def seed_default_config():
    """Seed global default configuration values for CreteXtract production deployment."""
    defaults = [
        # ============================================================================
        # LLM Configuration
        # ============================================================================
        {
            "key": "LLM_PROVIDER",
            "value": "gemini",
            "value_type": "string",
            "input_type": "select",
            "allowed_values": json.dumps(["openai", "gemini", "anthropic", "deepseek", "ollama"]),
            "default_value": "gemini",
            "category": "llm",
            "description": "Primary LLM provider for extraction tasks",
            "required": 1,
            "display_order": 1
        },
        {
            "key": "LLM_MODEL",
            "value": "gemini-2.0-flash",
            "value_type": "string",
            "input_type": "text",
            "default_value": "gemini-2.0-flash",
            "category": "llm",
            "description": "Model identifier for the selected LLM provider",
            "required": 1,
            "display_order": 2
        },
        {
            "key": "LLM_TEMPERATURE",
            "value": "0.1",
            "value_type": "number",
            "input_type": "number",
            "default_value": "0.1",
            "category": "llm",
            "description": "Temperature for LLM responses (0.0-1.0, lower = more deterministic)",
            "display_order": 3
        },
        {
            "key": "LLM_MAX_TOKENS",
            "value": "4096",
            "value_type": "number",
            "input_type": "number",
            "default_value": "4096",
            "category": "llm",
            "description": "Maximum tokens for LLM response generation",
            "display_order": 4
        },
        {
            "key": "LLM_TIMEOUT",
            "value": "120",
            "value_type": "number",
            "input_type": "number",
            "default_value": "120",
            "category": "llm",
            "description": "Timeout in seconds for LLM API calls",
            "display_order": 5
        },
        {
            "key": "OLLAMA_BASE_URL",
            "value": "http://localhost:11434",
            "value_type": "string",
            "input_type": "text",
            "default_value": "http://localhost:11434",
            "category": "llm",
            "description": "Base URL for Ollama local LLM server",
            "display_order": 6
        },
        # ============================================================================
        # API Keys
        # ============================================================================
        {
            "key": "OPENAI_API_KEY",
            "value": "",
            "value_type": "string",
            "input_type": "secret",
            "category": "api_keys",
            "description": "OpenAI API key for GPT models",
            "sensitive": 1,
            "display_order": 100
        },
        {
            "key": "GEMINI_API_KEY",
            "value": "",
            "value_type": "string",
            "input_type": "secret",
            "category": "api_keys",
            "description": "Google Gemini API key",
            "sensitive": 1,
            "display_order": 101
        },
        {
            "key": "ANTHROPIC_API_KEY",
            "value": "",
            "value_type": "string",
            "input_type": "secret",
            "category": "api_keys",
            "description": "Anthropic Claude API key",
            "sensitive": 1,
            "display_order": 102
        },
        {
            "key": "DEEPSEEK_API_KEY",
            "value": "",
            "value_type": "string",
            "input_type": "secret",
            "category": "api_keys",
            "description": "DeepSeek API key",
            "sensitive": 1,
            "display_order": 103
        },
        {
            "key": "DATALAB_API_KEY",
            "value": "",
            "value_type": "string",
            "input_type": "secret",
            "category": "api_keys",
            "description": "Datalab Marker API key for PDF-to-Markdown conversion",
            "sensitive": 1,
            "required": 1,
            "display_order": 104
        },
        # ============================================================================
        # PDF Processing & Extraction
        # ============================================================================
        {
            "key": "PDF_PROCESSOR",
            "value": "marker",
            "value_type": "string",
            "input_type": "select",
            "allowed_values": json.dumps(["marker", "pymupdf", "pdfplumber"]),
            "default_value": "marker",
            "category": "extraction",
            "description": "PDF processing backend for text extraction",
            "required": 1,
            "display_order": 10
        },
        {
            "key": "ENABLE_ROW_COUNTING",
            "value": "false",
            "value_type": "boolean",
            "input_type": "switch",
            "default_value": "false",
            "category": "extraction",
            "description": "Enable row counting pre-analysis phase before extraction",
            "display_order": 11
        },
        {
            "key": "MAX_RETRIES",
            "value": "3",
            "value_type": "number",
            "input_type": "number",
            "default_value": "3",
            "category": "extraction",
            "description": "Maximum retry attempts for failed extraction chunks",
            "display_order": 12
        },
        {
            "key": "RETRY_FOR_CACHED_EMPTY_LIST",
            "value": "true",
            "value_type": "boolean",
            "input_type": "switch",
            "default_value": "true",
            "category": "extraction",
            "description": "Retry extraction when cached LLM response returns empty list",
            "display_order": 13
        },
        {
            "key": "VALIDATION_RETRIES",
            "value": "1",
            "value_type": "number",
            "input_type": "number",
            "default_value": "1",
            "category": "extraction",
            "description": "Number of validation retry attempts per source (0 = no validation, 1 = validate once, 2+ = retry on failure)",
            "display_order": 14
        },
        {
            "key": "CHUNK_SIZE",
            "value": "10",
            "value_type": "number",
            "input_type": "number",
            "default_value": "10",
            "category": "extraction",
            "description": "Number of pages per extraction chunk",
            "display_order": 14
        },
        {
            "key": "CHUNK_OVERLAP",
            "value": "1",
            "value_type": "number",
            "input_type": "number",
            "default_value": "1",
            "category": "extraction",
            "description": "Number of overlapping pages between chunks",
            "display_order": 14
        },
        {
            "key": "PDF_PROCESSING_TIMEOUT",
            "value": "300",
            "value_type": "number",
            "input_type": "number",
            "default_value": "300",
            "category": "extraction",
            "description": "Timeout in seconds for PDF processing operations",
            "display_order": 15
        },
        {
            "key": "MAX_FILE_SIZE_MB",
            "value": "100",
            "value_type": "number",
            "input_type": "number",
            "default_value": "100",
            "category": "extraction",
            "description": "Maximum allowed PDF file size in megabytes",
            "display_order": 16
        },
        {
            "key": "EXTRACTION_CONCURRENCY",
            "value": "3",
            "value_type": "number",
            "input_type": "number",
            "default_value": "3",
            "category": "extraction",
            "description": "Number of concurrent extraction workers",
            "display_order": 17
        },
        {
            "key": "DEFAULT_EXTRACTION_PROMPT",
            "value": "Extract all structured data from this document into the specified schema format. Be thorough and accurate.",
            "value_type": "string",
            "input_type": "textarea",
            "default_value": "Extract all structured data from this document into the specified schema format. Be thorough and accurate.",
            "category": "extraction",
            "description": "Default system prompt for extraction tasks",
            "display_order": 18
        },
        # ============================================================================
        # Cache Configuration
        # ============================================================================
        {
            "key": "CACHE_ENABLED",
            "value": "true",
            "value_type": "boolean",
            "input_type": "switch",
            "default_value": "true",
            "category": "general",
            "description": "Enable caching for PDF processing and LLM responses",
            "display_order": 20
        },
        {
            "key": "CACHE_TTL_HOURS",
            "value": "168",
            "value_type": "number",
            "input_type": "number",
            "default_value": "168",
            "category": "general",
            "description": "Cache time-to-live in hours (168 = 7 days)",
            "display_order": 21
        },
        {
            "key": "CACHE_MAX_SIZE_MB",
            "value": "1000",
            "value_type": "number",
            "input_type": "number",
            "default_value": "1000",
            "category": "general",
            "description": "Maximum cache size in megabytes",
            "display_order": 22
        },
        # ============================================================================
        # Output & Export
        # ============================================================================
        {
            "key": "DEFAULT_OUTPUT_FORMAT",
            "value": "json",
            "value_type": "string",
            "input_type": "select",
            "allowed_values": json.dumps(["json", "csv", "excel"]),
            "default_value": "json",
            "category": "general",
            "description": "Default format for exported extraction results",
            "display_order": 30
        },
        {
            "key": "INCLUDE_METADATA",
            "value": "true",
            "value_type": "boolean",
            "input_type": "switch",
            "default_value": "true",
            "category": "general",
            "description": "Include extraction metadata in output files",
            "display_order": 31
        },
        {
            "key": "AUTO_CLEANUP_DAYS",
            "value": "30",
            "value_type": "number",
            "input_type": "number",
            "default_value": "30",
            "category": "general",
            "description": "Auto-delete completed runs after this many days (0 = disabled)",
            "display_order": 32
        },
        # ============================================================================
        # Advanced Settings
        # ============================================================================
        {
            "key": "LOG_LEVEL",
            "value": "INFO",
            "value_type": "string",
            "input_type": "select",
            "allowed_values": json.dumps(["DEBUG", "INFO", "WARNING", "ERROR"]),
            "default_value": "INFO",
            "category": "advanced",
            "description": "Application logging verbosity level",
            "display_order": 200
        },
        {
            "key": "ENABLE_TELEMETRY",
            "value": "false",
            "value_type": "boolean",
            "input_type": "switch",
            "default_value": "false",
            "category": "advanced",
            "description": "Enable anonymous usage telemetry",
            "display_order": 201
        },
        {
            "key": "API_RATE_LIMIT",
            "value": "100",
            "value_type": "number",
            "input_type": "number",
            "default_value": "100",
            "category": "advanced",
            "description": "Maximum API requests per minute per user",
            "display_order": 202
        },
        {
            "key": "WORKER_TIMEOUT",
            "value": "3600",
            "value_type": "number",
            "input_type": "number",
            "default_value": "3600",
            "category": "advanced",
            "description": "Maximum time in seconds for a single extraction job",
            "display_order": 203
        },
        {
            "key": "DB_BACKUP_ENABLED",
            "value": "true",
            "value_type": "boolean",
            "input_type": "switch",
            "default_value": "true",
            "category": "advanced",
            "description": "Enable automatic database backups",
            "display_order": 204
        },
        {
            "key": "DB_BACKUP_INTERVAL_HOURS",
            "value": "24",
            "value_type": "number",
            "input_type": "number",
            "default_value": "24",
            "category": "advanced",
            "description": "Database backup interval in hours",
            "display_order": 205
        }
    ]
    
    conn = get_db()
    cur = conn.cursor()
    now = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    
    for cfg in defaults:
        # Check if key already exists (NULL user_id = global)
        cur.execute("SELECT 1 FROM config WHERE key = ? AND user_id IS NULL", (cfg["key"],))
        if cur.fetchone():
            continue  # Skip if already exists
        
        # Insert global defaults (user_id = NULL)
        cur.execute("""
            INSERT INTO config (key, user_id, value, value_type, input_type, allowed_values, 
                default_value, category, description, sensitive, required, display_order, last_modified)
            VALUES (?, NULL, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            cfg["key"],
            cfg.get("value", ""),
            cfg.get("value_type", "string"),
            cfg.get("input_type", "text"),
            cfg.get("allowed_values"),
            cfg.get("default_value"),
            cfg.get("category", "general"),
            cfg.get("description", ""),
            cfg.get("sensitive", 0),
            cfg.get("required", 0),
            cfg.get("display_order", 0),
            now
        ))
    
    conn.commit()
    conn.close()
    print("[STARTUP] Default configuration seeded")

# Seed default config on startup
seed_default_config()

def cleanup_stale_running_tasks():
    """Mark any 'running' tasks as 'aborted' on server startup.
    
    This handles the case where the server was killed while tasks were running.
    """
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id, name FROM runs WHERE status = 'running'")
    stale_runs = cur.fetchall()
    
    if stale_runs:
        for row in stale_runs:
            run_id = row["id"]
            run_name = row["name"]
            cur.execute("UPDATE runs SET status = 'aborted' WHERE id = ?", (run_id,))
            print(f"[STARTUP] Marked stale run as aborted: {run_name} ({run_id})")
        conn.commit()
        print(f"[STARTUP] Cleaned up {len(stale_runs)} stale running task(s)")
    
    conn.close()

# Cleanup stale tasks on startup
cleanup_stale_running_tasks()

# ============================================================================
# Utility Functions
# ============================================================================

def log_message(message: str, level: str = "INFO", run_id: str = None):
    """Log a message to DB and buffer for SSE."""
    now = datetime.now(timezone.utc).isoformat()
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO logs (created_at, level, message, run_id) VALUES (?, ?, ?, ?)",
        (now, level, message, run_id)
    )
    conn.commit()
    log_id = cur.lastrowid
    conn.close()
    
    # Add to SSE buffer
    with log_buffer_lock:
        log_buffer.append({"id": log_id, "createdAt": now, "level": level, "message": message, "runId": run_id})
        if len(log_buffer) > 1000:
            log_buffer.pop(0)

def paginate(query_result, page: int, page_size: int):
    """Paginate a list of results."""
    total = len(query_result)
    start = (page - 1) * page_size
    end = start + page_size

    items = query_result[start:end]
    return {"items": items, "total": total, "page": page, "pageSize": page_size}


def _parse_runs_sort(sort_value):
    allowed = {
        "startDate": "start_date",
        "name": "name",
        "status": "status",
        "sourcesCount": "sources_count",
        "dataEntriesCount": "data_entries_count",
        "llmProvider": "llm_provider",
        "id": "id",
    }

    if sort_value is None:
        return "start_date", "DESC"

    raw = str(sort_value).strip()
    if not raw:
        return "start_date", "DESC"

    if ":" in raw:
        field, direction = raw.split(":", 1)
    else:
        field, direction = raw, "desc"

    field = field.strip()
    direction = direction.strip().lower()

    col = allowed.get(field)
    if not col:
        return "start_date", "DESC"

    if direction not in {"asc", "desc"}:
        direction = "desc"

    return col, direction.upper()

def row_to_dict(row):
    """Convert sqlite3.Row to dict."""
    return dict(row) if row else None

def camel_case(snake_str):
    """Convert snake_case to camelCase."""
    components = snake_str.split('_')
    return components[0] + ''.join(x.title() for x in components[1:])

def to_camel_dict(d):
    """Convert all keys in dict from snake_case to camelCase."""
    if isinstance(d, dict):
        return {camel_case(k): to_camel_dict(v) for k, v in d.items()}
    elif isinstance(d, list):
        return [to_camel_dict(i) for i in d]
    return d

def register_file(filepath: str, original_name: str, file_type: str, run_id: str = None, mime_type: str = None) -> str:
    """Register a file in the database and return its ID. Path is stored internally, never exposed."""
    file_id = str(uuid.uuid4())
    filename = os.path.basename(filepath)
    size_bytes = os.path.getsize(filepath) if os.path.exists(filepath) else 0
    now = datetime.now(timezone.utc).isoformat()
    
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO files (id, filename, original_name, mime_type, size_bytes, file_type, run_id, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (file_id, filename, original_name, mime_type, size_bytes, file_type, run_id, now))
    conn.commit()
    conn.close()
    
    return file_id

def get_file_internal_path(file_id: str) -> str:
    """Get the internal file path for a file ID. Returns None if not found."""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT filename, run_id, file_type FROM files WHERE id = ?", (file_id,))
    row = cur.fetchone()
    conn.close()
    
    if not row:
        return None
    
    filename = row["filename"]
    run_id = row["run_id"]
    file_type = row["file_type"]
    
    # Reconstruct path based on file type
    if file_type in ("pdf", "crawled_pdf"):
        return os.path.join(UPLOAD_FOLDER, run_id, "pdfs", filename)
    elif file_type == "schema":
        return os.path.join(UPLOAD_FOLDER, run_id, filename)
    elif file_type == "export":
        return os.path.join(EXPORTS_FOLDER, run_id, filename)
    elif file_type == "zip":
        return os.path.join(UPLOAD_FOLDER, run_id, filename)
    elif file_type in ("validation_prompt", "extraction_prompt"):
        return os.path.join(UPLOAD_FOLDER, run_id, filename)
    else:
        # Default: assume it's in the uploads folder for the run
        return os.path.join(UPLOAD_FOLDER, run_id, filename) if run_id else None

def spawn_extraction_process(run_id: str, pdfs_dir: str, excel_path: str, output_dir: str, 
                             instructions: str = "", llm_provider: str = "openai",
                             enable_row_counting: bool = False, user_id: str = None,
                             validation_prompt_path: str = None, validation_enabled: bool = False,
                             validation_max_retries: int = 3,
                             cache_flags: dict = None):
    """
    Spawn extract.py as subprocess for a run.
    Captures stdout/stderr to IPC directory for logging.
    Updates run status on completion.
    Supports validation via --validation-text argument.
    """
    def update_run_status(run_id: str, new_status: str, error_message: str = None, entries_count: int = None):
        """Helper to update run status."""
        conn = get_db()
        cur = conn.cursor()
        
        if entries_count is not None:
            cur.execute("UPDATE runs SET status = ?, data_entries_count = ? WHERE id = ?",
                       (new_status, entries_count, run_id))
        else:
            cur.execute("UPDATE runs SET status = ? WHERE id = ?", (new_status, run_id))
        
        conn.commit()
        conn.close()
    
    def run_extraction():
        run_ipc_dir = os.path.join(IPC_DIR, run_id)
        os.makedirs(run_ipc_dir, exist_ok=True)
        
        stdout_path = os.path.join(run_ipc_dir, "stdout.log")
        stderr_path = os.path.join(run_ipc_dir, "stderr.log")
        
        def _get_int_config_value(key: str, user_id_for_config: str, default_value: int) -> int:
            try:
                conn = get_db()
                cur = conn.cursor()

                if user_id_for_config:
                    cur.execute(
                        "SELECT value FROM config WHERE key = ? AND user_id = ?",
                        (key, user_id_for_config),
                    )
                    row = cur.fetchone()
                    if row and row["value"] is not None and str(row["value"]).strip() != "":
                        conn.close()
                        return int(float(row["value"]))

                cur.execute(
                    "SELECT value FROM config WHERE key = ? AND user_id IS NULL",
                    (key,),
                )
                row = cur.fetchone()
                conn.close()
                if row and row["value"] is not None and str(row["value"]).strip() != "":
                    return int(float(row["value"]))
            except Exception:
                try:
                    conn.close()
                except Exception:
                    pass
            return default_value

        def _get_bool_config_value(key: str, user_id_for_config: str, default_value: bool) -> bool:
            try:
                conn = get_db()
                cur = conn.cursor()

                if user_id_for_config:
                    cur.execute(
                        "SELECT value FROM config WHERE key = ? AND user_id = ?",
                        (key, user_id_for_config),
                    )
                    row = cur.fetchone()
                    if row and row["value"] is not None and str(row["value"]).strip() != "":
                        conn.close()
                        return str(row["value"]).lower() in ("true", "1", "yes")

                cur.execute(
                    "SELECT value FROM config WHERE key = ? AND user_id IS NULL",
                    (key,),
                )
                row = cur.fetchone()
                conn.close()
                if row and row["value"] is not None and str(row["value"]).strip() != "":
                    return str(row["value"]).lower() in ("true", "1", "yes")
            except Exception:
                try:
                    conn.close()
                except Exception:
                    pass
            return default_value

        max_retries = _get_int_config_value("MAX_RETRIES", user_id, 0)
        retry_for_cached_empty_list = _get_bool_config_value("RETRY_FOR_CACHED_EMPTY_LIST", user_id, True)
        validation_retries_config = _get_int_config_value("VALIDATION_RETRIES", user_id, 1)

        # Build command
        cmd = [
            sys.executable, EXTRACT_SCRIPT,
            "--pdfs", pdfs_dir,
            "--excel", excel_path,
            "--output-dir", output_dir,
            "--log-file-path", os.path.join(run_ipc_dir, "extraction.log")
        ]

        if max_retries > 0:
            cmd.extend(["--retries", str(max_retries)])
        
        # Add instructions if provided
        if instructions:
            instructions_file = os.path.join(run_ipc_dir, "instructions.txt")
            with open(instructions_file, "w", encoding="utf-8") as f:
                f.write(instructions)
            cmd.extend(["--instructions", instructions_file])
        
        # Add row counting flag if enabled
        if enable_row_counting:
            cmd.append("--enable-row-counting")
        
        # Add retry for cached empty list flag
        cmd.extend(["--retry-for-cached-empty-list", str(retry_for_cached_empty_list).lower()])
        
        # Add granular cache flags if provided
        if cache_flags:
            for flag_name in ["surya_read", "surya_write", "llm_read", "llm_write", 
                              "schema_read", "schema_write", "validation_read", "validation_write"]:
                if flag_name in cache_flags:
                    cli_flag = f"--cache-{flag_name.replace('_', '-')}"
                    cmd.extend([cli_flag, str(cache_flags[flag_name]).lower()])
        
        # Add validation if enabled and validation prompt path provided
        if validation_enabled and validation_prompt_path and os.path.exists(validation_prompt_path):
            cmd.extend(["--validation-text", validation_prompt_path])
            # Use run-specific validation_max_retries if set, otherwise use global config
            effective_validation_retries = validation_max_retries if validation_max_retries > 0 else validation_retries_config
            if effective_validation_retries > 0:
                # Remove existing --retries if present
                new_cmd = []
                skip_next = False
                for i, c in enumerate(cmd):
                    if skip_next:
                        skip_next = False
                        continue
                    if c == "--retries":
                        skip_next = True
                        continue
                    new_cmd.append(c)
                cmd = new_cmd
                cmd.extend(["--retries", str(effective_validation_retries)])
        
        # Set environment for LLM provider
        env = os.environ.copy()
        env["LLM_PROVIDER"] = llm_provider
        # Two-call LLM routing (row counting vs extraction)
        # Defaults: row counting -> gemini, extraction -> llm_provider
        env.setdefault("ROWCOUNT_PROVIDER", "gemini")
        env.setdefault("EXTRACT_PROVIDER", llm_provider)
        # Optional model overrides (leave unset unless explicitly configured)
        if "ROWCOUNT_MODEL" in os.environ:
            env["ROWCOUNT_MODEL"] = os.environ["ROWCOUNT_MODEL"]
        if "EXTRACT_MODEL" in os.environ:
            env["EXTRACT_MODEL"] = os.environ["EXTRACT_MODEL"]
        # Set user ID for cache sandboxing
        if user_id:
            env["CACHE_USER_ID"] = user_id
        # Ensure UTF-8 encoding for subprocess on Windows
        env["PYTHONIOENCODING"] = "utf-8"
        env["PYTHONUTF8"] = "1"
        
        log_message(f"Starting extraction process for run {run_id[:8]}", "INFO", run_id)
        
        # Update status to running
        update_run_status(run_id, "running")
        
        try:
            with open(stdout_path, "w", encoding="utf-8") as stdout_file, \
                 open(stderr_path, "w", encoding="utf-8") as stderr_file:
                
                process = subprocess.Popen(
                    cmd,
                    stdout=stdout_file,
                    stderr=stderr_file,
                    env=env,
                    cwd=os.path.dirname(EXTRACT_SCRIPT)
                )
                active_processes[run_id] = process
                
                # Wait for completion
                return_code = process.wait()
                
                # Remove from active processes
                if run_id in active_processes:
                    del active_processes[run_id]
                
                if return_code == 0:
                    # Count extracted entries from output
                    global_json = os.path.join(output_dir, "global_data.json")
                    entries_count = 0
                    if os.path.exists(global_json):
                        try:
                            with open(global_json, "r") as f:
                                data = json.load(f)
                                entries_count = len(data) if isinstance(data, list) else 0
                        except:
                            pass
                    
                    # Parse validation results if validation was enabled
                    validation_pass_rate = None
                    validation_accepted = None
                    validation_rejected = None
                    validation_report_path = os.path.join(output_dir, "validation", "validation_report.json")
                    if os.path.exists(validation_report_path):
                        try:
                            with open(validation_report_path, "r", encoding="utf-8") as f:
                                val_report = json.load(f)
                            validation_pass_rate = val_report.get("summary", {}).get("overall_pass_rate")
                            # Count accepted/rejected from row_results
                            row_results = val_report.get("row_results", [])
                            if row_results:
                                validation_accepted = sum(1 for r in row_results if r.get("row_accept_candidate", True))
                                validation_rejected = len(row_results) - validation_accepted
                        except:
                            pass
                    
                    # Update run with validation results
                    conn = get_db()
                    cur = conn.cursor()
                    cur.execute("""
                        UPDATE runs SET status = ?, data_entries_count = ?,
                        validation_pass_rate = ?, validation_accepted_count = ?, validation_rejected_count = ?
                        WHERE id = ?
                    """, ("completed", entries_count, validation_pass_rate, validation_accepted, validation_rejected, run_id))
                    conn.commit()
                    conn.close()
                    
                    log_message(f"Extraction completed: {entries_count} entries", "INFO", run_id)
                else:
                    # Read stderr for error message
                    error_msg = ""
                    try:
                        with open(stderr_path, "r") as f:
                            error_msg = f.read()[-1000:]  # Last 1000 chars
                    except:
                        pass
                    
                    update_run_status(run_id, "failed", error_message=error_msg)
                    log_message(f"Extraction failed with code {return_code}", "ERROR", run_id)
                
        except Exception as e:
            log_message(f"Extraction error: {str(e)}", "ERROR", run_id)
            update_run_status(run_id, "failed", error_message=str(e))
            
            if run_id in active_processes:
                del active_processes[run_id]
    
    # Start in background thread
    thread = threading.Thread(target=run_extraction, daemon=True)
    thread.start()
    return thread


def spawn_validation_only_process(run_id: str, pdfs_dir: str, excel_path: str, output_dir: str,
                                   validation_prompt_path: str, user_id: str = None,
                                   cache_flags: dict = None):
    """
    Spawn extract.py with --validation-only flag.
    Skips extraction, only runs validation on existing global_data.json.
    Logs are appended to existing IPC logs.
    """
    def update_run_status(run_id: str, new_status: str, error_message: str = None):
        """Helper to update run status."""
        conn = get_db()
        cur = conn.cursor()
        cur.execute("UPDATE runs SET status = ? WHERE id = ?", (new_status, run_id))
        conn.commit()
        conn.close()
    
    def run_validation():
        run_ipc_dir = os.path.join(IPC_DIR, run_id)
        os.makedirs(run_ipc_dir, exist_ok=True)
        
        # Append to existing logs (use 'a' mode)
        stdout_path = os.path.join(run_ipc_dir, "stdout.log")
        stderr_path = os.path.join(run_ipc_dir, "stderr.log")
        
        # Build command with --validation-only flag
        cmd = [
            sys.executable, EXTRACT_SCRIPT,
            "--pdfs", pdfs_dir,
            "--excel", excel_path,
            "--output-dir", output_dir,
            "--log-file-path", os.path.join(run_ipc_dir, "extraction.log"),
            "--validation-only",
            "--validation-text", validation_prompt_path
        ]
        
        # Add granular cache flags if provided
        if cache_flags:
            for flag_name in ["surya_read", "surya_write", "llm_read", "llm_write", 
                              "schema_read", "schema_write", "validation_read", "validation_write"]:
                if flag_name in cache_flags:
                    cli_flag = f"--cache-{flag_name.replace('_', '-')}"
                    cmd.extend([cli_flag, str(cache_flags[flag_name]).lower()])
        
        # Set environment
        env = os.environ.copy()
        if user_id:
            env["CACHE_USER_ID"] = user_id
        env["PYTHONIOENCODING"] = "utf-8"
        env["PYTHONUTF8"] = "1"
        
        log_message(f"Starting validation-only process for run {run_id[:8]}", "INFO", run_id)
        
        # Update status to running (validation phase)
        update_run_status(run_id, "validating")
        
        try:
            # Append to existing logs
            with open(stdout_path, "a", encoding="utf-8") as stdout_file, \
                 open(stderr_path, "a", encoding="utf-8") as stderr_file:
                
                # Add separator in logs
                stdout_file.write(f"\n\n{'='*80}\n")
                stdout_file.write(f"[VALIDATION-ONLY MODE] Started at {datetime.now(timezone.utc).isoformat()}\n")
                stdout_file.write(f"{'='*80}\n\n")
                stdout_file.flush()
                
                process = subprocess.Popen(
                    cmd,
                    stdout=stdout_file,
                    stderr=stderr_file,
                    env=env,
                    cwd=os.path.dirname(EXTRACT_SCRIPT)
                )
                active_processes[run_id] = process
                
                # Wait for completion
                return_code = process.wait()
                
                # Remove from active processes
                if run_id in active_processes:
                    del active_processes[run_id]
                
                if return_code == 0:
                    # Parse validation results
                    validation_pass_rate = None
                    validation_accepted = None
                    validation_rejected = None
                    validation_report_path = os.path.join(output_dir, "validation", "validation_report.json")
                    if os.path.exists(validation_report_path):
                        try:
                            with open(validation_report_path, "r", encoding="utf-8") as f:
                                val_report = json.load(f)
                            validation_pass_rate = val_report.get("summary", {}).get("overall_pass_rate")
                            row_results = val_report.get("row_results", [])
                            if row_results:
                                validation_accepted = sum(1 for r in row_results if r.get("row_accept_candidate", True))
                                validation_rejected = len(row_results) - validation_accepted
                        except:
                            pass
                    
                    # Update run with validation results
                    conn = get_db()
                    cur = conn.cursor()
                    cur.execute("""
                        UPDATE runs SET status = ?,
                        validation_pass_rate = ?, validation_accepted_count = ?, validation_rejected_count = ?
                        WHERE id = ?
                    """, ("completed", validation_pass_rate, validation_accepted, validation_rejected, run_id))
                    conn.commit()
                    conn.close()
                    
                    log_message(f"Validation completed: pass_rate={validation_pass_rate}, accepted={validation_accepted}, rejected={validation_rejected}", "INFO", run_id)
                else:
                    # Read stderr for error message
                    error_msg = ""
                    try:
                        with open(stderr_path, "r") as f:
                            error_msg = f.read()[-1000:]
                    except:
                        pass
                    
                    update_run_status(run_id, "failed", error_message=error_msg)
                    log_message(f"Validation failed with code {return_code}", "ERROR", run_id)
                
        except Exception as e:
            log_message(f"Validation error: {str(e)}", "ERROR", run_id)
            update_run_status(run_id, "failed", error_message=str(e))
            
            if run_id in active_processes:
                del active_processes[run_id]
    
    # Start in background thread
    thread = threading.Thread(target=run_validation, daemon=True)
    thread.start()
    return thread


# Path to HTML extraction script
EXTRACT_HTML_SCRIPT = os.path.join(os.path.dirname(__file__), "extract_html.py")


def spawn_html_extraction_process(run_id: str, excel_path: str, output_dir: str, 
                                  instructions: str = "", llm_provider: str = "openai",
                                  user_id: str = None):
    """
    Spawn extract_html.py as subprocess for Deep Research/Links runs.
    Extracts data from HTML content in sources store (physical table name: sources).
    """
    def update_run_status(run_id: str, new_status: str, error_message: str = None, entries_count: int = None):
        """Helper to update run status."""
        conn = get_db()
        cur = conn.cursor()
        
        if entries_count is not None:
            cur.execute("UPDATE runs SET status = ?, data_entries_count = ? WHERE id = ?",
                       (new_status, entries_count, run_id))
        else:
            cur.execute("UPDATE runs SET status = ? WHERE id = ?", (new_status, run_id))
        
        conn.commit()
        conn.close()
    
    def run_html_extraction():
        run_ipc_dir = os.path.join(IPC_DIR, run_id)
        os.makedirs(run_ipc_dir, exist_ok=True)
        
        stdout_path = os.path.join(run_ipc_dir, "stdout.log")
        stderr_path = os.path.join(run_ipc_dir, "stderr.log")
        
        # Build command
        cmd = [
            sys.executable, EXTRACT_HTML_SCRIPT,
            "--run-id", run_id,
            "--excel", excel_path,
            "--output-dir", output_dir,
            "--db-path", DB_PATH,
            "--log-file-path", os.path.join(run_ipc_dir, "extraction.log")
        ]
        
        # Add instructions if provided
        if instructions:
            instructions_file = os.path.join(run_ipc_dir, "instructions.txt")
            with open(instructions_file, "w", encoding="utf-8") as f:
                f.write(instructions)
            cmd.extend(["--instructions", instructions_file])
        
        # Set environment for LLM provider
        env = os.environ.copy()
        env["LLM_PROVIDER"] = llm_provider
        # Set user ID for cache sandboxing
        if user_id:
            env["CACHE_USER_ID"] = user_id
        env["PYTHONIOENCODING"] = "utf-8"
        env["PYTHONUTF8"] = "1"
        
        log_message(f"Starting HTML extraction process for run {run_id[:8]}", "INFO", run_id)
        
        # Update status to running
        update_run_status(run_id, "running")
        
        try:
            with open(stdout_path, "w", encoding="utf-8") as stdout_file, \
                 open(stderr_path, "w", encoding="utf-8") as stderr_file:
                
                process = subprocess.Popen(
                    cmd,
                    stdout=stdout_file,
                    stderr=stderr_file,
                    env=env,
                    cwd=os.path.dirname(EXTRACT_HTML_SCRIPT)
                )
                active_processes[run_id] = process
                
                # Wait for completion
                return_code = process.wait()
                
                # Remove from active processes
                if run_id in active_processes:
                    del active_processes[run_id]
                
                if return_code == 0:
                    # Count extracted entries from output
                    global_json = os.path.join(output_dir, "global_data.json")
                    entries_count = 0
                    if os.path.exists(global_json):
                        try:
                            with open(global_json, "r") as f:
                                data = json.load(f)
                                entries_count = len(data) if isinstance(data, list) else 0
                        except:
                            pass
                    
                    # Update status to completed
                    update_run_status(run_id, "completed", entries_count=entries_count)
                    log_message(f"HTML extraction completed: {entries_count} entries", "INFO", run_id)
                else:
                    # Read stderr for error message
                    error_msg = ""
                    try:
                        with open(stderr_path, "r") as f:
                            error_msg = f.read()[-1000:]
                    except:
                        pass
                    
                    update_run_status(run_id, "failed", error_message=error_msg)
                    log_message(f"HTML extraction failed with code {return_code}", "ERROR", run_id)
                
        except Exception as e:
            log_message(f"HTML extraction error: {str(e)}", "ERROR", run_id)
            update_run_status(run_id, "failed", error_message=str(e))
            
            if run_id in active_processes:
                del active_processes[run_id]
    
    # Start in background thread
    thread = threading.Thread(target=run_html_extraction, daemon=True)
    thread.start()
    return thread


# ============================================================================
# API Routes - Runs (14 endpoints - added /start)
# ============================================================================

@app.route("/runs", methods=["GET"])
@optional_auth
def list_runs():
    """List runs with pagination.
    
    If authenticated, returns only the user's runs.
    If not authenticated, returns all runs (for backward compatibility).
    Use ?all=true to get all runs regardless of ownership (if authenticated).
    """
    page = int(request.args.get("page", 1))
    page_size = int(request.args.get("pageSize", 10))
    q = request.args.get("q", "")
    sort = request.args.get("sort")
    show_all = request.args.get("all", "false").lower() == "true"
    
    conn = get_db()
    cur = conn.cursor()
    
    # Filter by user if authenticated (unless ?all=true)
    user = g.current_user
    order_col, order_dir = _parse_runs_sort(sort)
    offset = max(0, (page - 1) * page_size)

    where_clause = ""
    params = []
    if user and not show_all:
        where_clause = "WHERE user_id = ?"
        params.append(user["id"])
        if q:
            where_clause += " AND name LIKE ?"
            params.append(f"%{q}%")
    else:
        if q:
            where_clause = "WHERE name LIKE ?"
            params.append(f"%{q}%")

    cur.execute(f"SELECT COUNT(*) as total FROM runs {where_clause}", tuple(params))
    total = int(cur.fetchone()[0])

    cur.execute(
        f"SELECT * FROM runs {where_clause} ORDER BY {order_col} {order_dir} LIMIT ? OFFSET ?",
        tuple(params + [page_size, offset]),
    )
    rows = cur.fetchall()
    conn.close()
    
    results = []
    for row in rows:
        run = dict(row)
        result = to_camel_dict(run)
        result["searchMethods"] = json.loads(result.get("searchMethods") or "[]")
        result["searchQueries"] = json.loads(result.get("searchQueries") or "[]")
        
        # Get live data_entries_count from global_data.json if available
        output_dir = run.get("output_dir")
        if output_dir:
            global_json = os.path.join(output_dir, "global_data.json")
            if os.path.exists(global_json):
                try:
                    with open(global_json, "r", encoding="utf-8") as f:
                        data = json.load(f)
                        result["dataEntriesCount"] = len(data) if isinstance(data, list) else 0
                except:
                    pass
        
        results.append(result)

    return jsonify({"items": results, "total": total, "page": page, "pageSize": page_size})

@app.route("/runs/nuke", methods=["POST"])
def nuke_all_runs():
    """DELETE ALL RUNS, EXPORTS, UPLOADS, AND LOGS. NUCLEAR OPTION.
    
    This endpoint:
    1. Deletes all runs from database
    2. Deletes all exports from database and filesystem
    3. Deletes all upload directories
    4. Deletes all IPC directories
    5. Clears all logs
    6. Resets to clean slate
    
    USE WITH EXTREME CAUTION.
    """
    import shutil
    
    conn = get_db()
    cur = conn.cursor()
    
    # Get counts before deletion
    cur.execute("SELECT COUNT(*) FROM runs")
    runs_count = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM exports")
    exports_count = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM logs")
    logs_count = cur.fetchone()[0]
    
    # Delete all database records
    cur.execute("DELETE FROM runs")
    cur.execute("DELETE FROM exports")
    cur.execute("DELETE FROM sources")
    cur.execute("DELETE FROM logs")
    cur.execute("DELETE FROM domains")
    cur.execute("DELETE FROM files")
    conn.commit()
    conn.close()
    
    # Delete all upload directories
    uploads_deleted = 0
    if os.path.exists(UPLOAD_FOLDER):
        for item in os.listdir(UPLOAD_FOLDER):
            item_path = os.path.join(UPLOAD_FOLDER, item)
            if os.path.isdir(item_path):
                try:
                    shutil.rmtree(item_path)
                    uploads_deleted += 1
                except:
                    pass
    
    # Delete all export directories
    exports_deleted = 0
    if os.path.exists(EXPORTS_FOLDER):
        for item in os.listdir(EXPORTS_FOLDER):
            item_path = os.path.join(EXPORTS_FOLDER, item)
            try:
                if os.path.isdir(item_path):
                    shutil.rmtree(item_path)
                else:
                    os.remove(item_path)
                exports_deleted += 1
            except:
                pass
    
    # Delete all IPC directories
    ipc_deleted = 0
    if os.path.exists(IPC_DIR):
        for item in os.listdir(IPC_DIR):
            item_path = os.path.join(IPC_DIR, item)
            if os.path.isdir(item_path):
                try:
                    shutil.rmtree(item_path)
                    ipc_deleted += 1
                except:
                    pass
    
    log_message("NUCLEAR OPTION EXECUTED - All data deleted", "WARN")
    
    return jsonify({
        "message": "NUKED - All data deleted",
        "deleted": {
            "runs": runs_count,
            "exports": exports_count,
            "logs": logs_count,
            "upload_dirs": uploads_deleted,
            "export_files": exports_deleted,
            "ipc_dirs": ipc_deleted
        }
    })


@app.route("/runs/<run_id>", methods=["GET"])
def get_run(run_id):
    """Get run details by ID. Paths are NOT exposed - use file IDs instead."""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM runs WHERE id = ?", (run_id,))
    row = cur.fetchone()
    conn.close()
    
    if not row:
        return jsonify({"error": "Run not found"}), 404
    
    run = dict(row)
    
    # For running jobs, get live count from global_data.json
    data_entries_count = run["data_entries_count"]
    if run["status"] == "running" and run.get("output_dir"):
        global_json = os.path.join(run["output_dir"], "global_data.json")
        if os.path.exists(global_json):
            try:
                with open(global_json, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    data_entries_count = len(data) if isinstance(data, list) else 0
            except:
                pass
    
    # Build response without exposing internal paths
    result = {
        "id": run["id"],
        "name": run["name"],
        "sourceType": run.get("source_type", "pdf"),
        "status": run["status"],
        "startDate": run["start_date"],
        "sourcesCount": run.get("sources_count", 0),
        "dataEntriesCount": data_entries_count,
        "llmProvider": run["llm_provider"],
        "prompt": run.get("prompt", ""),
        "searchMethods": json.loads(run.get("search_methods") or "[]"),
        "searchQueries": json.loads(run.get("search_queries") or "[]"),
        "links": json.loads(run.get("links") or "[]"),
        "schemaFileId": run.get("schema_file_id"),
        "zipFileId": run.get("zip_file_id"),
        "enableRowCounting": bool(run.get("enable_row_counting", 0)),
        "deepResearchQuery": run.get("deep_research_query"),
        "deepResearchResult": run.get("deep_research_result"),
    }
    
    return jsonify(result)


@app.route("/runs", methods=["POST"])
@optional_auth
def create_run():
    """Create a new run with file uploads.
    
    Accepts multipart/form-data with:
    - pdfsZip: ZIP file containing PDF files (required)
    - excelSchema: Excel schema file (required)
    - name: run name (optional, default: "Untitled Run")
    - llmProvider: LLM provider (optional, default: from config)
    - prompt: extraction instructions (optional)
    
    If authenticated, the run will be associated with the current user.
    """
    import zipfile
    import traceback
    
    try:
        # Validate required files
        if "pdfsZip" not in request.files:
            return jsonify({"error": "pdfsZip file is required"}), 400
        
        if "excelSchema" not in request.files:
            return jsonify({"error": "excelSchema file is required"}), 400
        
        pdfs_zip = request.files["pdfsZip"]
        excel_file = request.files["excelSchema"]
        
        if pdfs_zip.filename == "":
            return jsonify({"error": "No ZIP file selected"}), 400
        
        if excel_file.filename == "":
            return jsonify({"error": "No Excel file selected"}), 400
        
        # Validate file extensions
        if not pdfs_zip.filename.lower().endswith('.zip'):
            return jsonify({"error": "pdfsZip must be a .zip file"}), 400
        
        if not excel_file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({"error": "excelSchema must be an Excel file (.xlsx or .xls)"}), 400
        
        run_id = str(uuid.uuid4())
        
        # Create run directories
        run_upload_dir = os.path.join(UPLOAD_FOLDER, run_id)
        pdfs_dir = os.path.join(run_upload_dir, "pdfs")
        output_dir = os.path.join(EXPORTS_FOLDER, run_id)
        
        os.makedirs(pdfs_dir, exist_ok=True)
        os.makedirs(output_dir, exist_ok=True)
        
        # Save and extract ZIP file
        zip_filename = secure_filename(pdfs_zip.filename)
        zip_path = os.path.join(run_upload_dir, zip_filename)
        pdfs_zip.save(zip_path)
        
        with zipfile.ZipFile(zip_path, 'r') as zf:
            # Extract only PDF files
            pdf_count = 0
            for name in zf.namelist():
                if name.lower().endswith('.pdf') and not name.startswith('__MACOSX'):
                    # Extract to pdfs_dir with flat structure
                    zf.extract(name, pdfs_dir)
                    # Move to root of pdfs_dir if nested
                    extracted_path = os.path.join(pdfs_dir, name)
                    if os.path.dirname(name):
                        flat_path = os.path.join(pdfs_dir, os.path.basename(name))
                        if extracted_path != flat_path:
                            os.rename(extracted_path, flat_path)
                    pdf_count += 1
        
        # Clean up empty subdirectories from extraction
        for root, dirs, files in os.walk(pdfs_dir, topdown=False):
            for d in dirs:
                try:
                    os.rmdir(os.path.join(root, d))
                except:
                    pass
        
        if pdf_count == 0:
            return jsonify({"error": "No PDF files found in ZIP"}), 400
        
        # Save Excel schema file
        excel_filename = secure_filename(excel_file.filename)
        excel_path = os.path.join(run_upload_dir, excel_filename)
        excel_file.save(excel_path)
        
        # Get form data
        name = request.form.get("name", "Untitled Run")
        llm_provider = request.form.get("llmProvider", LLM_PROVIDER)
        enable_row_counting = request.form.get("enableRowCounting", "false").lower() == "true"
        validation_enabled = request.form.get("validationEnabled", "false").lower() == "true"
        validation_max_retries = int(request.form.get("validationMaxRetries", "3"))
        now = datetime.now(timezone.utc).isoformat()
        
        # Parse cache flags from form data (JSON string or individual fields)
        cache_flags = None
        cache_flags_json = request.form.get("cacheFlags")
        if cache_flags_json:
            try:
                cache_flags = json.loads(cache_flags_json)
            except:
                pass
        if not cache_flags:
            # Try individual fields
            cache_flags = {}
            for flag in ["surya_read", "surya_write", "llm_read", "llm_write", 
                         "schema_read", "schema_write", "validation_read", "validation_write"]:
                form_key = f"cache_{flag.replace('_', '')}"  # e.g., cache_suryaread
                val = request.form.get(form_key)
                if val is not None:
                    cache_flags[flag] = val.lower() == "true"
            if not cache_flags:
                cache_flags = None
        
        # Register files in database (paths stored internally, IDs exposed to frontend)
        zip_file_id = register_file(zip_path, pdfs_zip.filename, "zip", run_id, "application/zip")
        schema_file_id = register_file(excel_path, excel_file.filename, "schema", run_id, 
                                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
        # Handle extraction prompt file upload (replaces textarea)
        extraction_prompt_file_id = None
        extraction_prompt_path = None
        if "extractionPrompt" in request.files:
            prompt_file = request.files["extractionPrompt"]
            if prompt_file.filename and prompt_file.filename.lower().endswith('.txt'):
                prompt_filename = secure_filename(prompt_file.filename)
                extraction_prompt_path = os.path.join(run_upload_dir, f"extraction_prompt_{prompt_filename}")
                prompt_file.save(extraction_prompt_path)
                extraction_prompt_file_id = register_file(extraction_prompt_path, prompt_file.filename, "extraction_prompt", run_id, "text/plain")
        
        # Handle validation prompt file upload
        validation_prompt_file_id = None
        validation_prompt_path = None
        if "validationPrompt" in request.files:
            val_prompt_file = request.files["validationPrompt"]
            if val_prompt_file.filename and val_prompt_file.filename.lower().endswith('.txt'):
                val_prompt_filename = secure_filename(val_prompt_file.filename)
                validation_prompt_path = os.path.join(run_upload_dir, f"validation_prompt_{val_prompt_filename}")
                val_prompt_file.save(validation_prompt_path)
                validation_prompt_file_id = register_file(validation_prompt_path, val_prompt_file.filename, "validation_prompt", run_id, "text/plain")
        
        # Read extraction prompt content for backward compatibility (prompt column)
        prompt = ""
        if extraction_prompt_path and os.path.exists(extraction_prompt_path):
            with open(extraction_prompt_path, "r", encoding="utf-8") as f:
                prompt = f.read()
        
        # Register individual PDFs
        for pdf_file in os.listdir(pdfs_dir):
            if pdf_file.lower().endswith('.pdf'):
                pdf_path = os.path.join(pdfs_dir, pdf_file)
                register_file(pdf_path, pdf_file, "pdf", run_id, "application/pdf")
        
        # Get current user if authenticated
        user_id = g.current_user["id"] if g.current_user else None
        
        # Create run using SQLite
        conn = get_db()
        cur = conn.cursor()

        # Create meta source (provenance) for this run
        meta_source_id = create_meta_source(run_id=run_id, method="pdf_upload", user_id=user_id, name=name)
        cur.execute("""
            INSERT INTO runs (id, name, status, start_date, llm_provider, pdfs_dir, excel_path, output_dir, 
                              prompt, search_methods, search_queries, links, table_file_url, per_link_prompt, 
                              sources_count, schema_file_id, zip_file_id, enable_row_counting, user_id, meta_source_id,
                              extraction_prompt_file_id, validation_prompt_file_id, validation_enabled, validation_max_retries,
                              cache_flags)
            VALUES (?, ?, 'waiting', ?, ?, ?, ?, ?, ?, '[]', '[]', '[]', '', '', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            run_id,
            name,
            now,
            llm_provider,
            pdfs_dir,
            excel_path,
            output_dir,
            prompt,
            pdf_count,
            schema_file_id,
            zip_file_id,
            1 if enable_row_counting else 0,
            user_id,
            meta_source_id,
            extraction_prompt_file_id,
            validation_prompt_file_id,
            1 if validation_enabled else 0,
            validation_max_retries,
            json.dumps(cache_flags) if cache_flags else None
        ))

        # Insert sources rows immediately for each PDF file (PENDING)
        cur.execute(
            "SELECT id, original_name, filename FROM files WHERE run_id = ? AND file_type IN ('pdf','crawled_pdf')",
            (run_id,),
        )
        for f in cur.fetchall():
            pdf_file_id = f["id"]
            title = f["original_name"] or f["filename"]
            # Use pdf_file_id as source_id for PDF-upload sources (stable identity)
            ensure_source_row(
                source_id=pdf_file_id,
                run_id=run_id,
                source_type="pdf",
                status="PENDING",
                url=None,
                title=title,
                crawl_job_id=None,
                pdf_file_id=pdf_file_id,
                meta_source_id=meta_source_id,
                cur=cur,
            )

        conn.commit()
        
        cur.execute("SELECT * FROM runs WHERE id = ?", (run_id,))
        row = cur.fetchone()
        conn.close()
        
        log_message(f"Run created: {name} ({pdf_count} PDFs uploaded)", "INFO", run_id)

        # Start background conversion for PDF sources so they transition to READY
        try:
            threading.Thread(target=process_pdf_sources_for_run, args=(run_id, user_id), daemon=True).start()
        except Exception:
            pass
        
        # Build response without exposing paths
        result = {
            "id": run_id,
            "name": name,
            "sourceType": "pdf",
            "status": "waiting",
            "startDate": now,
            "sourcesCount": pdf_count,
            "dataEntriesCount": 0,
            "llmProvider": llm_provider,
            "prompt": prompt,
            "searchMethods": [],
            "searchQueries": [],
            "schemaFileId": schema_file_id,
            "zipFileId": zip_file_id,
            "enableRowCounting": bool(enable_row_counting),
            "extractionPromptFileId": extraction_prompt_file_id,
            "validationPromptFileId": validation_prompt_file_id,
            "validationEnabled": validation_enabled,
            "validationMaxRetries": validation_max_retries
        }
        
        return jsonify(result), 201
        
    except zipfile.BadZipFile:
        return jsonify({"error": "Invalid ZIP file"}), 400
    except Exception as e:
        traceback.print_exc()
        log_message(f"Run creation failed: {str(e)}", "ERROR")
        return jsonify({"error": f"Failed to create run: {str(e)}"}), 500


@app.route("/runs/from-search", methods=["POST"])
@require_auth
def create_run_from_search():
    """Create a new run using Deep Research (Gemini search).
    
    This initiates a Gemini Deep Research query, extracts links from the results,
    and creates crawl jobs for the Chrome extension to fetch HTML content.
    
    Accepts multipart/form-data with:
    - excelSchema: Excel schema file (required)
    - query: search query for Gemini Deep Research (required)
    - name: run name (optional, default: "Deep Research Run")
    - llmProvider: LLM provider for extraction (optional)
    - prompt: extraction instructions (optional)
    """
    user_id = g.current_user["id"]
    
    # Validate required Excel schema file
    if "excelSchema" not in request.files:
        return jsonify({"error": "excelSchema file is required"}), 400
    
    excel_file = request.files["excelSchema"]
    if excel_file.filename == "":
        return jsonify({"error": "No Excel file selected"}), 400
    
    if not excel_file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({"error": "excelSchema must be an Excel file (.xlsx or .xls)"}), 400
    
    # Get form fields
    query = request.form.get("query", "").strip()
    if not query:
        return jsonify({"error": "Search query is required"}), 400
    
    name = request.form.get("name", "Deep Research Run").strip()
    llm_provider = request.form.get("llmProvider", LLM_PROVIDER)
    validation_enabled = request.form.get("validationEnabled", "false").lower() == "true"
    validation_max_retries = int(request.form.get("validationMaxRetries", "3"))
    
    # Get Gemini API key
    api_key = get_gemini_api_key(user_id)
    if not api_key:
        return jsonify({"error": "GEMINI_API_KEY not configured. Set it in Config > API Keys."}), 400
    
    run_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    
    # Create run directories
    run_upload_dir = os.path.join(UPLOAD_FOLDER, run_id)
    output_dir = os.path.join(EXPORTS_FOLDER, run_id)
    os.makedirs(run_upload_dir, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)
    
    # Save Excel schema file
    excel_filename = secure_filename(excel_file.filename)
    excel_path = os.path.join(run_upload_dir, excel_filename)
    excel_file.save(excel_path)
    
    # Register schema file in database
    schema_file_id = register_file(excel_path, excel_file.filename, "schema", run_id, 
                                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    
    # Handle extraction prompt file upload
    extraction_prompt_file_id = None
    extraction_prompt_path = None
    prompt = ""
    if "extractionPrompt" in request.files:
        prompt_file = request.files["extractionPrompt"]
        if prompt_file.filename and prompt_file.filename.lower().endswith('.txt'):
            prompt_filename = secure_filename(prompt_file.filename)
            extraction_prompt_path = os.path.join(run_upload_dir, f"extraction_prompt_{prompt_filename}")
            prompt_file.save(extraction_prompt_path)
            extraction_prompt_file_id = register_file(extraction_prompt_path, prompt_file.filename, "extraction_prompt", run_id, "text/plain")
            with open(extraction_prompt_path, "r", encoding="utf-8") as f:
                prompt = f.read()
    
    # Handle validation prompt file upload
    validation_prompt_file_id = None
    if "validationPrompt" in request.files:
        val_prompt_file = request.files["validationPrompt"]
        if val_prompt_file.filename and val_prompt_file.filename.lower().endswith('.txt'):
            val_prompt_filename = secure_filename(val_prompt_file.filename)
            validation_prompt_path = os.path.join(run_upload_dir, f"validation_prompt_{val_prompt_filename}")
            val_prompt_file.save(validation_prompt_path)
            validation_prompt_file_id = register_file(validation_prompt_path, val_prompt_file.filename, "validation_prompt", run_id, "text/plain")
    
    conn = get_db()
    cur = conn.cursor()
    
    # Create run with source_type='deep_research' and status='searching'
    meta_source_id = create_meta_source(run_id=run_id, method="google_deep_research", user_id=user_id, name=name, query=query)
    cur.execute("""
        INSERT INTO runs (id, name, source_type, status, start_date, llm_provider, excel_path, output_dir, 
                          prompt, deep_research_query, schema_file_id, sources_count, user_id, meta_source_id,
                          extraction_prompt_file_id, validation_prompt_file_id, validation_enabled, validation_max_retries)
        VALUES (?, ?, 'deep_research', 'searching', ?, ?, ?, ?, ?, ?, ?, 0, ?, ?, ?, ?, ?, ?)
    """, (run_id, name, now, llm_provider, excel_path, output_dir, prompt, query, schema_file_id, user_id, meta_source_id,
          extraction_prompt_file_id, validation_prompt_file_id, 1 if validation_enabled else 0, validation_max_retries))
    conn.commit()
    conn.close()
    
    log_message(f"Deep Research run created: {name}", "INFO", run_id)
    
    # Start Deep Research in background thread
    def run_deep_research():
        try:
            import requests
            import re
            
            # Update status to 'researching'
            conn2 = get_db()
            cur2 = conn2.cursor()
            cur2.execute("UPDATE runs SET status = 'researching' WHERE id = ?", (run_id,))
            conn2.commit()
            conn2.close()
            
            log_message(f"Starting Gemini Deep Research for query: {query}", "INFO", run_id)
            
            from cache_utils import get_gpt_cache, set_gpt_cache
            
            # Check cache first
            cache_model = "gemini-deep-research:gemini-2.0-flash"
            cached_response = get_gpt_cache("", query, cache_model)
            
            if cached_response:
                log_message(f"[CACHE HIT] Deep Research - using cached response", "INFO", run_id)
                response = cached_response
            else:
                log_message(f"[CACHE MISS] Deep Research - calling Gemini API...", "INFO", run_id)
                # Use Gemini REST API with Google Search grounding
                gemini_url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={api_key}"
                
                payload = {
                    "contents": [{
                        "parts": [{
                            "text": query
                        }]
                    }],
                    "tools": [{
                        "google_search": {}
                    }]
                }
                
                log_message(f"Gemini API request payload: {json.dumps(payload)[:500]}", "DEBUG", run_id)
                
                resp = requests.post(
                    gemini_url,
                    headers={"Content-Type": "application/json"},
                    json=payload,
                    timeout=120
                )
                resp.raise_for_status()
                response = resp.json()
                
                # Cache the response
                set_gpt_cache("", query, cache_model, response)
                log_message(f"[CACHE STORED] Deep Research response cached", "INFO", run_id)
            
            result_text = ""
            extracted_links = []
            
            # Log response structure for debugging
            log_message(f"Response keys: {list(response.keys())}", "DEBUG", run_id)
            if "candidates" in response:
                log_message(f"Candidates count: {len(response['candidates'])}", "DEBUG", run_id)
            
            # Extract text and links from response
            if "candidates" in response:
                for idx, candidate in enumerate(response["candidates"]):
                    content = candidate.get("content", {})
                    parts = content.get("parts", [])
                    log_message(f"Candidate {idx}: {len(parts)} parts, keys: {list(candidate.keys())}", "DEBUG", run_id)
                    
                    for part in parts:
                        if "text" in part:
                            result_text += part["text"] + "\n"
                    
                    # Extract links from grounding metadata
                    grounding_metadata = candidate.get("groundingMetadata", {})
                    log_message(f"Grounding metadata keys: {list(grounding_metadata.keys())}", "DEBUG", run_id)
                    
                    grounding_chunks = grounding_metadata.get("groundingChunks", [])
                    log_message(f"Grounding chunks count: {len(grounding_chunks)}", "DEBUG", run_id)
                    
                    for chunk in grounding_chunks:
                        web = chunk.get("web", {})
                        url = web.get("uri", "") or web.get("url", "")
                        title = web.get("title", "")
                        if url:
                            extracted_links.append({"url": url, "title": title})
                    
                    # Also check groundingSupports for additional URLs
                    grounding_supports = grounding_metadata.get("groundingSupports", [])
                    log_message(f"Grounding supports count: {len(grounding_supports)}", "DEBUG", run_id)
                    
                    for support in grounding_supports:
                        for chunk_idx in support.get("groundingChunkIndices", []):
                            if chunk_idx < len(grounding_chunks):
                                web = grounding_chunks[chunk_idx].get("web", {})
                                url = web.get("uri", "") or web.get("url", "")
                                if url and not any(l["url"] == url for l in extracted_links):
                                    extracted_links.append({"url": url, "title": web.get("title", "")})
                    
                    # Check for search entry point if no grounding chunks
                    search_entry_point = grounding_metadata.get("searchEntryPoint", {})
                    if search_entry_point:
                        log_message(f"Search entry point: {json.dumps(search_entry_point)[:200]}", "DEBUG", run_id)
            
            log_message(f"Links from grounding metadata: {len(extracted_links)}", "DEBUG", run_id)
            log_message(f"Result text length: {len(result_text)} chars", "DEBUG", run_id)
            if result_text:
                log_message(f"Result text preview: {result_text[:500]}...", "DEBUG", run_id)
            
            # Fallback: extract URLs from text if no grounding metadata
            if not extracted_links and result_text:
                log_message(f"No grounding links found, extracting URLs from text...", "INFO", run_id)
                url_pattern = r'https?://[^\s<>"{}|\\^`\[\]]+'
                urls = re.findall(url_pattern, result_text)
                log_message(f"URLs found in text: {len(urls)}", "DEBUG", run_id)
                for url in urls[:50]:  # Limit to 50 links
                    extracted_links.append({"url": url, "title": ""})
            
            # Deduplicate links
            seen_urls = set()
            unique_links = []
            for link in extracted_links:
                url = link.get("url", "")
                if url and url not in seen_urls:
                    seen_urls.add(url)
                    unique_links.append(link)
            
            log_message(f"Deep Research completed: {len(unique_links)} links extracted", "SUCCESS", run_id)
            
            # Update run with results
            conn3 = get_db()
            cur3 = conn3.cursor()
            cur3.execute("""
                UPDATE runs 
                SET status = 'crawling', deep_research_result = ?, links = ?
                WHERE id = ?
            """, (result_text, json.dumps(unique_links), run_id))
            conn3.commit()
            
            # Create crawl jobs for all links (HTML as PENDING, PDFs as PDF_PENDING)
            html_count = 0
            pdf_count = 0
            for link in unique_links:
                try:
                    url = link.get("url", "")
                    title = link.get("title", "")
                    if not url:
                        continue
                    
                    # Check if PDF
                    url_lower = url.lower()
                    if url_lower.endswith('.pdf') or '/pdf/' in url_lower or 'pdf?' in url_lower:
                        pdf_count += 1
                        job_id = str(uuid.uuid4())
                        cur3.execute("""
                            INSERT INTO crawl_jobs (id, run_id, user_id, url, title, status, created_at)
                            VALUES (?, ?, ?, ?, ?, 'PDF_PENDING', ?)
                        """, (job_id, run_id, user_id, url, title, now))

                        # Insert source immediately (PENDING)
                        ensure_source_row(
                            source_id=job_id,
                            run_id=run_id,
                            source_type="pdf",
                            status="PENDING",
                            url=url,
                            title=title,
                            crawl_job_id=job_id,
                            pdf_file_id=None,
                            meta_source_id=meta_source_id,
                            cur=cur3,
                        )
                    else:
                        job_id = str(uuid.uuid4())
                        cur3.execute("""
                            INSERT INTO crawl_jobs (id, run_id, user_id, url, title, status, created_at)
                            VALUES (?, ?, ?, ?, ?, 'PENDING', ?)
                        """, (job_id, run_id, user_id, url, title, now))
                        html_count += 1

                        # Insert source immediately (PENDING)
                        ensure_source_row(
                            source_id=job_id,
                            run_id=run_id,
                            source_type="link",
                            status="PENDING",
                            url=url,
                            title=title,
                            crawl_job_id=job_id,
                            pdf_file_id=None,
                            meta_source_id=meta_source_id,
                            cur=cur3,
                        )
                except Exception as link_err:
                    log_message(f"Failed to create crawl job for {url[:100]}: {str(link_err)}", "ERROR", run_id)
            
            cur3.execute("UPDATE runs SET sources_count = ? WHERE id = ?", (html_count + pdf_count, run_id))
            conn3.commit()
            conn3.close()
            
            log_message(f"Created {html_count} HTML crawl jobs and {pdf_count} PDF jobs", "SUCCESS", run_id)
            
            # Start background thread to download PDFs if any
            if pdf_count > 0:
                threading.Thread(
                    target=process_pdf_jobs_for_run,
                    args=(run_id, user_id),
                    daemon=True
                ).start()
                log_message(f"Started background PDF download for {pdf_count} PDF(s)", "INFO", run_id)
            
        except Exception as e:
            log_message(f"Deep Research failed: {str(e)}", "ERROR", run_id)
            conn_err = get_db()
            cur_err = conn_err.cursor()
            cur_err.execute("UPDATE runs SET status = 'failed' WHERE id = ?", (run_id,))
            conn_err.commit()
            conn_err.close()
    
    # Start background thread
    import threading
    thread = threading.Thread(target=run_deep_research, daemon=True)
    thread.start()
    
    return jsonify({
        "id": run_id,
        "name": name,
        "sourceType": "deep_research",
        "status": "searching",
        "query": query,
        "startDate": now,
        "sourcesCount": 0,
        "llmProvider": llm_provider,
        "prompt": prompt
    }), 201


@app.route("/runs/<run_id>/ipc", methods=["GET"])
def get_run_ipc(run_id):
    """Get IPC metadata for a run (no paths exposed)."""
    ipc_dir = os.path.join(IPC_DIR, run_id)
    metadata_path = os.path.join(ipc_dir, "metadata.json")
    
    metadata = {}
    if os.path.exists(metadata_path):
        with open(metadata_path, "r") as f:
            metadata = json.load(f)
    
    return jsonify({"runId": run_id, "metadata": metadata})


@app.route("/runs/from-links", methods=["POST"])
@require_auth
def create_run_from_links():
    """Create a new run from manually provided URLs.
    
    Accepts multipart/form-data with:
    - excelSchema: Excel schema file (required)
    - links: JSON array of URLs or newline-separated URLs (required)
    - name: run name (optional, default: "Manual Links Run")
    - llmProvider: LLM provider for extraction (optional)
    - prompt: extraction instructions (optional)
    """
    user_id = g.current_user["id"]
    
    # Validate required Excel schema file
    if "excelSchema" not in request.files:
        return jsonify({"error": "excelSchema file is required"}), 400
    
    excel_file = request.files["excelSchema"]
    if excel_file.filename == "":
        return jsonify({"error": "No Excel file selected"}), 400
    
    if not excel_file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({"error": "excelSchema must be an Excel file (.xlsx or .xls)"}), 400
    
    # Get and parse links
    links_raw = request.form.get("links", "").strip()
    if not links_raw:
        return jsonify({"error": "Links are required"}), 400
    
    # Parse links - support JSON array or newline-separated
    try:
        links_list = json.loads(links_raw)
        if not isinstance(links_list, list):
            links_list = [links_list]
    except json.JSONDecodeError:
        # Treat as newline-separated URLs
        links_list = [url.strip() for url in links_raw.split('\n') if url.strip()]
    
    # Validate URLs
    valid_links = []
    for link in links_list:
        if isinstance(link, dict):
            url = link.get("url", "")
            title = link.get("title", "")
        else:
            url = str(link).strip()
            title = ""
        
        if url and (url.startswith("http://") or url.startswith("https://")):
            valid_links.append({"url": url, "title": title})
    
    if not valid_links:
        return jsonify({"error": "No valid URLs provided. URLs must start with http:// or https://"}), 400
    
    # Get form fields
    name = request.form.get("name", "Manual Links Run").strip()
    llm_provider = request.form.get("llmProvider", LLM_PROVIDER)
    validation_enabled = request.form.get("validationEnabled", "false").lower() == "true"
    validation_max_retries = int(request.form.get("validationMaxRetries", "3"))
    
    run_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    
    # Create run directories
    run_upload_dir = os.path.join(UPLOAD_FOLDER, run_id)
    output_dir = os.path.join(EXPORTS_FOLDER, run_id)
    os.makedirs(run_upload_dir, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)
    
    # Save Excel schema file
    excel_filename = secure_filename(excel_file.filename)
    excel_path = os.path.join(run_upload_dir, excel_filename)
    excel_file.save(excel_path)
    
    # Register schema file in database
    schema_file_id = register_file(excel_path, excel_file.filename, "schema", run_id, 
                                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    
    # Handle extraction prompt file upload
    extraction_prompt_file_id = None
    prompt = ""
    if "extractionPrompt" in request.files:
        prompt_file = request.files["extractionPrompt"]
        if prompt_file.filename and prompt_file.filename.lower().endswith('.txt'):
            prompt_filename = secure_filename(prompt_file.filename)
            extraction_prompt_path = os.path.join(run_upload_dir, f"extraction_prompt_{prompt_filename}")
            prompt_file.save(extraction_prompt_path)
            extraction_prompt_file_id = register_file(extraction_prompt_path, prompt_file.filename, "extraction_prompt", run_id, "text/plain")
            with open(extraction_prompt_path, "r", encoding="utf-8") as f:
                prompt = f.read()
    
    # Handle validation prompt file upload
    validation_prompt_file_id = None
    if "validationPrompt" in request.files:
        val_prompt_file = request.files["validationPrompt"]
        if val_prompt_file.filename and val_prompt_file.filename.lower().endswith('.txt'):
            val_prompt_filename = secure_filename(val_prompt_file.filename)
            validation_prompt_path = os.path.join(run_upload_dir, f"validation_prompt_{val_prompt_filename}")
            val_prompt_file.save(validation_prompt_path)
            validation_prompt_file_id = register_file(validation_prompt_path, val_prompt_file.filename, "validation_prompt", run_id, "text/plain")
    
    conn = get_db()
    cur = conn.cursor()
    
    # Create run with source_type='links' and status='crawling'
    meta_source_id = create_meta_source(run_id=run_id, method="manual_links", user_id=user_id, name=name)
    cur.execute("""
        INSERT INTO runs (id, name, source_type, status, start_date, llm_provider, excel_path, output_dir, 
                          prompt, links, schema_file_id, sources_count, user_id, meta_source_id,
                          extraction_prompt_file_id, validation_prompt_file_id, validation_enabled, validation_max_retries)
        VALUES (?, ?, 'links', 'crawling', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (run_id, name, now, llm_provider, excel_path, output_dir, prompt, json.dumps(valid_links), 
          schema_file_id, len(valid_links), user_id, meta_source_id,
          extraction_prompt_file_id, validation_prompt_file_id, 1 if validation_enabled else 0, validation_max_retries))
    conn.commit()
    
    # Create crawl jobs for each URL
    html_count = 0
    pdf_count = 0
    for link in valid_links:
        url = link.get("url", "")
        title = link.get("title", "")
        if not url:
            continue
        
        # Check if PDF
        url_lower = url.lower()
        if url_lower.endswith('.pdf') or '/pdf/' in url_lower or 'pdf?' in url_lower:
            pdf_count += 1
            # Still create job but mark for PDF processing
            job_id = str(uuid.uuid4())
            cur.execute("""
                INSERT INTO crawl_jobs (id, run_id, user_id, url, title, status, created_at)
                VALUES (?, ?, ?, ?, ?, 'PDF_PENDING', ?)
            """, (job_id, run_id, user_id, url, title, now))

            ensure_source_row(
                source_id=job_id,
                run_id=run_id,
                source_type="pdf",
                status="PENDING",
                url=url,
                title=title,
                crawl_job_id=job_id,
                pdf_file_id=None,
                meta_source_id=meta_source_id,
            )
        else:
            job_id = str(uuid.uuid4())
            cur.execute("""
                INSERT INTO crawl_jobs (id, run_id, user_id, url, title, status, created_at)
                VALUES (?, ?, ?, ?, ?, 'PENDING', ?)
            """, (job_id, run_id, user_id, url, title, now))
            html_count += 1

            ensure_source_row(
                source_id=job_id,
                run_id=run_id,
                source_type="link",
                status="PENDING",
                url=url,
                title=title,
                crawl_job_id=job_id,
                pdf_file_id=None,
                meta_source_id=meta_source_id,
            )
    
    conn.commit()
    conn.close()
    
    log_message(f"Manual Links run created: {name} with {len(valid_links)} URLs ({html_count} HTML, {pdf_count} PDFs)", "INFO", run_id)
    
    # Start background thread to download PDFs if any
    if pdf_count > 0:
        threading.Thread(
            target=process_pdf_jobs_for_run,
            args=(run_id, user_id),
            daemon=True
        ).start()
        log_message(f"Started background PDF download for {pdf_count} PDF(s)", "INFO", run_id)
    
    return jsonify({
        "id": run_id,
        "name": name,
        "sourceType": "links",
        "status": "crawling",
        "links": valid_links,
        "startDate": now,
        "sourcesCount": len(valid_links),
        "htmlCount": html_count,
        "pdfCount": pdf_count,
        "llmProvider": llm_provider,
        "prompt": prompt
    }), 201


# =============================================================================
# Files API - Access files by ID (paths never exposed)
# =============================================================================

@app.route("/files", methods=["GET"])
def list_files():
    """List all files, optionally filtered by run_id or file_type."""
    run_id = request.args.get("runId")
    file_type = request.args.get("type")
    
    conn = get_db()
    cur = conn.cursor()
    
    query = "SELECT id, filename, original_name, mime_type, size_bytes, file_type, run_id, created_at FROM files WHERE 1=1"
    params = []
    
    if run_id:
        query += " AND run_id = ?"
        params.append(run_id)
    if file_type:
        query += " AND file_type = ?"
        params.append(file_type)
    
    query += " ORDER BY created_at DESC"
    cur.execute(query, params)
    rows = cur.fetchall()
    conn.close()
    
    files = []
    for row in rows:
        files.append({
            "id": row["id"],
            "filename": row["filename"],
            "originalName": row["original_name"],
            "mimeType": row["mime_type"],
            "sizeBytes": row["size_bytes"],
            "fileType": row["file_type"],
            "runId": row["run_id"],
            "createdAt": row["created_at"]
        })
    
    return jsonify({"items": files, "total": len(files)})


@app.route("/files/<file_id>", methods=["GET"])
def get_file_info(file_id):
    """Get file metadata by ID (no path exposed)."""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id, filename, original_name, mime_type, size_bytes, file_type, run_id, created_at FROM files WHERE id = ?", (file_id,))
    row = cur.fetchone()
    conn.close()
    
    if not row:
        return jsonify({"error": "File not found"}), 404
    
    return jsonify({
        "id": row["id"],
        "filename": row["filename"],
        "originalName": row["original_name"],
        "mimeType": row["mime_type"],
        "sizeBytes": row["size_bytes"],
        "fileType": row["file_type"],
        "runId": row["run_id"],
        "createdAt": row["created_at"]
    })


@app.route("/files/<file_id>/download", methods=["GET"])
def download_file(file_id):
    """Download a file by ID. Path is resolved internally."""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT filename, original_name, mime_type, file_type, run_id FROM files WHERE id = ?", (file_id,))
    row = cur.fetchone()
    conn.close()
    
    if not row:
        return jsonify({"error": "File not found"}), 404
    
    filename = row["filename"]
    original_name = row["original_name"] or filename
    mime_type = row["mime_type"] or "application/octet-stream"
    file_type = row["file_type"]
    run_id = row["run_id"]
    
    # Resolve internal path based on file type
    if file_type in ("pdf", "crawled_pdf"):
        filepath = os.path.join(UPLOAD_FOLDER, run_id, "pdfs", filename)
    elif file_type == "schema":
        filepath = os.path.join(UPLOAD_FOLDER, run_id, filename)
    elif file_type == "export":
        filepath = os.path.join(EXPORTS_FOLDER, run_id, filename)
    elif file_type == "zip":
        filepath = os.path.join(UPLOAD_FOLDER, run_id, filename)
    else:
        return jsonify({"error": "Unknown file type"}), 400
    
    if not os.path.exists(filepath):
        return jsonify({"error": "File not found on disk"}), 404
    
    return send_file(filepath, mimetype=mime_type, as_attachment=True, download_name=original_name)


@app.route("/runs/<run_id>/files", methods=["GET"])
def get_run_files(run_id):
    """Get all files for a run (no paths exposed)."""
    file_type = request.args.get("type")
    
    conn = get_db()
    cur = conn.cursor()
    
    query = "SELECT id, filename, original_name, mime_type, size_bytes, file_type, created_at FROM files WHERE run_id = ?"
    params = [run_id]
    
    if file_type:
        query += " AND file_type = ?"
        params.append(file_type)
    
    cur.execute(query, params)
    rows = cur.fetchall()
    conn.close()
    
    files = []
    for row in rows:
        files.append({
            "id": row["id"],
            "filename": row["filename"],
            "originalName": row["original_name"],
            "mimeType": row["mime_type"],
            "sizeBytes": row["size_bytes"],
            "fileType": row["file_type"],
            "createdAt": row["created_at"]
        })
    
    return jsonify({"items": files, "total": len(files), "runId": run_id})

@app.route("/runs/<run_id>/engine/status", methods=["GET"])
def get_engine_status(run_id):
    """Get engine status for a run."""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT status FROM runs WHERE id = ?", (run_id,))
    row = cur.fetchone()
    conn.close()
    
    if not row:
        return jsonify({"error": "Run not found"}), 404
    
    status = row["status"]
    crashed = status == "failed"
    
    return jsonify({
        "runId": run_id,
        "state": status,
        "crashed": crashed,
        "crashCount": 1 if crashed else 0,
        "crashes": []
    })

@app.route("/runs/<run_id>/engine/data", methods=["GET"])
def get_engine_data(run_id):
    """Get engine data for a run."""
    data_path = os.path.join(IPC_DIR, run_id, "data.json")
    
    if os.path.exists(data_path):
        with open(data_path, "r") as f:
            return jsonify(json.load(f))
    
    return jsonify({})

@app.route("/runs/<run_id>/logs", methods=["GET"])
def get_run_logs(run_id):
    """Get logs for a run."""
    tail_lines = int(request.args.get("tailLines", 500))
    
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "SELECT * FROM logs WHERE run_id = ? ORDER BY id DESC LIMIT ?",
        (run_id, tail_lines)
    )
    rows = [to_camel_dict(dict(r)) for r in cur.fetchall()]
    rows.reverse()  # Oldest first
    conn.close()
    
    content = "\n".join([f"[{r['createdAt']}] {r['level']}: {r['message']}" for r in rows])
    
    return jsonify({
        "runId": run_id,
        "content": content,
        "lines": rows,
        "total": len(rows)
    })

@app.route("/runs/<run_id>/engine/logs", methods=["GET"])
def get_engine_logs(run_id):
    """Get engine stdout/stderr/extraction logs for a run.
    
    Returns the actual process output from the spawned extraction subprocess.
    """
    import re
    
    run_ipc_dir = os.path.join(IPC_DIR, run_id)
    stdout_path = os.path.join(run_ipc_dir, "stdout.log")
    stderr_path = os.path.join(run_ipc_dir, "stderr.log")
    extraction_log_path = os.path.join(run_ipc_dir, "extraction.log")
    
    stdout = ""
    stderr = ""
    extraction_log = ""
    
    # Read raw log files
    if os.path.exists(stdout_path):
        with open(stdout_path, "r", encoding="utf-8", errors="replace") as f:
            stdout = f.read()[-65536:]  # Last 64KB
    
    if os.path.exists(stderr_path):
        with open(stderr_path, "r", encoding="utf-8", errors="replace") as f:
            stderr = f.read()[-65536:]
    
    if os.path.exists(extraction_log_path):
        with open(extraction_log_path, "r", encoding="utf-8", errors="replace") as f:
            extraction_log = f.read()[-65536:]
    
    # Parse extraction log into structured entries
    log_entries = []
    if extraction_log:
        for line in extraction_log.strip().split("\n"):
            if not line.strip():
                continue
            # Try to parse structured log format: TIMESTAMP - LEVEL - MESSAGE
            match = re.match(r'^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2},\d{3})\s*-\s*(\w+)\s*-\s*(.*)$', line)
            if match:
                log_entries.append({
                    "timestamp": match.group(1),
                    "level": match.group(2),
                    "message": match.group(3),
                    "source": "extraction"
                })
            else:
                # Fallback for unstructured lines
                log_entries.append({
                    "timestamp": "",
                    "level": "INFO",
                    "message": line,
                    "source": "extraction"
                })
    
    # Also parse stdout for any structured output
    if stdout:
        for line in stdout.strip().split("\n"):
            if not line.strip():
                continue
            log_entries.append({
                "timestamp": "",
                "level": "INFO",
                "message": line,
                "source": "stdout"
            })
    
    # Parse stderr as errors/warnings
    if stderr:
        for line in stderr.strip().split("\n"):
            if not line.strip():
                continue
            level = "ERROR" if "error" in line.lower() else "WARN"
            log_entries.append({
                "timestamp": "",
                "level": level,
                "message": line,
                "source": "stderr"
            })
    
    return jsonify({
        "runId": run_id,
        "stdout": stdout,
        "stderr": stderr,
        "extractionLog": extraction_log,
        "entries": log_entries,
        "total": len(log_entries)
    })


@app.route("/runs/<run_id>/cache", methods=["GET"])
def get_run_cache_stats(run_id):
    """Get cache hit/miss statistics for a run.
    
    Parses extraction logs and stdout for cache-related messages.
    Returns categorized cache events by provider (Surya, Gemini, OpenAI, etc.)
    """
    import re
    
    run_ipc_dir = os.path.join(IPC_DIR, run_id)
    stdout_path = os.path.join(run_ipc_dir, "stdout.log")
    extraction_log_path = os.path.join(run_ipc_dir, "extraction.log")
    
    cache_events = []
    
    # Patterns to match cache log lines
    cache_hit_pattern = re.compile(r'\[CACHE HIT\]\s*(\w+):\s*(.+)')
    cache_miss_pattern = re.compile(r'\[CACHE MISS\]\s*(\w+):\s*(.+?)(?:\s*→.*)?$')
    cache_skip_pattern = re.compile(r'\[CACHE SKIP\]\s*(\w+):\s*(.+)')
    surya_hit_pattern = re.compile(r'\[CACHE HIT\]\s*Surya:\s*(.+)')
    gpt_hit_pattern = re.compile(r'\[CACHE HIT\]\s*GPT response')
    
    def parse_log_file(filepath, source_name):
        if not os.path.exists(filepath):
            return
        with open(filepath, "r", encoding="utf-8", errors="replace") as f:
            for line_num, line in enumerate(f, 1):
                line = line.strip()
                if not line:
                    continue
                
                # Check for cache hit
                hit_match = cache_hit_pattern.search(line)
                if hit_match:
                    provider = hit_match.group(1).upper()
                    details = hit_match.group(2).strip()
                    cache_events.append({
                        "type": "hit",
                        "provider": provider,
                        "details": details,
                        "source": source_name,
                        "line": line_num,
                        "raw": line
                    })
                    continue
                
                # Check for cache miss
                miss_match = cache_miss_pattern.search(line)
                if miss_match:
                    provider = miss_match.group(1).upper()
                    details = miss_match.group(2).strip()
                    cache_events.append({
                        "type": "miss",
                        "provider": provider,
                        "details": details,
                        "source": source_name,
                        "line": line_num,
                        "raw": line
                    })
                    continue
                
                # Check for cache skip
                skip_match = cache_skip_pattern.search(line)
                if skip_match:
                    provider = skip_match.group(1).upper()
                    details = skip_match.group(2).strip()
                    cache_events.append({
                        "type": "skip",
                        "provider": provider,
                        "details": details,
                        "source": source_name,
                        "line": line_num,
                        "raw": line
                    })
                    continue
                
                # Legacy GPT response hit
                if gpt_hit_pattern.search(line):
                    cache_events.append({
                        "type": "hit",
                        "provider": "GPT",
                        "details": "GPT response",
                        "source": source_name,
                        "line": line_num,
                        "raw": line
                    })
    
    # Parse both log files
    parse_log_file(stdout_path, "stdout")
    parse_log_file(extraction_log_path, "extraction")
    
    # Compute summary statistics
    summary = {
        "totalHits": sum(1 for e in cache_events if e["type"] == "hit"),
        "totalMisses": sum(1 for e in cache_events if e["type"] == "miss"),
        "totalSkips": sum(1 for e in cache_events if e["type"] == "skip"),
        "byProvider": {}
    }
    
    for event in cache_events:
        provider = event["provider"]
        if provider not in summary["byProvider"]:
            summary["byProvider"][provider] = {"hits": 0, "misses": 0, "skips": 0}
        if event["type"] == "hit":
            summary["byProvider"][provider]["hits"] += 1
        elif event["type"] == "miss":
            summary["byProvider"][provider]["misses"] += 1
        elif event["type"] == "skip":
            summary["byProvider"][provider]["skips"] += 1
    
    # Calculate hit rate per provider
    for provider, stats in summary["byProvider"].items():
        total = stats["hits"] + stats["misses"] + stats["skips"]
        stats["total"] = total
        stats["hitRate"] = round(stats["hits"] / total * 100, 1) if total > 0 else 0
    
    return jsonify({
        "runId": run_id,
        "events": cache_events,
        "summary": summary,
        "total": len(cache_events)
    })


@app.route("/runs/<run_id>/api-analytics", methods=["GET"])
def get_run_api_analytics(run_id):
    """Get API call analytics for a run.
    
    Parses extraction logs for API call timing information.
    Returns per-provider statistics including call count, total time, and averages.
    """
    import re
    
    run_ipc_dir = os.path.join(IPC_DIR, run_id)
    stdout_path = os.path.join(run_ipc_dir, "stdout.log")
    extraction_log_path = os.path.join(run_ipc_dir, "extraction.log")
    
    api_calls = []
    
    # Pattern to match API call timing logs: [API CALL] PROVIDER: model completed in Xms
    api_call_pattern = re.compile(r'\[API CALL\]\s*(\w+):\s*([^\s]+)\s+completed in\s+(\d+)ms')
    
    def parse_log_file(filepath, source_name):
        if not os.path.exists(filepath):
            return
        with open(filepath, "r", encoding="utf-8", errors="replace") as f:
            for line_num, line in enumerate(f, 1):
                line = line.strip()
                if not line:
                    continue
                
                match = api_call_pattern.search(line)
                if match:
                    provider = match.group(1).upper()
                    model = match.group(2)
                    duration_ms = int(match.group(3))
                    api_calls.append({
                        "provider": provider,
                        "model": model,
                        "durationMs": duration_ms,
                        "source": source_name,
                        "line": line_num
                    })
    
    # Parse both log files
    parse_log_file(stdout_path, "stdout")
    parse_log_file(extraction_log_path, "extraction")
    
    # Compute summary statistics
    summary = {
        "totalCalls": len(api_calls),
        "totalTimeMs": sum(c["durationMs"] for c in api_calls),
        "avgTimeMs": 0,
        "byProvider": {}
    }
    
    if api_calls:
        summary["avgTimeMs"] = round(summary["totalTimeMs"] / len(api_calls))
    
    for call in api_calls:
        provider = call["provider"]
        if provider not in summary["byProvider"]:
            summary["byProvider"][provider] = {
                "calls": 0,
                "totalTimeMs": 0,
                "avgTimeMs": 0,
                "minTimeMs": float('inf'),
                "maxTimeMs": 0,
                "models": {}
            }
        
        stats = summary["byProvider"][provider]
        stats["calls"] += 1
        stats["totalTimeMs"] += call["durationMs"]
        stats["minTimeMs"] = min(stats["minTimeMs"], call["durationMs"])
        stats["maxTimeMs"] = max(stats["maxTimeMs"], call["durationMs"])
        
        # Track per-model stats
        model = call["model"]
        if model not in stats["models"]:
            stats["models"][model] = {"calls": 0, "totalTimeMs": 0}
        stats["models"][model]["calls"] += 1
        stats["models"][model]["totalTimeMs"] += call["durationMs"]
    
    # Calculate averages
    for provider, stats in summary["byProvider"].items():
        if stats["calls"] > 0:
            stats["avgTimeMs"] = round(stats["totalTimeMs"] / stats["calls"])
        if stats["minTimeMs"] == float('inf'):
            stats["minTimeMs"] = 0
        for model_stats in stats["models"].values():
            if model_stats["calls"] > 0:
                model_stats["avgTimeMs"] = round(model_stats["totalTimeMs"] / model_stats["calls"])
    
    return jsonify({
        "runId": run_id,
        "calls": api_calls,
        "summary": summary,
        "total": len(api_calls)
    })


@app.route("/runs/<run_id>/progress", methods=["GET"])
def get_run_progress(run_id):
    """Get extraction progress for a run.
    
    Reads progress.json from the run's output directory.
    Returns processed/total PDFs, current file, and entries extracted.
    """
    conn = get_db()
    cur = conn.cursor()
    row = None
    sources_count = 0
    try:
        cur.execute("SELECT output_dir, status, sources_count FROM runs WHERE id = ?", (run_id,))
        row = cur.fetchone()
        if row is not None:
            sources_count = row["sources_count"] if row["sources_count"] is not None else 0
    except Exception:
        try:
            cur.execute("SELECT output_dir, status, articles_count FROM runs WHERE id = ?", (run_id,))
            row = cur.fetchone()
            if row is not None:
                sources_count = row["articles_count"] if row["articles_count"] is not None else 0
        except Exception:
            cur.execute("SELECT output_dir, status FROM runs WHERE id = ?", (run_id,))
            row = cur.fetchone()
    conn.close()
    
    if not row:
        return jsonify({"error": "Run not found"}), 404
    
    output_dir = row["output_dir"]
    status = row["status"]
    sources_count = sources_count
    
    # Default progress response
    progress = {
        "runId": run_id,
        "processed": 0,
        "total": sources_count,
        "currentFile": "",
        "status": status,
        "entriesExtracted": 0,
        "percentComplete": 0.0,
        "updatedAt": None
    }
    
    # Try to read progress.json if it exists
    if output_dir:
        progress_path = os.path.join(output_dir, "progress.json")
        if os.path.exists(progress_path):
            try:
                with open(progress_path, "r", encoding="utf-8") as f:
                    file_progress = json.load(f)
                    progress.update({
                        "processed": file_progress.get("processed", 0),
                        "total": file_progress.get("total", sources_count),
                        "currentFile": file_progress.get("currentFile", ""),
                        "entriesExtracted": file_progress.get("entriesExtracted", 0),
                        "percentComplete": file_progress.get("percentComplete", 0.0),
                        "updatedAt": file_progress.get("updatedAt")
                    })
            except:
                pass
    
    return jsonify(progress)

@app.route("/runs/<run_id>/pause", methods=["POST"])
def pause_run(run_id):
    """Pause a run."""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE runs SET status = 'PAUSED' WHERE id = ?", (run_id,))
    conn.commit()
    conn.close()
    log_message(f"Run paused", "INFO", run_id)
    return "", 204

@app.route("/runs/<run_id>/resume", methods=["POST"])
def resume_run(run_id):
    """Resume a paused run."""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE runs SET status = 'running' WHERE id = ?", (run_id,))
    conn.commit()
    conn.close()
    log_message(f"Run resumed", "INFO", run_id)
    return "", 204

@app.route("/runs/<run_id>/start", methods=["POST"])
def start_run(run_id):
    """Start extraction process for a run.
    
    For PDF runs: Uses pdfsDir and excelPath stored in the run configuration.
    For Deep Research/Links runs: Uses sources store (physical table name: sources) and excelPath.
    Optional request body can override instructions.
    """
    # Use get_json with silent=True to handle empty body gracefully
    data = request.get_json(silent=True) or {}
    
    # Get run details
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM runs WHERE id = ?", (run_id,))
    row = cur.fetchone()
    conn.close()
    
    if not row:
        return jsonify({"error": "Run not found"}), 404
    
    run = dict(row)
    source_type = run.get("source_type", "pdf")
    
    # Check if already running
    if run_id in active_processes:
        return jsonify({"error": "Run is already in progress"}), 400
    
    excel_path = run.get("excel_path")
    output_dir = run.get("output_dir")
    
    if not excel_path:
        return jsonify({"error": "Run not properly configured: excelPath required"}), 400
    
    if not os.path.isfile(excel_path):
        return jsonify({"error": f"Excel file not found: {excel_path}"}), 400
    
    # Ensure output directory exists
    output_dir = output_dir or os.path.join(EXPORTS_FOLDER, run_id)
    os.makedirs(output_dir, exist_ok=True)
    
    # Get instructions from request or run
    instructions = data.get("instructions") or run.get("prompt", "")
    llm_provider = run.get("llm_provider") or LLM_PROVIDER
    
    # Route based on source type
    if source_type in ("deep_research", "links"):
        # Use extract_html.py for HTML content from sources store (physical table name: sources)
        spawn_html_extraction_process(
            run_id=run_id,
            excel_path=excel_path,
            output_dir=output_dir,
            instructions=instructions,
            llm_provider=llm_provider,
            user_id=run.get("user_id")
        )
        return jsonify({"message": "HTML extraction started", "runId": run_id, "sourceType": source_type}), 202
    else:
        # PDF source type - use original extract.py
        pdfs_dir = run.get("pdfs_dir")
        
        if not pdfs_dir:
            return jsonify({"error": "Run not properly configured: pdfsDir required for PDF runs"}), 400
        
        if not os.path.isdir(pdfs_dir):
            return jsonify({"error": f"PDF directory not found: {pdfs_dir}"}), 400
        
        enable_row_counting = bool(run.get("enable_row_counting", 0))
        validation_enabled = bool(run.get("validation_enabled", 0))
        validation_max_retries = run.get("validation_max_retries") or 3
        
        # Get validation prompt path from file ID
        validation_prompt_path = None
        validation_prompt_file_id = run.get("validation_prompt_file_id")
        if validation_prompt_file_id:
            conn2 = get_db()
            cur2 = conn2.cursor()
            cur2.execute("SELECT filename, run_id FROM files WHERE id = ?", (validation_prompt_file_id,))
            vp_row = cur2.fetchone()
            conn2.close()
            if vp_row:
                file_run_id = vp_row["run_id"] or run_id
                validation_prompt_path = os.path.join(UPLOAD_FOLDER, file_run_id, vp_row["filename"])
        
        # Parse cache flags from run
        cache_flags = None
        cache_flags_str = run.get("cache_flags")
        if cache_flags_str:
            try:
                cache_flags = json.loads(cache_flags_str)
            except:
                pass
        
        # Spawn extraction process
        spawn_extraction_process(
            run_id=run_id,
            pdfs_dir=pdfs_dir,
            excel_path=excel_path,
            output_dir=output_dir,
            instructions=instructions,
            llm_provider=llm_provider,
            enable_row_counting=enable_row_counting,
            user_id=run.get("user_id"),
            validation_prompt_path=validation_prompt_path,
            validation_enabled=validation_enabled,
            validation_max_retries=validation_max_retries,
            cache_flags=cache_flags
        )
        
        return jsonify({"message": "PDF extraction started", "runId": run_id, "sourceType": "pdf"}), 202


def normalize_retry_name(name: str) -> str:
    """Normalize run name to use (Retry N) format instead of repeated (Retry).
    
    Handles:
    - "Run Name (Retry) (Retry) (Retry)" -> "Run Name (Retry 3)"
    - "Run Name (Retry 5)" -> "Run Name (Retry 6)"
    - "Run Name" -> "Run Name (Retry 1)"
    
    Always stores in correct format, even if input is wrong format.
    """
    if not name:
        name = "Untitled Run"
    
    # Pattern 1: Count repeated (Retry) at the end
    repeated_pattern = r'^(.+?)((?:\s*\(Retry\))+)\s*$'
    match = re.match(repeated_pattern, name)
    if match:
        base_name = match.group(1).strip()
        retry_count = match.group(2).count('(Retry)')
        return f"{base_name} (Retry {retry_count + 1})"
    
    # Pattern 2: Already has (Retry N) format
    numbered_pattern = r'^(.+?)\s*\(Retry\s+(\d+)\)\s*$'
    match = re.match(numbered_pattern, name)
    if match:
        base_name = match.group(1).strip()
        current_count = int(match.group(2))
        return f"{base_name} (Retry {current_count + 1})"
    
    # No retry suffix - this is the first retry
    return f"{name} (Retry 1)"


@app.route("/runs/<run_id>/retry", methods=["POST"])
@require_auth
def retry_run(run_id):
    """Retry a run by creating a NEW run with the same inputs.

    This creates a fresh run_id, copies stored artifacts (schema/zip) and recreates
    crawl jobs or deep research as needed.
    """
    import zipfile

    user_id = g.current_user["id"]
    payload = request.get_json(silent=True) or {}
    auto_start = bool(payload.get("autoStart", True))

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM runs WHERE id = ?", (run_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return jsonify({"error": "Run not found"}), 404

    src = dict(row)
    if src.get("user_id") != user_id:
        conn.close()
        return jsonify({"error": "Forbidden"}), 403

    source_type = src.get("source_type", "pdf")
    src_excel_path = src.get("excel_path")
    src_pdfs_dir = src.get("pdfs_dir")
    src_zip_file_id = src.get("zip_file_id")
    src_schema_file_id = src.get("schema_file_id")

    has_schema = src_excel_path and os.path.isfile(src_excel_path)
    if not has_schema:
        conn.close()
        return jsonify({"error": "Source run is missing schema file on disk"}), 400

    new_run_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    run_upload_dir = os.path.join(UPLOAD_FOLDER, new_run_id)
    output_dir = os.path.join(EXPORTS_FOLDER, new_run_id)
    os.makedirs(run_upload_dir, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)

    # Copy schema
    schema_filename = secure_filename(os.path.basename(src_excel_path))
    new_excel_path = os.path.join(run_upload_dir, schema_filename)
    shutil.copy2(src_excel_path, new_excel_path)
    schema_file_id = register_file(
        new_excel_path,
        os.path.basename(src_excel_path),
        "schema",
        new_run_id,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    llm_provider = src.get("llm_provider") or LLM_PROVIDER
    prompt = src.get("prompt") or ""
    enable_row_counting = 1 if bool(src.get("enable_row_counting", 0)) else 0
    name = normalize_retry_name(src.get("name") or "Untitled Run")
    
    # Copy validation and prompt settings from source run
    src_extraction_prompt_file_id = src.get("extraction_prompt_file_id")
    src_validation_prompt_file_id = src.get("validation_prompt_file_id")
    validation_enabled = 1 if bool(src.get("validation_enabled", 0)) else 0
    validation_max_retries = src.get("validation_max_retries") or 3
    src_cache_flags = src.get("cache_flags")
    
    # Copy extraction prompt file if exists
    new_extraction_prompt_file_id = None
    new_extraction_prompt_path = None
    if src_extraction_prompt_file_id:
        src_ep_path = get_file_internal_path(src_extraction_prompt_file_id)
        if src_ep_path and os.path.isfile(src_ep_path):
            ep_filename = secure_filename(os.path.basename(src_ep_path))
            new_extraction_prompt_path = os.path.join(run_upload_dir, ep_filename)
            shutil.copy2(src_ep_path, new_extraction_prompt_path)
            new_extraction_prompt_file_id = register_file(new_extraction_prompt_path, ep_filename, "extraction_prompt", new_run_id, "text/plain")
    
    # Copy validation prompt file if exists
    new_validation_prompt_file_id = None
    new_validation_prompt_path = None
    if src_validation_prompt_file_id:
        src_vp_path = get_file_internal_path(src_validation_prompt_file_id)
        if src_vp_path and os.path.isfile(src_vp_path):
            vp_filename = secure_filename(os.path.basename(src_vp_path))
            new_validation_prompt_path = os.path.join(run_upload_dir, vp_filename)
            shutil.copy2(src_vp_path, new_validation_prompt_path)
            new_validation_prompt_file_id = register_file(new_validation_prompt_path, vp_filename, "validation_prompt", new_run_id, "text/plain")

    if source_type == "pdf":
        if not src_pdfs_dir or not os.path.isdir(src_pdfs_dir):
            conn.close()
            return jsonify({"error": "Source run is missing PDFs directory on disk"}), 400

        # Copy zip if present, otherwise reconstruct a new zip from PDFs folder
        new_zip_file_id = None
        new_zip_path = None
        if src_zip_file_id:
            src_zip_path = get_file_internal_path(src_zip_file_id)
            if src_zip_path and os.path.isfile(src_zip_path):
                zip_filename = secure_filename(os.path.basename(src_zip_path))
                new_zip_path = os.path.join(run_upload_dir, zip_filename)
                shutil.copy2(src_zip_path, new_zip_path)
        if new_zip_path is None:
            zip_filename = "pdfs.zip"
            new_zip_path = os.path.join(run_upload_dir, zip_filename)
            with zipfile.ZipFile(new_zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                for fn in os.listdir(src_pdfs_dir):
                    if fn.lower().endswith(".pdf"):
                        zf.write(os.path.join(src_pdfs_dir, fn), arcname=fn)

        new_zip_file_id = register_file(new_zip_path, os.path.basename(new_zip_path), "zip", new_run_id, "application/zip")

        pdfs_dir = os.path.join(run_upload_dir, "pdfs")
        os.makedirs(pdfs_dir, exist_ok=True)
        pdf_count = 0
        with zipfile.ZipFile(new_zip_path, "r") as zf:
            for member in zf.namelist():
                if member.lower().endswith(".pdf") and not member.startswith("__MACOSX"):
                    zf.extract(member, pdfs_dir)
                    extracted_path = os.path.join(pdfs_dir, member)
                    if os.path.dirname(member):
                        flat_path = os.path.join(pdfs_dir, os.path.basename(member))
                        if extracted_path != flat_path:
                            os.rename(extracted_path, flat_path)
                    pdf_count += 1
        for root, dirs, _files in os.walk(pdfs_dir, topdown=False):
            for d in dirs:
                try:
                    os.rmdir(os.path.join(root, d))
                except Exception:
                    pass
        if pdf_count == 0:
            conn.close()
            return jsonify({"error": "No PDF files found to retry"}), 400

        # Register individual PDFs
        for pdf_file in os.listdir(pdfs_dir):
            if pdf_file.lower().endswith(".pdf"):
                register_file(os.path.join(pdfs_dir, pdf_file), pdf_file, "pdf", new_run_id, "application/pdf")

        # Create meta source for provenance
        meta_source_id = create_meta_source(run_id=new_run_id, method="retry", user_id=user_id, name=name)

        cur.execute(
            """
            INSERT INTO runs (id, name, source_type, status, start_date, llm_provider, pdfs_dir, excel_path, output_dir,
                              prompt, search_methods, search_queries, links, table_file_url, per_link_prompt,
                              sources_count, schema_file_id, zip_file_id, enable_row_counting, user_id, meta_source_id,
                              extraction_prompt_file_id, validation_prompt_file_id, validation_enabled, validation_max_retries, cache_flags)
            VALUES (?, ?, 'pdf', 'waiting', ?, ?, ?, ?, ?, ?, '[]', '[]', '[]', '', '', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                new_run_id,
                name,
                now,
                llm_provider,
                pdfs_dir,
                new_excel_path,
                output_dir,
                prompt,
                pdf_count,
                schema_file_id,
                new_zip_file_id,
                enable_row_counting,
                user_id,
                meta_source_id,
                new_extraction_prompt_file_id,
                new_validation_prompt_file_id,
                validation_enabled,
                validation_max_retries,
                src_cache_flags,
            ),
        )

        # Insert source rows for each PDF file (same pattern as create_run)
        cur.execute(
            "SELECT id, original_name, filename FROM files WHERE run_id = ? AND file_type IN ('pdf','crawled_pdf')",
            (new_run_id,),
        )
        for f in cur.fetchall():
            pdf_file_id = f["id"]
            title = f["original_name"] or f["filename"]
            ensure_source_row(
                source_id=pdf_file_id,
                run_id=new_run_id,
                source_type="pdf",
                status="PENDING",
                url=None,
                title=title,
                crawl_job_id=None,
                pdf_file_id=pdf_file_id,
                meta_source_id=meta_source_id,
                cur=cur,
            )

        conn.commit()
        conn.close()

        log_message(f"Run retried from {run_id}: {name}", "INFO", new_run_id)

        # Start background conversion for PDF sources so they transition to READY
        try:
            threading.Thread(target=process_pdf_sources_for_run, args=(new_run_id, user_id), daemon=True).start()
        except Exception:
            pass

        if auto_start:
            # Parse cache flags from source run
            cache_flags_dict = None
            if src_cache_flags:
                try:
                    cache_flags_dict = json.loads(src_cache_flags) if isinstance(src_cache_flags, str) else src_cache_flags
                except Exception:
                    pass
            
            spawn_extraction_process(
                run_id=new_run_id,
                pdfs_dir=pdfs_dir,
                excel_path=new_excel_path,
                output_dir=output_dir,
                instructions=prompt,
                llm_provider=llm_provider,
                enable_row_counting=bool(enable_row_counting),
                user_id=user_id,
                validation_prompt_path=new_validation_prompt_path,
                validation_enabled=bool(validation_enabled),
                validation_max_retries=validation_max_retries,
                cache_flags=cache_flags_dict,
            )

        return jsonify(
            {
                "id": new_run_id,
                "name": name,
                "sourceType": "pdf",
                "status": "running" if auto_start else "waiting",
                "startDate": now,
                "sourcesCount": pdf_count,
                "dataEntriesCount": 0,
                "llmProvider": llm_provider,
                "prompt": prompt,
                "searchMethods": [],
                "searchQueries": [],
                "schemaFileId": schema_file_id,
                "zipFileId": new_zip_file_id,
                "enableRowCounting": bool(enable_row_counting),
                "validationEnabled": bool(validation_enabled),
                "validationMaxRetries": validation_max_retries,
            }
        ), 201

    if source_type == "links":
        # Clone links run and recreate crawl jobs
        links_raw = src.get("links") or "[]"
        try:
            links = json.loads(links_raw)
        except Exception:
            links = []
        if not isinstance(links, list) or not links:
            conn.close()
            return jsonify({"error": "Source run has no links to retry"}), 400

        cur.execute(
            """
            INSERT INTO runs (id, name, source_type, status, start_date, llm_provider, excel_path, output_dir,
                              prompt, links, schema_file_id, sources_count, user_id,
                              extraction_prompt_file_id, validation_prompt_file_id, validation_enabled, validation_max_retries, cache_flags)
            VALUES (?, ?, 'links', 'crawling', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (new_run_id, name, now, llm_provider, new_excel_path, output_dir, prompt, json.dumps(links), schema_file_id, len(links), user_id,
             new_extraction_prompt_file_id, new_validation_prompt_file_id, validation_enabled, validation_max_retries, src_cache_flags),
        )
        conn.commit()

        # Recreate crawl_jobs (same logic as create_run_from_links)
        html_count = 0
        pdf_count = 0
        for link in links:
            url = link.get("url", "") if isinstance(link, dict) else str(link).strip()
            title = link.get("title", "") if isinstance(link, dict) else ""
            if not url:
                continue
            url_lower = url.lower()
            job_id = str(uuid.uuid4())
            if url_lower.endswith('.pdf') or '/pdf/' in url_lower or 'pdf?' in url_lower:
                pdf_count += 1
                cur.execute(
                    """
                    INSERT INTO crawl_jobs (id, run_id, user_id, url, title, status, created_at)
                    VALUES (?, ?, ?, ?, ?, 'PDF_PENDING', ?)
                    """,
                    (job_id, new_run_id, user_id, url, title, now),
                )
            else:
                cur.execute(
                    """
                    INSERT INTO crawl_jobs (id, run_id, user_id, url, title, status, created_at)
                    VALUES (?, ?, ?, ?, ?, 'PENDING', ?)
                    """,
                    (job_id, new_run_id, user_id, url, title, now),
                )
                html_count += 1
        conn.commit()
        conn.close()

        if pdf_count > 0:
            threading.Thread(target=process_pdf_jobs_for_run, args=(new_run_id, user_id), daemon=True).start()

        log_message(f"Links run retried from {run_id}: {name}", "INFO", new_run_id)
        return jsonify(
            {
                "id": new_run_id,
                "name": name,
                "sourceType": "links",
                "status": "crawling",
                "startDate": now,
                "sourcesCount": len(links),
                "dataEntriesCount": 0,
                "llmProvider": llm_provider,
                "prompt": prompt,
                "schemaFileId": schema_file_id,
                "validationEnabled": bool(validation_enabled),
                "validationMaxRetries": validation_max_retries,
            }
        ), 201

    if source_type == "deep_research":
        query = src.get("deep_research_query") or ""
        if not query.strip():
            conn.close()
            return jsonify({"error": "Source run has no deep research query to retry"}), 400
        api_key = get_gemini_api_key(user_id)
        if not api_key:
            conn.close()
            return jsonify({"error": "GEMINI_API_KEY not configured. Set it in Config > API Keys."}), 400

        cur.execute(
            """
            INSERT INTO runs (id, name, source_type, status, start_date, llm_provider, excel_path, output_dir,
                              prompt, deep_research_query, schema_file_id, sources_count, user_id,
                              extraction_prompt_file_id, validation_prompt_file_id, validation_enabled, validation_max_retries, cache_flags)
            VALUES (?, ?, 'deep_research', 'searching', ?, ?, ?, ?, ?, ?, ?, 0, ?, ?, ?, ?, ?, ?)
            """,
            (new_run_id, name, now, llm_provider, new_excel_path, output_dir, prompt, query, schema_file_id, user_id,
             new_extraction_prompt_file_id, new_validation_prompt_file_id, validation_enabled, validation_max_retries, src_cache_flags),
        )
        conn.commit()
        conn.close()

        def run_deep_research_retry():
            try:
                import requests
                import re
                from cache_utils import get_gpt_cache, set_gpt_cache

                conn2 = get_db()
                cur2 = conn2.cursor()
                cur2.execute("UPDATE runs SET status = 'researching' WHERE id = ?", (new_run_id,))
                conn2.commit()
                conn2.close()

                log_message(f"Starting Gemini Deep Research for query: {query}", "INFO", new_run_id)

                # Check cache first
                cache_model = "gemini-deep-research:gemini-2.0-flash"
                cached_response = get_gpt_cache("", query, cache_model)
                
                if cached_response:
                    log_message(f"[CACHE HIT] Deep Research - using cached response", "INFO", new_run_id)
                    data = cached_response
                else:
                    log_message(f"[CACHE MISS] Deep Research - calling Gemini API...", "INFO", new_run_id)
                    gemini_url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={api_key}"
                    payload = {
                        "contents": [{"parts": [{"text": query}]}],
                        "tools": [{"google_search": {}}],
                    }
                    resp = requests.post(gemini_url, json=payload, timeout=300)
                    resp.raise_for_status()
                    data = resp.json()
                    
                    # Cache the response
                    set_gpt_cache("", query, cache_model, data)
                    log_message(f"[CACHE STORED] Deep Research response cached", "INFO", new_run_id)

                # Log response structure
                log_message(f"Response keys: {list(data.keys())}", "DEBUG", new_run_id)
                
                # Extract text and links from grounding metadata (same as main function)
                result_text = ""
                extracted_links = []
                
                if "candidates" in data:
                    log_message(f"Candidates count: {len(data['candidates'])}", "DEBUG", new_run_id)
                    for idx, candidate in enumerate(data["candidates"]):
                        content = candidate.get("content", {})
                        parts = content.get("parts", [])
                        log_message(f"Candidate {idx}: {len(parts)} parts, keys: {list(candidate.keys())}", "DEBUG", new_run_id)
                        
                        for part in parts:
                            if "text" in part:
                                result_text += part["text"] + "\n"
                        
                        # Extract links from grounding metadata
                        grounding_metadata = candidate.get("groundingMetadata", {})
                        log_message(f"Grounding metadata keys: {list(grounding_metadata.keys())}", "DEBUG", new_run_id)
                        
                        grounding_chunks = grounding_metadata.get("groundingChunks", [])
                        log_message(f"Grounding chunks count: {len(grounding_chunks)}", "DEBUG", new_run_id)
                        
                        for chunk in grounding_chunks:
                            web = chunk.get("web", {})
                            url = web.get("uri", "") or web.get("url", "")
                            title = web.get("title", "")
                            if url:
                                extracted_links.append({"url": url, "title": title})
                
                log_message(f"Links from grounding metadata: {len(extracted_links)}", "DEBUG", new_run_id)
                log_message(f"Result text length: {len(result_text)} chars", "DEBUG", new_run_id)
                if result_text:
                    log_message(f"Result text preview: {result_text[:500]}...", "DEBUG", new_run_id)
                
                # Fallback: extract URLs from text if no grounding metadata
                if not extracted_links and result_text:
                    log_message(f"No grounding links found, extracting URLs from text...", "INFO", new_run_id)
                    urls = re.findall(r"https?://[^\s\]\)\"']+", result_text)
                    urls = list(dict.fromkeys(urls))
                    log_message(f"URLs found in text: {len(urls)}", "DEBUG", new_run_id)
                    for url in urls[:50]:
                        extracted_links.append({"url": url, "title": ""})
                
                # Deduplicate
                seen_urls = set()
                unique_links = []
                for link in extracted_links:
                    url = link.get("url", "")
                    if url and url not in seen_urls:
                        seen_urls.add(url)
                        unique_links.append(link)
                
                log_message(f"Found {len(unique_links)} unique URLs from Deep Research", "INFO", new_run_id)

                conn3 = get_db()
                cur3 = conn3.cursor()
                cur3.execute("UPDATE runs SET deep_research_result = ?, status = 'crawling', sources_count = ? WHERE id = ?", (result_text, len(unique_links), new_run_id))

                for link in unique_links:
                    url = link.get("url", "")
                    title = link.get("title", "")
                    job_id = str(uuid.uuid4())
                    cur3.execute(
                        """
                        INSERT INTO crawl_jobs (id, run_id, user_id, url, title, status, created_at)
                        VALUES (?, ?, ?, ?, ?, 'PENDING', ?)
                        """,
                        (job_id, new_run_id, user_id, url, title, now),
                    )
                    # Insert source row
                    ensure_source_row(
                        source_id=job_id,
                        run_id=new_run_id,
                        source_type="link",
                        status="PENDING",
                        url=url,
                        title=title,
                        crawl_job_id=job_id,
                        pdf_file_id=None,
                        meta_source_id=None,
                        cur=cur3,
                    )
                conn3.commit()
                conn3.close()
                
                log_message(f"Created {len(unique_links)} crawl jobs for extension to process", "SUCCESS", new_run_id)
            except Exception as e:
                conn4 = get_db()
                cur4 = conn4.cursor()
                cur4.execute("UPDATE runs SET status = 'failed' WHERE id = ?", (new_run_id,))
                conn4.commit()
                conn4.close()
                log_message(f"Deep Research retry failed: {str(e)}", "ERROR", new_run_id)

        threading.Thread(target=run_deep_research_retry, daemon=True).start()
        return jsonify(
            {
                "id": new_run_id,
                "name": name,
                "sourceType": "deep_research",
                "status": "searching",
                "startDate": now,
                "sourcesCount": 0,
                "dataEntriesCount": 0,
                "llmProvider": llm_provider,
                "prompt": prompt,
                "schemaFileId": schema_file_id,
                "deepResearchQuery": query,
                "validationEnabled": bool(validation_enabled),
                "validationMaxRetries": validation_max_retries,
            }
        ), 201

    return jsonify({"error": f"Unsupported source type for retry: {source_type}"}), 400


@app.route("/runs/<run_id>/stop", methods=["POST"])
def stop_run(run_id):
    """Stop a running extraction process."""
    # Try to terminate running process
    if run_id in active_processes:
        try:
            process = active_processes[run_id]
            process.terminate()
            process.wait(timeout=5)
        except:
            try:
                process.kill()
            except:
                pass
        finally:
            if run_id in active_processes:
                del active_processes[run_id]
        log_message(f"Run process terminated", "INFO", run_id)
    
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE runs SET status = 'completed' WHERE id = ?", (run_id,))
    conn.commit()
    conn.close()
    log_message(f"Run stopped", "INFO", run_id)
    return "", 204


@app.route("/runs/<run_id>/data", methods=["GET"])
def get_run_data(run_id):
    """Get extracted data (global_data.json) for a run.
    
    Returns the actual extracted JSON data. NO PATHS EXPOSED.
    Missing values are normalized to null for backward compatibility.
    """
    from missing_utils import normalize_data_list
    
    page = int(request.args.get("page", 1))
    page_size = int(request.args.get("pageSize", 100))
    sort = request.args.get("sort")

    if page < 1:
        page = 1
    if page_size < 1:
        page_size = 1
    if page_size > 500:
        page_size = 500

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT output_dir FROM runs WHERE id = ?", (run_id,))
    row = cur.fetchone()
    conn.close()
    
    if not row:
        return jsonify({"error": "Run not found"}), 404
    
    output_dir = row["output_dir"]
    if not output_dir:
        return jsonify({"error": "Run has no output directory configured"}), 400
    
    global_json_path = os.path.join(output_dir, "global_data.json")
    
    if not os.path.exists(global_json_path):
        return jsonify({
            "exists": False,
            "data": [],
            "count": 0,
            "page": page,
            "pageSize": page_size,
            "fields": []
        })
    
    try:
        with open(global_json_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        
        # Ensure data is a list
        if not isinstance(data, list):
            data = [data] if data else []
        
        # Normalize data - convert all missing indicators to null (backward compatibility)
        data = normalize_data_list(data)

        fields = set()
        for item in data:
            if isinstance(item, dict):
                for k in item.keys():
                    if k not in {"__source", "__url"}:
                        fields.add(k)

        total = len(data)
        items = data

        if sort:
            raw = str(sort).strip()
            if raw:
                if ":" in raw:
                    sort_field, sort_dir = raw.split(":", 1)
                else:
                    sort_field, sort_dir = raw, "desc"
                sort_field = sort_field.strip()
                sort_dir = sort_dir.strip().lower()
                reverse = sort_dir != "asc"

                def _sort_key(x):
                    if not isinstance(x, dict):
                        return ""
                    v = x.get(sort_field)
                    if v is None:
                        return ""
                    if isinstance(v, (int, float)):
                        return v
                    if isinstance(v, bool):
                        return int(v)
                    if isinstance(v, (dict, list)):
                        try:
                            return json.dumps(v, ensure_ascii=False, sort_keys=True)
                        except Exception:
                            return str(v)
                    return str(v).lower()

                try:
                    items = sorted(items, key=_sort_key, reverse=reverse)
                except Exception:
                    items = items

        start = (page - 1) * page_size
        end = start + page_size
        paged = items[start:end]

        return jsonify({
            "exists": True,
            "data": paged,
            "count": total,
            "page": page,
            "pageSize": page_size,
            "fields": sorted(list(fields))
        })
    except json.JSONDecodeError as e:
        return jsonify({"error": f"Invalid JSON in output file: {e}"}), 500
    except Exception as e:
        return jsonify({"error": f"Failed to read output file: {e}"}), 500


@app.route("/runs/<run_id>/inspection", methods=["GET"])
def get_run_inspection(run_id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT output_dir FROM runs WHERE id = ?", (run_id,))
    row = cur.fetchone()
    conn.close()

    if not row:
        return jsonify({"error": "Run not found"}), 404

    output_dir = row["output_dir"]
    if not output_dir:
        return jsonify({"error": "Run has no output directory configured"}), 400

    global_json_path = os.path.join(output_dir, "global_data.json")
    if not os.path.exists(global_json_path):
        return jsonify({
            "exists": False,
            "rows": 0,
            "fields": [],
            "overall": {"applicable": 0, "total": 0, "ratio": 0.0},
            "perField": {},
        })

    from missing_utils import is_missing
    def _is_applicable(v):
        return not is_missing(v)

    try:
        with open(global_json_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        if not isinstance(data, list):
            data = [data] if data else []

        fields_set = set()
        for item in data:
            if isinstance(item, dict):
                for k in item.keys():
                    if k not in {"__source", "__url"}:
                        fields_set.add(k)

        fields = sorted(list(fields_set))
        rows = len(data)
        total_cells = rows * len(fields)

        per_field = {}
        for field in fields:
            per_field[field] = {"applicable": 0, "total": rows, "ratio": 0.0}

        overall_applicable = 0
        if rows > 0 and fields:
            for item in data:
                row_obj = item if isinstance(item, dict) else {}
                for field in fields:
                    if _is_applicable(row_obj.get(field)):
                        overall_applicable += 1
                        per_field[field]["applicable"] += 1

            for field in fields:
                denom = per_field[field]["total"]
                per_field[field]["ratio"] = (per_field[field]["applicable"] / denom) if denom else 0.0

        overall_ratio = (overall_applicable / total_cells) if total_cells else 0.0
        return jsonify({
            "exists": True,
            "rows": rows,
            "fields": fields,
            "overall": {"applicable": overall_applicable, "total": total_cells, "ratio": overall_ratio},
            "perField": per_field,
        })
    except json.JSONDecodeError as e:
        return jsonify({"error": f"Invalid JSON in output file: {e}"}), 500
    except Exception as e:
        return jsonify({"error": f"Failed to read output file: {e}"}), 500


def _humanize_constraint(python_expr: str) -> str:
    """Convert pandas python_expression to human-readable constraint."""
    if not python_expr:
        return ""
    
    # Replace common pandas patterns with readable text
    result = python_expr
    
    # Replace df['column'] with just column name
    result = re.sub(r"df\['([^']+)'\]", r"\1", result)
    result = re.sub(r'df\["([^"]+)"\]', r"\1", result)
    
    # Replace pd.to_numeric(..., errors='coerce') with "numeric value of"
    result = re.sub(r"pd\.to_numeric\(([^,]+),\s*errors='coerce'\)", r"(numeric) \1", result)
    
    # Replace .between(a, b) with "between a and b"
    result = re.sub(r"\.between\(([^,]+),\s*([^)]+)\)", r" is between \1 and \2", result)
    
    # Replace .fillna(0) with cleaner text
    result = re.sub(r"\.fillna\(0\)", "", result)
    
    # Replace .notna() with "is not empty"
    result = re.sub(r"\.notna\(\)", " is not empty", result)
    
    # Replace .isna() with "is empty"
    result = re.sub(r"\.isna\(\)", " is empty", result)
    
    # Replace .str.contains(...) with "contains"
    result = re.sub(r"\.str\.contains\('([^']+)'[^)]*\)", r" contains '\1'", result)
    
    # Replace >= 0 with ">= 0"
    result = re.sub(r"\s*>=\s*0", " >= 0", result)
    
    # Replace > 0 with "> 0"  
    result = re.sub(r"\s*>\s*0", " > 0", result)
    
    # Replace != '' with "is not empty"
    result = re.sub(r"\s*!=\s*''", " is not empty", result)
    
    # Replace & with "AND"
    result = re.sub(r"\s*&\s*", " AND ", result)
    
    # Replace | with "OR"
    result = re.sub(r"\s*\|\s*", " OR ", result)
    
    # Clean up extra whitespace
    result = re.sub(r"\s+", " ", result).strip()
    
    return result


@app.route("/runs/<run_id>/validation", methods=["GET"])
def get_run_validation(run_id):
    """Get validation results for a run.
    
    Returns validation report, rules summary, row flags, and validation prompt.
    """
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT output_dir, validation_enabled, validation_pass_rate, validation_accepted_count, validation_rejected_count, validation_prompt_file_id FROM runs WHERE id = ?", (run_id,))
    row = cur.fetchone()
    conn.close()

    if not row:
        return jsonify({"error": "Run not found"}), 404

    output_dir = row["output_dir"]
    if not output_dir:
        return jsonify({"exists": False, "message": "Run has no output directory"})

    # Check if validation was enabled
    if not row["validation_enabled"]:
        return jsonify({"exists": False, "message": "Validation was not enabled for this run"})

    validation_dir = os.path.join(output_dir, "validation")
    validation_report_path = os.path.join(validation_dir, "validation_report.json")
    validation_config_path = os.path.join(output_dir, "validation_config.json")
    row_flags_path = os.path.join(validation_dir, "row_flags.csv")

    if not os.path.exists(validation_report_path):
        # Still return validation prompt if it exists, even without results
        validation_prompt = None
        validation_prompt_file_id = row["validation_prompt_file_id"]
        if validation_prompt_file_id:
            prompt_path = get_file_internal_path(validation_prompt_file_id)
            if prompt_path and os.path.isfile(prompt_path):
                try:
                    with open(prompt_path, "r", encoding="utf-8") as f:
                        validation_prompt = f.read()
                except Exception:
                    pass
        
        return jsonify({
            "exists": False,
            "message": "Validation report not found. Run may still be in progress.",
            "validationPrompt": validation_prompt,
            "validationEnabled": True
        })

    try:
        # Load validation report
        with open(validation_report_path, "r", encoding="utf-8") as f:
            report = json.load(f)

        # Load generated config if exists
        generated_config = None
        if os.path.exists(validation_config_path):
            with open(validation_config_path, "r", encoding="utf-8") as f:
                generated_config = json.load(f)

        # Load row flags if exists
        row_flags = []
        if os.path.exists(row_flags_path):
            import csv
            with open(row_flags_path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                row_flags = list(reader)

        # Build rules summary with human-readable constraints
        rules_summary = []
        
        # Create a lookup from generated config for rule details
        config_rules_lookup = {}
        if generated_config and "rules" in generated_config:
            for rule in generated_config["rules"]:
                config_rules_lookup[rule.get("rule_id")] = rule
        
        for result in report.get("validation_results", []):
            rule_id = result.get("rule_id")
            config_rule = config_rules_lookup.get(rule_id, {})
            
            # Convert python_expression to human-readable constraint
            python_expr = config_rule.get("python_expression", "")
            constraint = _humanize_constraint(python_expr) if python_expr else None
            
            severity = result.get("severity") or config_rule.get("severity", "warning")
            raw_passed = result.get("passed", False)
            # Only errors count as true failures; warnings always "pass" for acceptance purposes
            effective_passed = raw_passed if severity == "error" else True
            
            rules_summary.append({
                "ruleId": rule_id,
                "name": config_rule.get("name") or (result.get("message", "").split(":")[0] if result.get("message") else ""),
                "description": config_rule.get("description"),
                "columns": config_rule.get("columns", []),
                "constraint": constraint,
                "passed": effective_passed,  # Only errors can fail
                "rawPassed": raw_passed,  # Original pass/fail for display
                "severity": severity,
                "details": result.get("details", {}),
                "affectedRows": result.get("affected_rows", [])
            })

        # Load validation prompt if exists
        validation_prompt = None
        validation_prompt_file_id = row["validation_prompt_file_id"]
        if validation_prompt_file_id:
            prompt_path = get_file_internal_path(validation_prompt_file_id)
            if prompt_path and os.path.isfile(prompt_path):
                try:
                    with open(prompt_path, "r", encoding="utf-8") as f:
                        validation_prompt = f.read()
                except Exception:
                    pass

        return jsonify({
            "exists": True,
            "summary": {
                "overallPassRate": row["validation_pass_rate"] or report.get("summary", {}).get("overall_pass_rate", 0),
                "totalRows": report.get("total_rows", 0),
                "acceptedRows": row["validation_accepted_count"],
                "rejectedRows": row["validation_rejected_count"],
                "totalRules": report.get("summary", {}).get("total_rules", 0),
                "enabledRules": report.get("summary", {}).get("enabled_rules", 0)
            },
            "rules": rules_summary,
            "generatedConfig": generated_config,
            "rowFlags": row_flags[:100],  # Limit to first 100 rows for performance
            "rowFlagsTotal": len(row_flags),
            "validationPrompt": validation_prompt
        })
    except json.JSONDecodeError as e:
        return jsonify({"error": f"Invalid JSON in validation report: {e}"}), 500
    except Exception as e:
        return jsonify({"error": f"Failed to read validation results: {e}"}), 500


@app.route("/runs/<run_id>/validation/report", methods=["GET"])
def download_validation_report(run_id):
    """Download validation report as PDF or Excel file.
    
    Query params:
        format: 'pdf' (default) or 'excel'
    """
    from openpyxl import Workbook
    import io
    import csv
    import tempfile
    
    report_format = request.args.get("format", "pdf").lower()
    
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM runs WHERE id = ?", (run_id,))
    row = cur.fetchone()
    conn.close()

    if not row:
        return jsonify({"error": "Run not found"}), 404

    run_data = dict(row)
    output_dir = run_data.get("output_dir")
    run_name = run_data.get("name") or run_id[:8]
    
    if not output_dir or not run_data.get("validation_enabled"):
        return jsonify({"error": "Validation not enabled for this run"}), 400

    validation_dir = os.path.join(output_dir, "validation")
    validation_report_path = os.path.join(validation_dir, "validation_report.json")
    validation_config_path = os.path.join(output_dir, "validation_config.json")
    row_flags_path = os.path.join(validation_dir, "row_flags.csv")
    validated_data_path = os.path.join(validation_dir, "validated_data.csv")
    global_data_path = os.path.join(output_dir, "global_data.json")
    schema_mapping_path = os.path.join(output_dir, "schema_mapping.json")

    if not os.path.exists(validation_report_path):
        return jsonify({"error": "Validation report not found"}), 404

    try:
        # Load validation data
        with open(validation_report_path, "r", encoding="utf-8") as f:
            report = json.load(f)
        
        generated_config = None
        if os.path.exists(validation_config_path):
            with open(validation_config_path, "r", encoding="utf-8") as f:
                generated_config = json.load(f)
        
        # Handle PDF format
        if report_format == "pdf":
            from report_generator import generate_validation_report_pdf
            
            # Load extracted data
            extracted_data = []
            if os.path.exists(global_data_path):
                with open(global_data_path, "r", encoding="utf-8") as f:
                    extracted_data = json.load(f)
                    if not isinstance(extracted_data, list):
                        extracted_data = [extracted_data] if extracted_data else []
            
            # Load schema fields
            schema_fields = None
            if os.path.exists(schema_mapping_path):
                with open(schema_mapping_path, "r", encoding="utf-8") as f:
                    schema_mapping = json.load(f)
                    if 'fields' in schema_mapping:
                        schema_fields = schema_mapping['fields']
                    elif 'fieldDefs' in schema_mapping:
                        schema_fields = [f.get('name') for f in schema_mapping['fieldDefs'] if f.get('name')]
            
            # Generate PDF
            safe_name = "".join(c if c.isalnum() or c in "-_" else "_" for c in run_name)
            pdf_path = os.path.join(tempfile.gettempdir(), f"validation_report_{safe_name}.pdf")
            
            generate_validation_report_pdf(
                run_data=run_data,
                extracted_data=extracted_data,
                validation_report=report,
                validation_config=generated_config,
                output_path=pdf_path,
                schema_fields=schema_fields
            )
            
            return send_file(
                pdf_path,
                mimetype="application/pdf",
                as_attachment=True,
                download_name=f"validation_report_{safe_name}.pdf"
            )
        
        # Handle Excel format (existing logic)
        
        # Create Excel workbook
        wb = Workbook()
        
        # Calculate passed/failed from validation_results
        validation_results = report.get("validation_results", [])
        total_rules = len(validation_results)
        passed_rules = sum(1 for r in validation_results if r.get("passed", False))
        failed_rules = total_rules - passed_rules
        # Only count errors as true failures
        error_failures = sum(1 for r in validation_results if r.get("severity") == "error" and not r.get("passed", False))
        warning_failures = sum(1 for r in validation_results if r.get("severity") == "warning" and not r.get("passed", False))
        
        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"
        summary_data = report.get("summary", {})
        ws_summary.append(["Validation Report Summary"])
        ws_summary.append([])
        ws_summary.append(["Metric", "Value"])
        ws_summary.append(["Total Rows", report.get("total_rows", 0)])
        ws_summary.append(["Overall Pass Rate", f"{summary_data.get('overall_pass_rate', 0) * 100:.1f}%"])
        ws_summary.append(["Total Rules", total_rules])
        ws_summary.append(["Passed Rules", passed_rules])
        ws_summary.append(["Failed Rules (Errors)", error_failures])
        ws_summary.append(["Warnings", warning_failures])
        
        # Sheet 2: Rules by Severity (Errors vs Warnings)
        ws_rules = wb.create_sheet("Rules Summary")
        ws_rules.append(["Rule ID", "Name", "Severity", "Status", "Affected Rows", "Description", "Columns", "Constraint"])
        
        config_rules_lookup = {}
        if generated_config and "rules" in generated_config:
            for rule in generated_config["rules"]:
                config_rules_lookup[rule.get("rule_id")] = rule
        
        for result in report.get("validation_results", []):
            rule_id = result.get("rule_id")
            config_rule = config_rules_lookup.get(rule_id, {})
            severity = result.get("severity", "warning")
            raw_passed = result.get("passed", False)
            # Only errors count as FAIL
            status = "PASS" if raw_passed else ("FAIL" if severity == "error" else "WARN")
            affected_count = len(result.get("affected_rows", []))
            
            ws_rules.append([
                rule_id,
                config_rule.get("name", ""),
                severity.upper(),
                status,
                affected_count,
                config_rule.get("description", ""),
                ", ".join(config_rule.get("columns", [])),
                _humanize_constraint(config_rule.get("python_expression", ""))
            ])
        
        # Sheet 3: Constraints by Column
        ws_columns = wb.create_sheet("By Column")
        ws_columns.append(["Column", "Total Rules", "Errors Failed", "Warnings", "Status", "Rule IDs"])
        
        column_to_rules = {}
        for result in report.get("validation_results", []):
            rule_id = result.get("rule_id")
            config_rule = config_rules_lookup.get(rule_id, {})
            for col in config_rule.get("columns", []):
                if col not in column_to_rules:
                    column_to_rules[col] = []
                column_to_rules[col].append({
                    "rule_id": rule_id,
                    "severity": result.get("severity", "warning"),
                    "passed": result.get("passed", False)
                })
        
        for col in sorted(column_to_rules.keys()):
            rules = column_to_rules[col]
            error_fails = sum(1 for r in rules if r["severity"] == "error" and not r["passed"])
            warnings = sum(1 for r in rules if r["severity"] == "warning" and not r["passed"])
            status = "FAIL" if error_fails > 0 else ("WARN" if warnings > 0 else "PASS")
            rule_ids = ", ".join(r["rule_id"] for r in rules)
            ws_columns.append([col, len(rules), error_fails, warnings, status, rule_ids])
        
        # Sheet 4: Row Flags with Success Rate
        if os.path.exists(row_flags_path):
            ws_rows = wb.create_sheet("Row Flags")
            import csv
            with open(row_flags_path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                rows_data = list(reader)
            
            if rows_data:
                # Get flag columns (boolean-like)
                flag_cols = [k for k in rows_data[0].keys() if k != "row_index"]
                headers = ["Row", "Passed", "Failed", "Success Rate"] + flag_cols
                ws_rows.append(headers)
                
                for row_data in rows_data:
                    row_idx = row_data.get("row_index", "")
                    passed = sum(1 for k in flag_cols if str(row_data.get(k, "")).lower() in ("true", "1"))
                    failed = sum(1 for k in flag_cols if str(row_data.get(k, "")).lower() in ("false", "0"))
                    total = passed + failed
                    success_rate = f"{(passed / total * 100):.1f}%" if total > 0 else "N/A"
                    
                    row_values = [row_idx, passed, failed, success_rate]
                    for k in flag_cols:
                        val = row_data.get(k, "")
                        if str(val).lower() == "true":
                            row_values.append("✓")
                        elif str(val).lower() == "false":
                            row_values.append("✗")
                        else:
                            row_values.append(val)
                    ws_rows.append(row_values)
        
        # Sheet 5: Validated Data
        if os.path.exists(validated_data_path):
            ws_validated = wb.create_sheet("Validated Data")
            import csv
            with open(validated_data_path, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                for row_data in reader:
                    ws_validated.append(row_data)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        safe_name = "".join(c if c.isalnum() or c in "-_" else "_" for c in run_name)
        filename = f"validation_report_{safe_name}.xlsx"
        
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({"error": f"Failed to generate validation report: {e}"}), 500


@app.route("/runs/<run_id>/validation/upload", methods=["POST"])
@require_auth
def upload_validation_prompt(run_id):
    """Upload a validation prompt for an existing run (post-extraction).
    
    This allows enabling validation after extraction has completed.
    """
    user_id = g.current_user["id"]
    
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id, user_id, output_dir, status FROM runs WHERE id = ?", (run_id,))
    row = cur.fetchone()
    
    if not row:
        conn.close()
        return jsonify({"error": "Run not found"}), 404
    
    if row["user_id"] != user_id:
        conn.close()
        return jsonify({"error": "Forbidden"}), 403
    
    output_dir = row["output_dir"]
    
    # Check if validation prompt file is provided
    if "validationPrompt" not in request.files:
        conn.close()
        return jsonify({"error": "No validation prompt file provided"}), 400
    
    val_prompt_file = request.files["validationPrompt"]
    if not val_prompt_file.filename or not val_prompt_file.filename.lower().endswith('.txt'):
        conn.close()
        return jsonify({"error": "Validation prompt must be a .txt file"}), 400
    
    # Save validation prompt file
    run_upload_dir = os.path.join(UPLOAD_FOLDER, run_id)
    os.makedirs(run_upload_dir, exist_ok=True)
    
    val_prompt_filename = secure_filename(val_prompt_file.filename)
    validation_prompt_path = os.path.join(run_upload_dir, f"validation_prompt_{val_prompt_filename}")
    val_prompt_file.save(validation_prompt_path)
    
    # Register file
    validation_prompt_file_id = register_file(validation_prompt_path, val_prompt_file.filename, "validation_prompt", run_id, "text/plain")
    
    # Get optional max retries from form data
    validation_max_retries = int(request.form.get("validationMaxRetries", 3))
    
    # Update run with validation settings
    cur.execute("""
        UPDATE runs 
        SET validation_prompt_file_id = ?, validation_enabled = 1, validation_max_retries = ?
        WHERE id = ?
    """, (validation_prompt_file_id, validation_max_retries, run_id))
    conn.commit()
    conn.close()
    
    log_message(f"Validation prompt uploaded: {val_prompt_file.filename}", "INFO", run_id)
    
    return jsonify({
        "success": True,
        "message": "Validation prompt uploaded successfully",
        "validationPromptFileId": validation_prompt_file_id,
        "validationEnabled": True,
        "validationMaxRetries": validation_max_retries
    })


@app.route("/runs/<run_id>/validation/run", methods=["POST"])
@require_auth
def run_validation_only(run_id):
    """Run validation on extracted data without re-extracting.
    
    This spawns extract.py with --validation-only flag, which:
    - Skips extraction phase
    - Loads existing global_data.json
    - Generates validation config from the uploaded prompt
    - Runs validation and saves results
    - Appends logs to existing IPC logs (visible in Logs tab)
    """
    user_id = g.current_user["id"]
    
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, user_id, output_dir, pdfs_dir, status, validation_prompt_file_id, 
               validation_enabled, validation_max_retries, schema_file_id,
               cache_surya_read, cache_surya_write, cache_llm_read, cache_llm_write,
               cache_schema_read, cache_schema_write, cache_validation_read, cache_validation_write
        FROM runs WHERE id = ?
    """, (run_id,))
    row = cur.fetchone()
    
    if not row:
        conn.close()
        return jsonify({"error": "Run not found"}), 404
    
    if row["user_id"] != user_id:
        conn.close()
        return jsonify({"error": "Forbidden"}), 403
    
    output_dir = row["output_dir"]
    pdfs_dir = row["pdfs_dir"]
    validation_prompt_file_id = row["validation_prompt_file_id"]
    schema_file_id = row["schema_file_id"]
    
    # Check if validation prompt exists
    if not validation_prompt_file_id:
        conn.close()
        return jsonify({"error": "No validation prompt uploaded. Upload one first via /validation/upload"}), 400
    
    # Get validation prompt path
    validation_prompt_path = get_file_internal_path(validation_prompt_file_id)
    if not validation_prompt_path or not os.path.isfile(validation_prompt_path):
        conn.close()
        return jsonify({"error": "Validation prompt file not found"}), 400
    
    # Get schema file path
    excel_path = get_file_internal_path(schema_file_id) if schema_file_id else None
    if not excel_path or not os.path.isfile(excel_path):
        conn.close()
        return jsonify({"error": "Schema file not found"}), 400
    
    # Check if extracted data exists
    global_json_path = os.path.join(output_dir, "global_data.json")
    if not os.path.isfile(global_json_path):
        conn.close()
        return jsonify({"error": "No extracted data found. Run extraction first."}), 400
    
    # Build cache flags from run settings
    cache_flags = {
        "surya_read": row["cache_surya_read"] if row["cache_surya_read"] is not None else True,
        "surya_write": row["cache_surya_write"] if row["cache_surya_write"] is not None else True,
        "llm_read": row["cache_llm_read"] if row["cache_llm_read"] is not None else True,
        "llm_write": row["cache_llm_write"] if row["cache_llm_write"] is not None else True,
        "schema_read": row["cache_schema_read"] if row["cache_schema_read"] is not None else True,
        "schema_write": row["cache_schema_write"] if row["cache_schema_write"] is not None else True,
        "validation_read": row["cache_validation_read"] if row["cache_validation_read"] is not None else True,
        "validation_write": row["cache_validation_write"] if row["cache_validation_write"] is not None else True,
    }
    
    conn.close()
    
    # Spawn validation-only process (uses extract.py with --validation-only flag)
    spawn_validation_only_process(
        run_id=run_id,
        pdfs_dir=pdfs_dir,
        excel_path=excel_path,
        output_dir=output_dir,
        validation_prompt_path=validation_prompt_path,
        user_id=user_id,
        cache_flags=cache_flags
    )
    
    log_message("Validation-only process started", "INFO", run_id)
    
    return jsonify({
        "success": True,
        "message": "Validation started. Check the Logs tab for progress and Validation tab for results.",
        "runId": run_id
    })


@app.route("/runs/<run_id>/validation/rerun", methods=["POST"])
@require_auth
def rerun_validation_logic(run_id):
    """Re-run validation logic using existing validation_config.json.
    
    This does NOT regenerate the config from LLM - it only re-applies
    the existing validation rules to the extracted data.
    Useful after fixing validation engine bugs or tweaking config manually.
    """
    user_id = g.current_user["id"]
    
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id, user_id, output_dir, validation_enabled FROM runs WHERE id = ?", (run_id,))
    row = cur.fetchone()
    
    if not row:
        conn.close()
        return jsonify({"error": "Run not found"}), 404
    
    if row["user_id"] != user_id:
        conn.close()
        return jsonify({"error": "Forbidden"}), 403
    
    output_dir = row["output_dir"]
    conn.close()
    
    if not output_dir:
        return jsonify({"error": "Run has no output directory"}), 400
    
    # Check required files exist
    global_json_path = os.path.join(output_dir, "global_data.json")
    validation_config_path = os.path.join(output_dir, "validation_config.json")
    
    if not os.path.isfile(global_json_path):
        return jsonify({"error": "No extracted data found (global_data.json missing)"}), 400
    
    if not os.path.isfile(validation_config_path):
        return jsonify({"error": "No validation config found. Run validation first to generate config."}), 400
    
    try:
        # Load extracted data
        with open(global_json_path, "r", encoding="utf-8") as f:
            extracted_data = json.load(f)
        
        if not isinstance(extracted_data, list):
            extracted_data = [extracted_data] if extracted_data else []
        
        # Load validation config
        with open(validation_config_path, "r", encoding="utf-8") as f:
            validation_config = json.load(f)
        
        # Import validation modules
        import pandas as pd
        from validation.rule_types import ValidationConfig, RuleDefinition, RuleScope, RuleSeverity
        from validation.rule_engine import RuleEngine
        
        # Convert config to ValidationConfig object
        rules = []
        for rule_dict in validation_config.get("rules", []):
            rules.append(RuleDefinition(
                rule_id=rule_dict.get("rule_id", ""),
                name=rule_dict.get("name", ""),
                description=rule_dict.get("description", ""),
                scope=RuleScope(rule_dict.get("scope", "row")),
                severity=RuleSeverity(rule_dict.get("severity", "warning")),
                columns=rule_dict.get("columns", []),
                python_expression=rule_dict.get("python_expression"),
                enabled=rule_dict.get("enabled", True),
                filter_condition=rule_dict.get("filter_condition"),
                flag_column=rule_dict.get("flag_column")
            ))
        
        config = ValidationConfig(
            name=validation_config.get("name", "Validation"),
            description=validation_config.get("description", ""),
            rules=rules,
            filter_condition=validation_config.get("filter_condition"),
            paper_group_column=validation_config.get("paper_group_column")
        )
        
        # Create DataFrame and run validation
        df = pd.DataFrame(extracted_data)
        engine = RuleEngine(config)
        report = engine.validate(df)
        
        # Save validation results
        validation_dir = os.path.join(output_dir, "validation")
        os.makedirs(validation_dir, exist_ok=True)
        
        # Save report
        report_dict = report.to_dict()
        with open(os.path.join(validation_dir, "validation_report.json"), "w", encoding="utf-8") as f:
            json.dump(report_dict, f, indent=2, default=str)
        
        # Save row flags as CSV
        if report.row_results:
            row_flags_df = pd.DataFrame(report.row_results)
            row_flags_df.insert(0, "row_index", range(len(row_flags_df)))
            row_flags_df.to_csv(os.path.join(validation_dir, "row_flags.csv"), index=False)
        
        # Save validated data (rows that pass all error-severity rules)
        error_rules = [r for r in report.all_results if r.severity == RuleSeverity.ERROR]
        if error_rules:
            failed_rows = set()
            for r in error_rules:
                if not r.passed:
                    failed_rows.update(r.affected_rows)
            accepted_data = [row for i, row in enumerate(extracted_data) if i not in failed_rows]
        else:
            accepted_data = extracted_data
        
        with open(os.path.join(validation_dir, "validated_data.json"), "w", encoding="utf-8") as f:
            json.dump(accepted_data, f, indent=2, default=str)
        
        validated_df = pd.DataFrame(accepted_data)
        validated_df.to_csv(os.path.join(validation_dir, "validated_data.csv"), index=False)
        
        # Generate summary
        summary_lines = [
            f"Validation Re-run Complete",
            f"Total rows: {len(extracted_data)}",
            f"Accepted rows: {len(accepted_data)}",
            f"Rejected rows: {len(extracted_data) - len(accepted_data)}",
            f"Pass rate: {report.summary.get('overall_pass_rate', 0):.1%}"
        ]
        with open(os.path.join(validation_dir, "validation_summary.txt"), "w", encoding="utf-8") as f:
            f.write("\n".join(summary_lines))
        
        log_message(f"Validation re-run complete: {len(accepted_data)}/{len(extracted_data)} rows accepted", "INFO", run_id)
        
        return jsonify({
            "success": True,
            "message": f"Validation re-run complete. {len(accepted_data)}/{len(extracted_data)} rows accepted.",
            "totalRows": len(extracted_data),
            "acceptedRows": len(accepted_data),
            "passRate": report.summary.get("overall_pass_rate", 0)
        })
        
    except Exception as e:
        log_message(f"Validation re-run failed: {e}", "ERROR", run_id)
        return jsonify({"error": f"Validation re-run failed: {e}"}), 500


@app.route("/runs/<run_id>/validated-data", methods=["GET"])
def get_run_validated_data(run_id):
    """Get validated/filtered data (validated_data.json) for a run.
    
    Returns only the rows that passed validation (accepted rows).
    Missing values are normalized to null for backward compatibility.
    """
    from missing_utils import normalize_data_list
    
    page = int(request.args.get("page", 1))
    page_size = int(request.args.get("pageSize", 100))

    if page < 1:
        page = 1
    if page_size < 1:
        page_size = 1
    if page_size > 500:
        page_size = 500

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT output_dir, validation_enabled FROM runs WHERE id = ?", (run_id,))
    row = cur.fetchone()
    conn.close()

    if not row:
        return jsonify({"error": "Run not found"}), 404

    output_dir = row["output_dir"]
    if not output_dir:
        return jsonify({"exists": False, "message": "Run has no output directory"})

    if not row["validation_enabled"]:
        return jsonify({"exists": False, "message": "Validation was not enabled for this run"})

    validated_json_path = os.path.join(output_dir, "validated_data.json")

    if not os.path.exists(validated_json_path):
        return jsonify({
            "exists": False,
            "message": "Validated data not found. Run may still be in progress.",
            "data": [],
            "count": 0
        })

    try:
        with open(validated_json_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        if not isinstance(data, list):
            data = [data] if data else []
        
        # Normalize data - convert all missing indicators to null (backward compatibility)
        data = normalize_data_list(data)

        fields = set()
        for item in data:
            if isinstance(item, dict):
                for k in item.keys():
                    if k not in {"__source", "__url", "row_accept_candidate"}:
                        fields.add(k)

        total = len(data)
        start = (page - 1) * page_size
        end = start + page_size
        paged = data[start:end]

        return jsonify({
            "exists": True,
            "data": paged,
            "count": total,
            "page": page,
            "pageSize": page_size,
            "fields": sorted(list(fields))
        })
    except json.JSONDecodeError as e:
        return jsonify({"error": f"Invalid JSON in validated data: {e}"}), 500
    except Exception as e:
        return jsonify({"error": f"Failed to read validated data: {e}"}), 500


@app.route("/runs/<run_id>/schema-mapping", methods=["GET"])
def get_run_schema_mapping(run_id):
    """Get schema mapping (schema_mapping.json) for a run.

    Returns the persisted schema mapping JSON. NO PATHS EXPOSED.
    """
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT output_dir FROM runs WHERE id = ?", (run_id,))
    row = cur.fetchone()
    conn.close()

    if not row:
        return jsonify({"error": "Run not found"}), 404

    output_dir = row["output_dir"]
    if not output_dir:
        return jsonify({"error": "Run has no output directory configured"}), 400

    schema_mapping_path = os.path.join(output_dir, "schema_mapping.json")
    if not os.path.exists(schema_mapping_path):
        return jsonify({"exists": False, "mapping": None})

    try:
        with open(schema_mapping_path, "r", encoding="utf-8") as f:
            mapping = json.load(f)
        return jsonify({"exists": True, "mapping": mapping})
    except json.JSONDecodeError as e:
        return jsonify({"error": f"Invalid JSON in schema mapping file: {e}"}), 500
    except Exception as e:
        return jsonify({"error": f"Failed to read schema mapping file: {e}"}), 500


@app.route("/runs/<run_id>/data/download", methods=["GET"])
def download_run_data(run_id):
    """Download extracted data in various formats: json, csv, excel.
    
    Query params:
        format: json | csv | excel (default: json)
    
    Note: Missing values are normalized:
        - JSON: null
        - CSV/Excel: N/A
    """
    import csv
    import io
    from missing_utils import normalize_data_list, is_missing
    
    format_type = request.args.get("format", "json").lower()
    
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT output_dir, name FROM runs WHERE id = ?", (run_id,))
    row = cur.fetchone()
    conn.close()
    
    if not row:
        return jsonify({"error": "Run not found"}), 404
    
    output_dir = row["output_dir"]
    run_name = row["name"] or run_id
    
    if not output_dir:
        return jsonify({"error": "Run has no output directory configured"}), 400
    
    global_json_path = os.path.join(output_dir, "global_data.json")
    
    if not os.path.exists(global_json_path):
        return jsonify({"error": "No extracted data available"}), 404
    
    try:
        with open(global_json_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        
        if not isinstance(data, list):
            data = [data] if data else []
        
        if not data:
            return jsonify({"error": "Extracted data is empty"}), 404
        
        # Normalize data - convert all missing indicators to null (backward compatibility)
        data = normalize_data_list(data)
        
        # Generate filename
        safe_name = "".join(c if c.isalnum() or c in "-_" else "_" for c in run_name)
        
        if format_type == "json":
            response = app.response_class(
                response=json.dumps(data, indent=2, ensure_ascii=False),
                status=200,
                mimetype="application/json"
            )
            response.headers["Content-Disposition"] = f"attachment; filename={safe_name}_data.json"
            return response
        
        elif format_type == "csv":
            # Get all unique keys from all entries
            all_keys = set()
            for entry in data:
                all_keys.update(entry.keys())
            all_keys = sorted(all_keys)
            
            # Convert null values to 'N/A' for CSV export
            def format_csv_value(v):
                if v is None:
                    return "N/A"
                return v
            
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=all_keys, extrasaction='ignore')
            writer.writeheader()
            for entry in data:
                formatted_entry = {k: format_csv_value(v) for k, v in entry.items()}
                writer.writerow(formatted_entry)
            
            response = app.response_class(
                response=output.getvalue(),
                status=200,
                mimetype="text/csv"
            )
            response.headers["Content-Disposition"] = f"attachment; filename={safe_name}_data.csv"
            return response
        
        elif format_type == "excel":
            try:
                import pandas as pd
                
                df = pd.DataFrame(data)
                # Replace None/NaN with 'N/A' for Excel export
                df = df.fillna("N/A")
                excel_buffer = io.BytesIO()
                df.to_excel(excel_buffer, index=False, engine='openpyxl')
                excel_buffer.seek(0)
                
                response = app.response_class(
                    response=excel_buffer.getvalue(),
                    status=200,
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                response.headers["Content-Disposition"] = f"attachment; filename={safe_name}_data.xlsx"
                return response
            except ImportError:
                return jsonify({"error": "Excel export requires pandas and openpyxl packages"}), 500
        
        else:
            return jsonify({"error": f"Unsupported format: {format_type}. Use json, csv, or excel"}), 400
            
    except json.JSONDecodeError as e:
        return jsonify({"error": f"Invalid JSON in output file: {e}"}), 500
    except Exception as e:
        return jsonify({"error": f"Failed to generate download: {e}"}), 500


@app.route("/runs/<run_id>/export", methods=["POST"])
def export_run(run_id):
    """Export run results as JSON."""
    now = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"export_{run_id[:8]}_{now}.json"
    filepath = os.path.join(EXPORTS_FOLDER, filename)
    
    # Create export file (placeholder)
    with open(filepath, "w") as f:
        json.dump({"runId": run_id, "exportedAt": datetime.now(timezone.utc).isoformat()}, f)
    
    # Record in DB
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO exports (run_id, created_at, file_path) VALUES (?, ?, ?)",
        (run_id, datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"), filepath)
    )
    conn.commit()
    conn.close()
    
    log_message(f"Run exported: {filename}", "INFO", run_id)
    return jsonify({"url": f"/exports/{cur.lastrowid}/download"})


@app.route("/runs/<run_id>/export-pdf", methods=["POST"])
def export_run_pdf(run_id):
    """Export comprehensive PDF report for a run.
    
    Includes:
    - Run metadata and configuration
    - Extraction summary
    - Validation results (if enabled)
    - Extracted data tables
    - Validated data tables (if enabled)
    """
    try:
        from report_generator import generate_report_from_run_dir
    except ImportError as e:
        return jsonify({"error": f"Report generator not available: {e}. Install reportlab: pip install reportlab"}), 500
    
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, name, status, source_type, sources_count, data_entries_count,
               llm_provider, start_date, output_dir, validation_enabled
        FROM runs WHERE id = ?
    """, (run_id,))
    row = cur.fetchone()
    conn.close()
    
    if not row:
        return jsonify({"error": "Run not found"}), 404
    
    run_data = dict(row)
    output_dir = run_data.get('output_dir')
    
    if not output_dir or not os.path.exists(output_dir):
        return jsonify({"error": "Run output directory not found. Run may not have completed."}), 400
    
    # Generate PDF
    now = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    safe_name = "".join(c for c in run_data.get('name', 'run')[:30] if c.isalnum() or c in (' ', '-', '_')).strip()
    safe_name = safe_name.replace(' ', '_')
    filename = f"report_{safe_name}_{run_id[:8]}_{now}.pdf"
    filepath = os.path.join(EXPORTS_FOLDER, filename)
    
    try:
        generate_report_from_run_dir(
            run_id=run_id,
            run_data=run_data,
            output_dir=output_dir,
            report_output_path=filepath
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Failed to generate PDF report: {e}"}), 500
    
    # Record in DB
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO exports (run_id, created_at, file_path) VALUES (?, ?, ?)",
        (run_id, datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"), filepath)
    )
    export_id = cur.lastrowid
    conn.commit()
    conn.close()
    
    log_message(f"PDF report generated: {filename}", "INFO", run_id)
    return jsonify({
        "url": f"/exports/{export_id}/download",
        "filename": filename,
        "exportId": export_id
    })


@app.route("/runs/<run_id>/export-zip", methods=["POST"])
@optional_auth
def export_run_zip(run_id):
    """Export COMPLETE run package as ZIP file with ALL related data.
    
    Includes EVERYTHING:
    - All files from run output directory (data files, schemas, validation, etc.)
    - All files from run uploads directory (PDFs, schema Excel, prompts, etc.)
    - All IPC logs (extraction.log, stdout.log, stderr.log)
    - Excel workbook with ALL database records (run, sources, logs, files, crawl_jobs, etc.)
    - Generated PDF report
    - JSON dumps of all database records
    """
    import zipfile
    import io
    
    try:
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500
    
    # Get FULL run data (all columns)
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM runs WHERE id = ?", (run_id,))
    row = cur.fetchone()
    
    if not row:
        conn.close()
        return jsonify({"error": "Run not found"}), 404
    
    run_data = dict(row)
    output_dir = run_data.get('output_dir')
    run_name = run_data.get('name', 'run').replace(' ', '_')[:30]
    pdfs_dir = run_data.get('pdfs_dir')
    
    if not output_dir or not os.path.isdir(output_dir):
        conn.close()
        return jsonify({"error": "Run output directory not found"}), 404
    
    # Get ALL related objects from DB
    # Sources (all columns)
    cur.execute("SELECT * FROM sources WHERE run_id = ?", (run_id,))
    sources = [dict(r) for r in cur.fetchall()]
    
    # Logs (ALL logs, not limited)
    cur.execute("SELECT * FROM logs WHERE run_id = ? ORDER BY created_at", (run_id,))
    logs = [dict(r) for r in cur.fetchall()]
    
    # Meta sources
    cur.execute("SELECT * FROM meta_sources WHERE run_id = ?", (run_id,))
    meta_sources = [dict(r) for r in cur.fetchall()]
    
    # Exports
    cur.execute("SELECT * FROM exports WHERE run_id = ?", (run_id,))
    exports = [dict(r) for r in cur.fetchall()]
    
    # Files (all registered files for this run)
    cur.execute("SELECT * FROM files WHERE run_id = ?", (run_id,))
    files = [dict(r) for r in cur.fetchall()]
    
    # Crawl jobs
    cur.execute("SELECT * FROM crawl_jobs WHERE run_id = ?", (run_id,))
    crawl_jobs = [dict(r) for r in cur.fetchall()]
    
    conn.close()
    
    # Paths for additional directories
    run_upload_dir = os.path.join(UPLOAD_FOLDER, run_id)
    run_ipc_dir = os.path.join(IPC_DIR, run_id)
    
    # Create ZIP file in memory
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        # 1. Add all files from OUTPUT directory (extracted data, validation, etc.)
        for root, dirs, files_list in os.walk(output_dir):
            for file in files_list:
                file_path = os.path.join(root, file)
                arcname = os.path.relpath(file_path, output_dir)
                zf.write(file_path, f"output/{arcname}")
        
        # 2. Add all files from UPLOADS directory (PDFs, schema, prompts, etc.)
        if os.path.isdir(run_upload_dir):
            for root, dirs, files_list in os.walk(run_upload_dir):
                for file in files_list:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, run_upload_dir)
                    zf.write(file_path, f"uploads/{arcname}")
        
        # 3. Add all IPC logs (extraction.log, stdout.log, stderr.log)
        if os.path.isdir(run_ipc_dir):
            for file in os.listdir(run_ipc_dir):
                file_path = os.path.join(run_ipc_dir, file)
                if os.path.isfile(file_path):
                    zf.write(file_path, f"logs/{file}")
        
        # 4. Add JSON dumps of ALL database records for programmatic access
        db_dump = {
            "run": run_data,
            "sources": sources,
            "logs": logs,
            "meta_sources": meta_sources,
            "exports": exports,
            "files": files,
            "crawl_jobs": crawl_jobs,
            "export_timestamp": datetime.now(timezone.utc).isoformat(),
            "run_id": run_id
        }
        zf.writestr("database/full_dump.json", json.dumps(db_dump, indent=2, default=str))
        zf.writestr("database/run.json", json.dumps(run_data, indent=2, default=str))
        zf.writestr("database/sources.json", json.dumps(sources, indent=2, default=str))
        zf.writestr("database/logs.json", json.dumps(logs, indent=2, default=str))
        zf.writestr("database/meta_sources.json", json.dumps(meta_sources, indent=2, default=str))
        zf.writestr("database/files.json", json.dumps(files, indent=2, default=str))
        zf.writestr("database/crawl_jobs.json", json.dumps(crawl_jobs, indent=2, default=str))
        
        # 5. Create Excel workbook with run metadata and related objects
        wb = openpyxl.Workbook()
        
        # Sheet 1: Run Overview
        ws_run = wb.active
        ws_run.title = "Run Overview"
        ws_run.append(["Property", "Value"])
        for key, value in run_data.items():
            ws_run.append([str(key), str(value) if value is not None else ""])
        
        # Sheet 2: Sources
        ws_sources = wb.create_sheet("Sources")
        if sources:
            headers = list(sources[0].keys())
            ws_sources.append(headers)
            for src in sources:
                ws_sources.append([str(src.get(h, '')) for h in headers])
        else:
            ws_sources.append(["No sources found"])
        
        # Sheet 3: Meta Sources
        ws_meta = wb.create_sheet("Meta Sources")
        if meta_sources:
            headers = list(meta_sources[0].keys())
            ws_meta.append(headers)
            for ms in meta_sources:
                ws_meta.append([str(ms.get(h, '')) for h in headers])
        else:
            ws_meta.append(["No meta sources found"])
        
        # Sheet 4: Logs
        ws_logs = wb.create_sheet("Logs")
        if logs:
            headers = list(logs[0].keys())
            ws_logs.append(headers)
            for log in logs:
                ws_logs.append([str(log.get(h, '')) for h in headers])
        else:
            ws_logs.append(["No logs found"])
        
        # Sheet 5: Exports
        ws_exports = wb.create_sheet("Exports")
        if exports:
            headers = list(exports[0].keys())
            ws_exports.append(headers)
            for exp in exports:
                ws_exports.append([str(exp.get(h, '')) for h in headers])
        else:
            ws_exports.append(["No exports found"])
        
        # Sheet 6: Files
        ws_files = wb.create_sheet("Files")
        if files:
            headers = list(files[0].keys())
            ws_files.append(headers)
            for f in files:
                ws_files.append([str(f.get(h, '')) for h in headers])
        else:
            ws_files.append(["No files found"])
        
        # Sheet 7: Crawl Jobs
        ws_crawl = wb.create_sheet("Crawl Jobs")
        if crawl_jobs:
            headers = list(crawl_jobs[0].keys())
            ws_crawl.append(headers)
            for cj in crawl_jobs:
                ws_crawl.append([str(cj.get(h, '')) for h in headers])
        else:
            ws_crawl.append(["No crawl jobs found"])
        
        # Sheet 8: Extracted Data (global_data.json)
        ws_extracted = wb.create_sheet("Extracted Data")
        global_json_path = os.path.join(output_dir, "global_data.json")
        if os.path.exists(global_json_path):
            try:
                with open(global_json_path, "r", encoding="utf-8") as f:
                    extracted_data = json.load(f)
                if isinstance(extracted_data, list) and extracted_data:
                    all_keys = set()
                    for row in extracted_data:
                        if isinstance(row, dict):
                            all_keys.update(row.keys())
                    headers = sorted(list(all_keys))
                    ws_extracted.append(headers)
                    for row in extracted_data:
                        if isinstance(row, dict):
                            ws_extracted.append([str(row.get(h, '')) for h in headers])
                else:
                    ws_extracted.append(["No extracted data rows found"])
            except Exception as e:
                ws_extracted.append([f"Error reading extracted data: {e}"])
        else:
            ws_extracted.append(["Extracted data not available"])
        
        # Sheet 9: Validated Data
        ws_validated = wb.create_sheet("Validated Data")
        validated_json_path = os.path.join(output_dir, "validated_data.json")
        if os.path.exists(validated_json_path):
            try:
                with open(validated_json_path, "r", encoding="utf-8") as f:
                    validated_data = json.load(f)
                if isinstance(validated_data, list) and validated_data:
                    all_keys = set()
                    for row in validated_data:
                        if isinstance(row, dict):
                            all_keys.update(row.keys())
                    headers = sorted(list(all_keys))
                    ws_validated.append(headers)
                    for row in validated_data:
                        if isinstance(row, dict):
                            ws_validated.append([str(row.get(h, '')) for h in headers])
                else:
                    ws_validated.append(["No validated data rows found"])
            except Exception as e:
                ws_validated.append([f"Error reading validated data: {e}"])
        else:
            ws_validated.append(["Validated data not available (validation may not be enabled or complete)"])
        
        # Validation Report Sheets (if validation was enabled)
        validation_dir = os.path.join(output_dir, "validation")
        validation_report_path = os.path.join(validation_dir, "validation_report.json")
        validation_config_path = os.path.join(output_dir, "validation_config.json")
        row_flags_path = os.path.join(validation_dir, "row_flags.csv")
        
        if os.path.exists(validation_report_path):
            try:
                with open(validation_report_path, "r", encoding="utf-8") as f:
                    val_report = json.load(f)
                
                val_config = None
                if os.path.exists(validation_config_path):
                    with open(validation_config_path, "r", encoding="utf-8") as f:
                        val_config = json.load(f)
                
                # Calculate passed/failed from validation_results
                val_results = val_report.get("validation_results", [])
                val_total_rules = len(val_results)
                val_passed_rules = sum(1 for r in val_results if r.get("passed", False))
                val_error_failures = sum(1 for r in val_results if r.get("severity") == "error" and not r.get("passed", False))
                val_warning_failures = sum(1 for r in val_results if r.get("severity") == "warning" and not r.get("passed", False))
                
                # Sheet 10: Validation Summary
                ws_val_summary = wb.create_sheet("Validation Summary")
                val_summary = val_report.get("summary", {})
                ws_val_summary.append(["Validation Report Summary"])
                ws_val_summary.append([])
                ws_val_summary.append(["Metric", "Value"])
                ws_val_summary.append(["Total Rows", val_report.get("total_rows", 0)])
                ws_val_summary.append(["Overall Pass Rate", f"{val_summary.get('overall_pass_rate', 0) * 100:.1f}%"])
                ws_val_summary.append(["Total Rules", val_total_rules])
                ws_val_summary.append(["Passed Rules", val_passed_rules])
                ws_val_summary.append(["Failed Rules (Errors)", val_error_failures])
                ws_val_summary.append(["Warnings", val_warning_failures])
                
                # Sheet 11: Validation Rules
                ws_val_rules = wb.create_sheet("Validation Rules")
                ws_val_rules.append(["Rule ID", "Name", "Severity", "Status", "Affected Rows", "Description", "Columns", "Constraint"])
                
                config_rules_lookup = {}
                if val_config and "rules" in val_config:
                    for rule in val_config["rules"]:
                        config_rules_lookup[rule.get("rule_id")] = rule
                
                for result in val_report.get("validation_results", []):
                    rule_id = result.get("rule_id")
                    config_rule = config_rules_lookup.get(rule_id, {})
                    severity = result.get("severity", "warning")
                    raw_passed = result.get("passed", False)
                    status = "PASS" if raw_passed else ("FAIL" if severity == "error" else "WARN")
                    affected_count = len(result.get("affected_rows", []))
                    
                    ws_val_rules.append([
                        rule_id,
                        config_rule.get("name", ""),
                        severity.upper(),
                        status,
                        affected_count,
                        config_rule.get("description", ""),
                        ", ".join(config_rule.get("columns", [])),
                        _humanize_constraint(config_rule.get("python_expression", ""))
                    ])
                
                # Sheet 12: Validation by Column
                ws_val_cols = wb.create_sheet("Validation By Column")
                ws_val_cols.append(["Column", "Total Rules", "Errors Failed", "Warnings", "Status", "Rule IDs"])
                
                column_to_rules = {}
                for result in val_report.get("validation_results", []):
                    rule_id = result.get("rule_id")
                    config_rule = config_rules_lookup.get(rule_id, {})
                    for col in config_rule.get("columns", []):
                        if col not in column_to_rules:
                            column_to_rules[col] = []
                        column_to_rules[col].append({
                            "rule_id": rule_id,
                            "severity": result.get("severity", "warning"),
                            "passed": result.get("passed", False)
                        })
                
                for col in sorted(column_to_rules.keys()):
                    rules = column_to_rules[col]
                    error_fails = sum(1 for r in rules if r["severity"] == "error" and not r["passed"])
                    warnings = sum(1 for r in rules if r["severity"] == "warning" and not r["passed"])
                    status = "FAIL" if error_fails > 0 else ("WARN" if warnings > 0 else "PASS")
                    rule_ids = ", ".join(r["rule_id"] for r in rules)
                    ws_val_cols.append([col, len(rules), error_fails, warnings, status, rule_ids])
                
                # Sheet 13: Row Flags with Success Rate
                if os.path.exists(row_flags_path):
                    ws_row_flags = wb.create_sheet("Row Validation Flags")
                    import csv
                    with open(row_flags_path, "r", encoding="utf-8") as f:
                        reader = csv.DictReader(f)
                        rows_data = list(reader)
                    
                    if rows_data:
                        flag_cols = [k for k in rows_data[0].keys() if k != "row_index"]
                        headers = ["Row", "Passed", "Failed", "Success Rate"] + flag_cols
                        ws_row_flags.append(headers)
                        
                        for row_data in rows_data:
                            row_idx = row_data.get("row_index", "")
                            passed = sum(1 for k in flag_cols if str(row_data.get(k, "")).lower() in ("true", "1"))
                            failed = sum(1 for k in flag_cols if str(row_data.get(k, "")).lower() in ("false", "0"))
                            total = passed + failed
                            success_rate = f"{(passed / total * 100):.1f}%" if total > 0 else "N/A"
                            
                            row_values = [row_idx, passed, failed, success_rate]
                            for k in flag_cols:
                                val = row_data.get(k, "")
                                if str(val).lower() == "true":
                                    row_values.append("PASS")
                                elif str(val).lower() == "false":
                                    row_values.append("FAIL")
                                else:
                                    row_values.append(val)
                            ws_row_flags.append(row_values)
            except Exception as e:
                ws_val_err = wb.create_sheet("Validation Error")
                ws_val_err.append([f"Error loading validation data: {e}"])
        
        # Save Excel to ZIP
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        zf.writestr("run_metadata.xlsx", excel_buffer.read())
        
        # Generate and add PDF report
        try:
            from report_generator import generate_report_from_run_dir
            import tempfile
            
            pdf_path = os.path.join(tempfile.gettempdir(), f"report_{run_id}.pdf")
            generate_report_from_run_dir(run_id, run_data, output_dir, pdf_path)
            
            if os.path.exists(pdf_path):
                with open(pdf_path, 'rb') as pf:
                    zf.writestr("run_report.pdf", pf.read())
                os.remove(pdf_path)
        except Exception as e:
            # Add error note if PDF generation fails
            zf.writestr("pdf_report_error.txt", f"Failed to generate PDF report: {e}")
        
        # Generate and add Validation PDF report (if validation exists)
        if os.path.exists(validation_report_path):
            try:
                from report_generator import generate_validation_report_pdf
                import tempfile
                
                # Load required data for validation PDF
                with open(validation_report_path, "r", encoding="utf-8") as f:
                    val_report_data = json.load(f)
                
                val_config_data = None
                if os.path.exists(validation_config_path):
                    with open(validation_config_path, "r", encoding="utf-8") as f:
                        val_config_data = json.load(f)
                
                # Load extracted data
                extracted_for_pdf = []
                global_json_for_pdf = os.path.join(output_dir, "global_data.json")
                if os.path.exists(global_json_for_pdf):
                    with open(global_json_for_pdf, "r", encoding="utf-8") as f:
                        extracted_for_pdf = json.load(f)
                        if not isinstance(extracted_for_pdf, list):
                            extracted_for_pdf = [extracted_for_pdf] if extracted_for_pdf else []
                
                # Load schema fields
                schema_fields_for_pdf = None
                schema_mapping_for_pdf = os.path.join(output_dir, "schema_mapping.json")
                if os.path.exists(schema_mapping_for_pdf):
                    with open(schema_mapping_for_pdf, "r", encoding="utf-8") as f:
                        sm = json.load(f)
                        if 'fields' in sm:
                            schema_fields_for_pdf = sm['fields']
                        elif 'fieldDefs' in sm:
                            schema_fields_for_pdf = [fd.get('name') for fd in sm['fieldDefs'] if fd.get('name')]
                
                val_pdf_path = os.path.join(tempfile.gettempdir(), f"validation_report_{run_id}.pdf")
                generate_validation_report_pdf(
                    run_data=run_data,
                    extracted_data=extracted_for_pdf,
                    validation_report=val_report_data,
                    validation_config=val_config_data,
                    output_path=val_pdf_path,
                    schema_fields=schema_fields_for_pdf
                )
                
                if os.path.exists(val_pdf_path):
                    with open(val_pdf_path, 'rb') as vpf:
                        zf.writestr("validation_report.pdf", vpf.read())
                    os.remove(val_pdf_path)
            except Exception as e:
                zf.writestr("validation_pdf_error.txt", f"Failed to generate validation PDF report: {e}")
    
    zip_buffer.seek(0)
    
    # Save ZIP to exports folder
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{run_name}_{timestamp}_complete.zip"
    filepath = os.path.join(EXPORTS_FOLDER, filename)
    
    with open(filepath, 'wb') as f:
        f.write(zip_buffer.read())
    
    # Record in DB
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO exports (run_id, created_at, file_path) VALUES (?, ?, ?)",
        (run_id, datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"), filepath)
    )
    export_id = cur.lastrowid
    conn.commit()
    conn.close()
    
    log_message(f"ZIP export generated: {filename}", "INFO", run_id)
    return jsonify({
        "url": f"/exports/{export_id}/download",
        "filename": filename,
        "exportId": export_id
    })


# ============================================================================
# API Routes - Config (per-user configuration)
# ============================================================================

@app.route("/config", methods=["GET"])
@optional_auth
def list_config():
    """List config entries for current user.
    
    Returns merged config: global defaults + user overrides.
    User-specific values take precedence over global defaults.
    """
    user = g.current_user
    user_id = user["id"] if user else None
    
    conn = get_db()
    cur = conn.cursor()
    
    if user_id:
        # Get global defaults and user overrides, user values take precedence
        cur.execute("""
            SELECT 
                COALESCE(u.key, g.key) as key,
                COALESCE(u.user_id, g.user_id) as user_id,
                COALESCE(u.value, g.value) as value,
                COALESCE(u.value_type, g.value_type) as value_type,
                COALESCE(u.input_type, g.input_type) as input_type,
                COALESCE(g.allowed_values, u.allowed_values) as allowed_values,
                COALESCE(g.default_value, u.default_value) as default_value,
                COALESCE(g.category, u.category) as category,
                COALESCE(g.description, u.description) as description,
                COALESCE(u.sensitive, g.sensitive) as sensitive,
                COALESCE(g.required, u.required) as required,
                COALESCE(g.display_order, u.display_order) as display_order,
                COALESCE(u.last_modified, g.last_modified) as last_modified,
                CASE WHEN u.key IS NOT NULL THEN 1 ELSE 0 END as is_user_override
            FROM config g
            LEFT JOIN config u ON g.key = u.key AND u.user_id = ?
            WHERE g.user_id IS NULL
            ORDER BY COALESCE(g.display_order, u.display_order), COALESCE(g.key, u.key)
        """, (user_id,))
    else:
        # Just return global defaults
        cur.execute("SELECT *, 0 as is_user_override FROM config WHERE user_id IS NULL ORDER BY display_order, key")
    
    rows = []
    for r in cur.fetchall():
        row = to_camel_dict(dict(r))
        row["sensitive"] = bool(row.get("sensitive"))
        row["required"] = bool(row.get("required"))
        row["isUserOverride"] = bool(row.get("isUserOverride"))
        # Parse allowed_values JSON
        if row.get("allowedValues"):
            try:
                row["allowedValues"] = json.loads(row["allowedValues"])
            except:
                row["allowedValues"] = []
        else:
            row["allowedValues"] = []
        rows.append(row)
    
    conn.close()
    return jsonify(rows)

@app.route("/config", methods=["POST"])
@optional_auth
def upsert_config():
    """Create or update a config entry for current user."""
    data = request.json
    now = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    user = g.current_user
    user_id = user["id"] if user else None
    
    key = data.get("key")
    if not key:
        return jsonify({"error": "Key is required"}), 400
    
    # Validate against allowed_values if present
    conn = get_db()
    cur = conn.cursor()
    
    # Get global config for this key to check allowed_values
    cur.execute("SELECT allowed_values FROM config WHERE key = ? AND user_id IS NULL", (key,))
    global_row = cur.fetchone()
    
    if global_row and global_row["allowed_values"]:
        try:
            allowed = json.loads(global_row["allowed_values"])
            if allowed and data.get("value") not in allowed:
                conn.close()
                return jsonify({"error": f"Value must be one of: {allowed}"}), 400
        except:
            pass
    
    # Check if config exists (handle NULL user_id properly)
    if user_id:
        cur.execute("SELECT 1 FROM config WHERE key = ? AND user_id = ?", (key, user_id))
    else:
        cur.execute("SELECT 1 FROM config WHERE key = ? AND user_id IS NULL", (key,))
    
    exists = cur.fetchone() is not None
    
    if exists:
        # Update existing config
        if user_id:
            cur.execute("""
                UPDATE config SET value = ?, value_type = ?, input_type = ?, allowed_values = ?,
                    default_value = ?, category = ?, description = ?, sensitive = ?, required = ?,
                    display_order = ?, last_modified = ?
                WHERE key = ? AND user_id = ?
            """, (
                data.get("value", ""),
                data.get("valueType", "string"),
                data.get("inputType", "text"),
                json.dumps(data.get("allowedValues")) if data.get("allowedValues") else None,
                data.get("defaultValue"),
                data.get("category", "general"),
                data.get("description", ""),
                1 if data.get("sensitive") else 0,
                1 if data.get("required") else 0,
                data.get("displayOrder", 0),
                now,
                key,
                user_id
            ))
        else:
            cur.execute("""
                UPDATE config SET value = ?, value_type = ?, input_type = ?, allowed_values = ?,
                    default_value = ?, category = ?, description = ?, sensitive = ?, required = ?,
                    display_order = ?, last_modified = ?
                WHERE key = ? AND user_id IS NULL
            """, (
                data.get("value", ""),
                data.get("valueType", "string"),
                data.get("inputType", "text"),
                json.dumps(data.get("allowedValues")) if data.get("allowedValues") else None,
                data.get("defaultValue"),
                data.get("category", "general"),
                data.get("description", ""),
                1 if data.get("sensitive") else 0,
                1 if data.get("required") else 0,
                data.get("displayOrder", 0),
                now,
                key
            ))
    else:
        # Insert new config
        cur.execute("""
            INSERT INTO config (key, user_id, value, value_type, input_type, allowed_values, 
                default_value, category, description, sensitive, required, display_order, last_modified)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            key,
            user_id,
            data.get("value", ""),
            data.get("valueType", "string"),
            data.get("inputType", "text"),
            json.dumps(data.get("allowedValues")) if data.get("allowedValues") else None,
            data.get("defaultValue"),
            data.get("category", "general"),
            data.get("description", ""),
            1 if data.get("sensitive") else 0,
            1 if data.get("required") else 0,
            data.get("displayOrder", 0),
            now
        ))
    conn.commit()
    
    # Fetch the updated row
    if user_id:
        cur.execute("SELECT * FROM config WHERE key = ? AND user_id = ?", (key, user_id))
    else:
        cur.execute("SELECT * FROM config WHERE key = ? AND user_id IS NULL", (key,))
    row = cur.fetchone()
    conn.close()
    
    if not row:
        return jsonify({"error": "Failed to save config"}), 500
    
    result = to_camel_dict(dict(row))
    result["sensitive"] = bool(result.get("sensitive"))
    result["required"] = bool(result.get("required"))
    if result.get("allowedValues"):
        try:
            result["allowedValues"] = json.loads(result["allowedValues"])
        except:
            result["allowedValues"] = []
    return jsonify(result)

@app.route("/config/<key>", methods=["DELETE"])
@optional_auth
def delete_config(key):
    """Delete a user's config override (reverts to global default)."""
    user = g.current_user
    user_id = user["id"] if user else None
    
    conn = get_db()
    cur = conn.cursor()
    
    if user_id:
        # Only delete user override, not global default
        cur.execute("DELETE FROM config WHERE key = ? AND user_id = ?", (key, user_id))
    else:
        # Delete global config (admin only in future)
        cur.execute("DELETE FROM config WHERE key = ? AND user_id IS NULL", (key,))
    
    conn.commit()
    conn.close()
    return "", 204

@app.route("/config/<key>/reset", methods=["POST"])
@optional_auth
def reset_config(key):
    """Reset a config key to its default value."""
    user = g.current_user
    user_id = user["id"] if user else None
    
    conn = get_db()
    cur = conn.cursor()
    
    # Get global default
    cur.execute("SELECT default_value FROM config WHERE key = ? AND user_id IS NULL", (key,))
    row = cur.fetchone()
    
    if not row:
        conn.close()
        return jsonify({"error": "Config key not found"}), 404
    
    default_value = row["default_value"] or ""
    now = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    
    if user_id:
        # Delete user override to revert to global
        cur.execute("DELETE FROM config WHERE key = ? AND user_id = ?", (key, user_id))
    else:
        # Reset global to default
        cur.execute("UPDATE config SET value = ?, last_modified = ? WHERE key = ? AND user_id IS NULL",
                   (default_value, now, key))
    
    conn.commit()
    conn.close()
    return jsonify({"message": "Config reset to default", "value": default_value})

@app.route("/config/import", methods=["POST"])
@optional_auth
def import_config():
    """Import config from JSON for current user."""
    data = request.json.get("data", {})
    now = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    user = g.current_user
    user_id = user["id"] if user else None
    
    conn = get_db()
    cur = conn.cursor()
    
    for key, value in data.items():
        # Only update value, preserve other metadata from global
        cur.execute("""
            INSERT INTO config (key, user_id, value, last_modified)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(key, user_id) DO UPDATE SET value = excluded.value, last_modified = excluded.last_modified
        """, (key, user_id, value if isinstance(value, str) else json.dumps(value), now))
    
    conn.commit()
    conn.close()
    
    return "", 204

@app.route("/config/categories", methods=["GET"])
def list_config_categories():
    """List all config categories with counts."""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT category, COUNT(*) as count 
        FROM config 
        WHERE user_id IS NULL 
        GROUP BY category 
        ORDER BY 
            CASE category 
                WHEN 'general' THEN 1 
                WHEN 'llm' THEN 2 
                WHEN 'extraction' THEN 3 
                WHEN 'api_keys' THEN 4 
                WHEN 'advanced' THEN 5 
                ELSE 6 
            END
    """)
    rows = [{"category": r["category"], "count": r["count"]} for r in cur.fetchall()]
    conn.close()
    return jsonify(rows)

# ============================================================================
# API Routes - Upload (1 endpoint)
# ============================================================================

@app.route("/upload", methods=["POST"])
def upload_file():
    """Handle file upload."""
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400
    
    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "No file selected"}), 400
    
    filename = secure_filename(file.filename)
    unique_name = f"{uuid.uuid4().hex[:8]}_{filename}"
    filepath = os.path.join(UPLOAD_FOLDER, unique_name)
    file.save(filepath)
    
    log_message(f"File uploaded: {filename}", "INFO")
    return jsonify({"url": f"/uploads/{unique_name}"})

@app.route("/uploads/<filename>", methods=["GET"])
def serve_upload(filename):
    """Serve an uploaded file."""
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    if not os.path.exists(filepath):
        return jsonify({"error": "File not found"}), 404
    return send_file(filepath)

# ============================================================================
# API Routes - Sources (backed by sources table) (2 endpoints)
# ============================================================================

def _list_sources_impl():
    page = int(request.args.get("page", 1))
    page_size = int(request.args.get("pageSize", 10))
    q = request.args.get("q", "")
    domain = request.args.get("domain", "")

    conn = get_db()
    cur = conn.cursor()

    query = "SELECT * FROM sources WHERE 1=1"
    params = []
    if q:
        query += " AND url LIKE ?"
        params.append(f"%{q}%")
    if domain:
        query += " AND domain = ?"
        params.append(domain)

    cur.execute(query, params)
    rows = [to_camel_dict(dict(r)) for r in cur.fetchall()]
    conn.close()

    return jsonify(paginate(rows, page, page_size))


def _get_source_impl(source_id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM sources WHERE id = ?", (source_id,))
    row = cur.fetchone()
    conn.close()

    if not row:
        return jsonify({"error": "Source not found"}), 404

    return jsonify(to_camel_dict(dict(row)))

@app.route("/sources", methods=["GET"], endpoint="sources_list")
def sources_list():
    """List sources with pagination."""
    return _list_sources_impl()

@app.route("/sources/<source_id>", methods=["GET"], endpoint="sources_get")
def sources_get(source_id):
    """Get source details."""

    return _get_source_impl(source_id)


@app.route("/runs/<run_id>/sources", methods=["GET"], endpoint="run_sources_list")
@optional_auth
def run_sources_list(run_id):
    page = int(request.args.get("page", 1))
    page_size = int(request.args.get("pageSize", 50))
    q = request.args.get("q", "")
    domain = request.args.get("domain", "")

    user = get_current_user()
    conn = get_db()
    cur = conn.cursor()

    run_row = None
    try:
        if user:
            cur.execute("SELECT id, source_type, user_id FROM runs WHERE id = ? AND user_id = ?", (run_id, user["id"]))
        else:
            cur.execute("SELECT id, source_type, user_id FROM runs WHERE id = ?", (run_id,))
        run_row = cur.fetchone()
    except Exception:
        run_row = None

    if not run_row:
        conn.close()
        return jsonify({"error": "Run not found"}), 404

    run_source_type = (run_row["source_type"] or "pdf") if run_row else "pdf"

    # For PDF runs: attempt to materialize HTML for PENDING/PROCESSING pdf sources in-place
    if run_source_type == "pdf":
        try:
            from pdf_converter import convert_pdf_to_text
        except Exception:
            convert_pdf_to_text = None

        if convert_pdf_to_text is not None:
            try:
                cur.execute(
                    """
                    SELECT id, pdf_file_id, title
                    FROM sources
                    WHERE run_id = ? AND source_type = 'pdf' AND (status IS NULL OR status != 'READY') AND pdf_file_id IS NOT NULL
                    """,
                    (run_id,),
                )
                for s in cur.fetchall():
                    pdf_file_id = s["pdf_file_id"]
                    pdf_path = get_file_internal_path(pdf_file_id)
                    if not pdf_path or not os.path.exists(pdf_path):
                        continue
                    try:
                        pdf_text = convert_pdf_to_text(pdf_path, use_cache=True)
                    except Exception:
                        continue
                    if not pdf_text or len(str(pdf_text).strip()) == 0:
                        continue
                    title = s["title"] or os.path.basename(pdf_path)
                    html_content = f"""<!DOCTYPE html>
<html>
<head><title>{title}</title></head>
<body>
<source>
<h1>{title}</h1>
<div class=\"pdf-content\">\n{pdf_text}\n</div>
</source>
</body>
</html>"""
                    now = datetime.now(timezone.utc).isoformat()
                    cur.execute(
                        """
                        UPDATE sources
                        SET html_content = ?, content_type = 'pdf', status = 'READY', error = NULL, updated_at = ?
                        WHERE id = ?
                        """,
                        (html_content, now, s["id"]),
                    )
                conn.commit()
            except Exception:
                pass

    # Return document sources from sources table (documents only)
    sql = "SELECT * FROM sources WHERE run_id = ?"
    params = [run_id]
    if q:
        sql += " AND (title LIKE ? OR url LIKE ?)"
        params.append(f"%{q}%")
        params.append(f"%{q}%")
    if domain:
        sql += " AND domain = ?"
        params.append(domain)
    sql += " ORDER BY created_at DESC"

    cur.execute(sql, params)
    source_rows = cur.fetchall()
    
    # Get run output_dir to find metadata files
    cur.execute("SELECT output_dir FROM runs WHERE id = ?", (run_id,))
    run_info = cur.fetchone()
    output_dir = run_info["output_dir"] if run_info else None
    
    rows = []
    for r in source_rows:
        d = to_camel_dict(dict(r))
        d["pdfDownloadUrl"] = f"/files/{d.get('pdfFileId')}/download" if d.get("pdfFileId") else None
        
        # Try to load source metadata from extraction output
        if output_dir:
            title = d.get("title") or ""
            # Try to find metadata file by title (PDF filename without extension)
            if title:
                base_name = os.path.splitext(title)[0]
                metadata_path = os.path.join(output_dir, "sources", f"{base_name}_metadata.json")
                if os.path.isfile(metadata_path):
                    try:
                        with open(metadata_path, 'r', encoding='utf-8') as f:
                            metadata = json.load(f)
                        d["rowCount"] = metadata.get("row_count")
                        d["rowCountLogic"] = metadata.get("row_count_logic")
                        d["rowCountReasoning"] = metadata.get("row_count_reasoning")
                        d["rowCountCandidates"] = metadata.get("all_candidates")
                        d["rejectionReason"] = metadata.get("rejection_reason")
                        d["extractedRows"] = metadata.get("extracted_rows")
                        d["rowCountMismatch"] = metadata.get("row_count_mismatch", False)
                        d["expectedRowCount"] = metadata.get("expected_row_count")
                    except Exception:
                        pass
        
        rows.append(d)
    conn.close()
    return jsonify(paginate(rows, page, page_size))


@app.route("/sources/<source_id>/preview", methods=["GET"], endpoint="source_preview")
@optional_auth
def source_preview(source_id):
    user = get_current_user()
    conn = get_db()
    cur = conn.cursor()

    if user:
        cur.execute(
            """
            SELECT s.*
            FROM sources s
            LEFT JOIN runs r ON r.id = s.run_id
            WHERE s.id = ? AND (r.user_id = ? OR r.user_id IS NULL)
            """,
            (source_id, user["id"]),
        )
    else:
        cur.execute("SELECT * FROM sources WHERE id = ?", (source_id,))

    row = cur.fetchone()
    if row:
        conn.close()
        d = dict(row)
        content_type = d.get("content_type") or "html"
        html = d.get("html_content") or ""
        pdf_file_id = d.get("pdf_file_id")

        resp = {
            "id": d.get("id"),
            "runId": d.get("run_id"),
            "url": d.get("url"),
            "domain": d.get("domain"),
            "title": d.get("title"),
            "sourceType": d.get("source_type"),
            "status": d.get("status"),
            "error": d.get("error"),
            "metaSourceId": d.get("meta_source_id"),
            "contentType": content_type,
            "createdAt": d.get("created_at"),
            "htmlContent": html,
            "pdfFileId": pdf_file_id,
            "pdfDownloadUrl": f"/files/{pdf_file_id}/download" if pdf_file_id else None,
        }
        return jsonify(resp)

    # Fallback: treat source_id as a file_id for PDF sources (pdf runs)
    file_row = None
    try:
        if user:
            cur.execute(
                """
                SELECT f.*
                FROM files f
                LEFT JOIN runs r ON r.id = f.run_id
                WHERE f.id = ? AND (r.user_id = ? OR r.user_id IS NULL)
                """,
                (source_id, user["id"]),
            )
        else:
            cur.execute("SELECT * FROM files WHERE id = ?", (source_id,))
        file_row = cur.fetchone()
    except Exception:
        file_row = None
    conn.close()

    if not file_row:
        return jsonify({"error": "Source not found"}), 404

    d = dict(file_row)
    original_name = d.get("original_name") or d.get("filename")
    resp = {
        "id": d.get("id"),
        "runId": d.get("run_id"),
        "url": None,
        "domain": "",
        "title": original_name,
        "sourceType": "pdf",
        "status": "READY",
        "error": None,
        "metaSourceId": None,
        "contentType": "pdf",
        "createdAt": d.get("created_at"),
        "previewText": "",
        "pdfFileId": d.get("id"),
        "pdfDownloadUrl": f"/files/{d.get('id')}/download" if d.get("id") else None,
    }
    return jsonify(resp)

# ============================================================================
# API Routes - Exports (3 endpoints)
# ============================================================================

@app.route("/exports", methods=["GET"])
def list_exports():
    """List exports with pagination. NO PATHS EXPOSED."""
    page = int(request.args.get("page", 1))
    page_size = int(request.args.get("pageSize", 10))
    
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id, run_id, created_at, file_path FROM exports ORDER BY created_at DESC")
    rows = cur.fetchall()
    conn.close()
    
    # Build response without exposing full paths
    items = []
    for row in rows:
        filename = os.path.basename(row["file_path"]) if row["file_path"] else "export.json"
        items.append({
            "id": row["id"],
            "runId": row["run_id"],
            "createdAt": row["created_at"],
            "filename": filename
        })
    
    return jsonify(paginate(items, page, page_size))

@app.route("/exports/<int:export_id>/download", methods=["GET", "POST"])
def download_export(export_id):
    """Download an export file directly. NO PATH EXPOSED."""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT file_path FROM exports WHERE id = ?", (export_id,))
    row = cur.fetchone()
    conn.close()
    
    if not row:
        return jsonify({"error": "Export not found"}), 404
    
    filepath = row["file_path"]
    if not os.path.exists(filepath):
        return jsonify({"error": "Export file not found on disk"}), 404
    
    filename = os.path.basename(filepath)
    return send_file(filepath, as_attachment=True, download_name=filename)

@app.route("/exports/<int:export_id>", methods=["DELETE"])
def delete_export(export_id):
    """Delete an export."""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT file_path FROM exports WHERE id = ?", (export_id,))
    row = cur.fetchone()
    
    if row and os.path.exists(row["file_path"]):
        os.remove(row["file_path"])
    
    cur.execute("DELETE FROM exports WHERE id = ?", (export_id,))
    conn.commit()
    conn.close()
    
    return "", 204

# ============================================================================
# API Routes - Domains (1 endpoint)
# ============================================================================

@app.route("/domains", methods=["GET"])
def list_domains():
    """List domains with pagination."""
    page = int(request.args.get("page", 1))
    page_size = int(request.args.get("pageSize", 10))
    
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM domains ORDER BY visited_count DESC")
    rows = [to_camel_dict(dict(r)) for r in cur.fetchall()]
    conn.close()
    
    return jsonify(paginate(rows, page, page_size))

# ============================================================================
# API Routes - Cache (5 endpoints) - Uses real cache_utils.py
# ============================================================================

@app.route("/cache/providers", methods=["GET"])
def list_cache_providers():
    """List all cache providers from real cache directory."""
    # Use real cache_utils.get_cache_stats()
    stats = get_cache_stats()
    now = datetime.now(timezone.utc).isoformat()
    
    providers = []
    for name, data in stats.items():
        providers.append({
            "id": name,
            "name": name.title(),
            "type": "LLM" if name == "gpt" else "SCRAPING" if name == "surya" else "SCHEMA",
            "entriesCount": data.get("count", 0),
            "totalSizeBytes": int(data.get("size_mb", 0) * 1024 * 1024),
            "hitRate": 0.0,  # Would need to track this separately
            "lastAccessed": now
        })
    
    return jsonify(providers)

@app.route("/cache/entries", methods=["GET"])
def list_cache_entries():
    """List cache entries from real cache directory."""
    page = int(request.args.get("page", 1))
    page_size = int(request.args.get("pageSize", 10))
    provider_id = request.args.get("providerId", "")
    
    entries = []
    subdirs = [provider_id] if provider_id else ["surya", "gpt", "schema"]
    
    for subdir in subdirs:
        cache_path = CACHE_DIR / subdir
        if cache_path.exists():
            for f in cache_path.glob("*"):
                if f.is_file() and not f.name.endswith('.meta.json'):
                    try:
                        stat = f.stat()
                        entries.append({
                            "id": f"{subdir}/{f.name}",
                            "providerId": subdir,
                            "key": f.stem,
                            "sizeBytes": stat.st_size,
                            "hitCount": 0,
                            "createdDate": datetime.fromtimestamp(stat.st_ctime, timezone.utc).isoformat(),
                            "lastAccessed": datetime.fromtimestamp(stat.st_atime, timezone.utc).isoformat(),
                            "status": "ACTIVE"
                        })
                    except:
                        pass
    
    # Sort by last accessed descending
    entries.sort(key=lambda x: x["lastAccessed"], reverse=True)
    
    return jsonify(paginate(entries, page, page_size))

@app.route("/cache/entries/<path:entry_id>", methods=["DELETE"])
def delete_cache_entry(entry_id):
    """Delete a cache entry from disk."""
    try:
        cache_file = CACHE_DIR / entry_id
        if cache_file.exists():
            cache_file.unlink()
            # Also delete meta file if exists
            meta_file = cache_file.with_suffix(cache_file.suffix + '.meta.json')
            if meta_file.exists():
                meta_file.unlink()
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    return "", 204

@app.route("/cache/providers/<provider_id>/clear", methods=["POST"])
def clear_cache_provider(provider_id):
    """Clear all entries for a cache provider using real cache_utils."""
    count = clear_cache(provider_id)
    log_message(f"Cleared {count} cache entries from {provider_id}", "INFO")
    return "", 204

@app.route("/cache/clear-all", methods=["POST"])
def clear_all_cache():
    """Clear all cache entries using real cache_utils."""
    total = 0
    for subdir in ["surya", "gpt", "schema"]:
        total += clear_cache(subdir)
    log_message(f"Cleared {total} total cache entries", "INFO")
    return "", 204


# ============================================================================
# API Routes - Logs (1 endpoint)
# ============================================================================

@app.route("/logs", methods=["GET"])
def list_logs():
    """List logs with pagination."""
    page = int(request.args.get("page", 1))
    page_size = int(request.args.get("pageSize", 50))
    level = request.args.get("level", "")
    q = request.args.get("q", "")
    
    conn = get_db()
    cur = conn.cursor()
    
    query = "SELECT * FROM logs WHERE 1=1"
    params = []
    if level and level != "ALL":
        query += " AND level = ?"
        params.append(level)
    if q:
        query += " AND message LIKE ?"
        params.append(f"%{q}%")
    query += " ORDER BY id DESC"
    
    cur.execute(query, params)
    rows = [to_camel_dict(dict(r)) for r in cur.fetchall()]
    conn.close()
    
    return jsonify(paginate(rows, page, page_size))

# ============================================================================
# API Routes - Server Logs SSE (2 endpoints)
# ============================================================================

@app.route("/server-logs/tail", methods=["GET"])
def server_logs_tail():
    """Get recent server logs."""
    max_lines = int(request.args.get("maxLines", 200))
    
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM logs ORDER BY id DESC LIMIT ?", (max_lines,))
    rows = [to_camel_dict(dict(r)) for r in cur.fetchall()]
    rows.reverse()
    conn.close()
    
    return jsonify({"items": rows})

@app.route("/server-logs/stream", methods=["GET"])
def server_logs_stream():
    """SSE stream of live logs."""
    def generate():
        last_id = 0
        while True:
            with log_buffer_lock:
                new_logs = [log for log in log_buffer if log["id"] > last_id]
                if new_logs:
                    last_id = new_logs[-1]["id"]
            
            for log in new_logs:
                yield f"data: {log['createdAt']} {log['message']}\n\n"
            
            time.sleep(0.5)
    
    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no"
        }
    )

# ============================================================================
# API Routes - Auth
# ============================================================================

@app.route("/signup", methods=["POST"])
def signup():
    """Register a new user."""
    data = request.json or {}
    email = data.get("email", "").strip().lower()
    password = data.get("password", "")
    
    if not email or not password:
        return jsonify({"error": "Email and password are required"}), 400
    
    if len(password) < 6:
        return jsonify({"error": "Password must be at least 6 characters"}), 400
    
    conn = get_db()
    cur = conn.cursor()
    
    # Check if email already exists
    cur.execute("SELECT id FROM users WHERE email = ?", (email,))
    if cur.fetchone():
        conn.close()
        return jsonify({"error": "Email already registered"}), 409
    
    # Create user
    user_id = str(uuid.uuid4())
    password_hash = hash_password(password)
    now = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    
    cur.execute("""
        INSERT INTO users (id, email, password_hash, created_at, is_active)
        VALUES (?, ?, ?, ?, 1)
    """, (user_id, email, password_hash, now))
    conn.commit()
    conn.close()
    
    # Create token
    token = create_token(user_id, email)
    
    log_message(f"User registered: {email}", "INFO")
    
    return jsonify({
        "id": user_id,
        "email": email,
        "createdAt": now,
        "isActive": True,
        "token": token
    }), 201

@app.route("/signin", methods=["POST"])
def signin():
    """User login."""
    data = request.json or {}
    email = data.get("email", "").strip().lower()
    password = data.get("password", "")
    
    if not email or not password:
        return jsonify({"error": "Email and password are required"}), 400
    
    conn = get_db()
    cur = conn.cursor()
    
    cur.execute("SELECT id, email, password_hash, created_at, is_active FROM users WHERE email = ?", (email,))
    row = cur.fetchone()
    conn.close()
    
    if not row:
        return jsonify({"error": "Invalid email or password"}), 401
    
    user = dict(row)
    
    if not verify_password(password, user["password_hash"]):
        return jsonify({"error": "Invalid email or password"}), 401
    
    if not user["is_active"]:
        return jsonify({"error": "Account is disabled"}), 403
    
    # Create token
    token = create_token(user["id"], user["email"])
    
    log_message(f"User signed in: {email}", "INFO")
    
    return jsonify({
        "id": user["id"],
        "email": user["email"],
        "createdAt": user["created_at"],
        "isActive": bool(user["is_active"]),
        "token": token
    })

@app.route("/signout", methods=["POST"])
def signout():
    """User logout."""
    # Token-based auth doesn't need server-side logout
    # Client just discards the token
    return "", 204

@app.route("/me", methods=["GET"])
@require_auth
def get_me():
    """Get current authenticated user."""
    user = g.current_user
    
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id, email, created_at, is_active FROM users WHERE id = ?", (user["id"],))
    row = cur.fetchone()
    conn.close()
    
    if not row:
        return jsonify({"error": "User not found"}), 404
    
    user_data = dict(row)
    return jsonify({
        "id": user_data["id"],
        "email": user_data["email"],
        "createdAt": user_data["created_at"],
        "isActive": bool(user_data["is_active"])
    })

# ============================================================================
# Deep Research API - Gemini Deep Research Integration
# ============================================================================

DEEP_RESEARCH_BASE_URL = "https://generativelanguage.googleapis.com/v1beta/interactions"

def get_gemini_api_key(user_id=None):
    """Get Gemini API key from user config, global config, or environment.
    
    Priority: user-specific config > global config > environment variable
    """
    conn = get_db()
    cur = conn.cursor()
    
    # First try user-specific config if user_id provided
    if user_id:
        cur.execute("SELECT value FROM config WHERE key = 'GEMINI_API_KEY' AND user_id = ?", (user_id,))
        row = cur.fetchone()
        if row and row["value"]:
            conn.close()
            return row["value"]
    
    # Fall back to global config
    cur.execute("SELECT value FROM config WHERE key = 'GEMINI_API_KEY' AND user_id IS NULL")
    row = cur.fetchone()
    conn.close()
    if row and row["value"]:
        return row["value"]
    
    # Fall back to environment variable
    return os.environ.get("GEMINI_API_KEY", "")

def deep_research_start(query: str, api_key: str) -> dict:
    """Start a Deep Research task with Gemini API."""
    import requests as req
    headers = {
        "Content-Type": "application/json",
        "x-goog-api-key": api_key
    }
    payload = {
        "input": query,
        "agent": "deep-research-pro-preview-12-2025",
        "background": True
    }
    resp = req.post(DEEP_RESEARCH_BASE_URL, headers=headers, json=payload, timeout=60)
    resp.raise_for_status()
    return resp.json()

def deep_research_status(interaction_id: str, api_key: str) -> dict:
    """Check status of a Deep Research task."""
    import requests as req
    headers = {"x-goog-api-key": api_key}
    url = f"{DEEP_RESEARCH_BASE_URL}/{interaction_id}"
    resp = req.get(url, headers=headers, timeout=30)
    resp.raise_for_status()
    return resp.json()

def extract_links_with_gemini(report_text: str, api_key: str) -> dict:
    """Extract links from research report using Gemini."""
    import requests as req
    import re
    
    gemini_url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash-exp:generateContent?key={api_key}"
    
    extraction_prompt = f"""
    You are a scientific data extraction assistant.
    Input: A research report containing analysis of scientific papers and links.
    Task: Extract all relevant external links (URLs) identified in the report that match the research criteria.
    Output: A valid JSON object with a single key "extracted_links" containing a list of objects. Each object should have "url", "title" (if available), and "relevance_score" (inferred 0-100).
    
    Report Content:
    {report_text[:50000]}
    
    Return ONLY raw JSON. No markdown formatting.
    """
    
    payload = {
        "contents": [{
            "parts": [{"text": extraction_prompt}]
        }]
    }
    
    resp = req.post(gemini_url, headers={"Content-Type": "application/json"}, json=payload, timeout=120)
    resp.raise_for_status()
    data = resp.json()
    
    try:
        gemini_text = data['candidates'][0]['content']['parts'][0]['text']
        gemini_text = re.sub(r'```json\s*', '', gemini_text)
        gemini_text = re.sub(r'```\s*', '', gemini_text)
        return json.loads(gemini_text)
    except Exception:
        return {"extracted_links": [], "error": "Failed to parse extraction response"}

def poll_deep_research(run_id: str, interaction_id: str, api_key: str):
    """Background thread to poll Deep Research status and update database."""
    import requests as req
    
    max_wait = 3600  # 1 hour
    poll_interval = 20
    start_time = time.time()
    logs = []
    
    def add_log(msg, level="INFO"):
        timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
        logs.append(f"[{timestamp}] [{level}] {msg}")
    
    add_log(f"Starting polling for interaction {interaction_id}")
    
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE deep_research_runs SET status = 'running', started_at = ? WHERE id = ?",
                (datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"), run_id))
    conn.commit()
    conn.close()
    
    while time.time() - start_time < max_wait:
        try:
            status = deep_research_status(interaction_id, api_key)
            state = status.get("state", status.get("status", "UNKNOWN")).upper()
            add_log(f"Current state: {state}")
            
            if state == "COMPLETED":
                add_log("Research completed!", "SUCCESS")
                
                # Extract result text
                result_text = None
                if "output" in status:
                    if isinstance(status["output"], dict):
                        result_text = status["output"].get("text", "")
                    elif isinstance(status["output"], str):
                        result_text = status["output"]
                
                if not result_text and "outputs" in status:
                    texts = []
                    for output in status["outputs"]:
                        if isinstance(output, dict):
                            texts.append(output.get("text", ""))
                        elif isinstance(output, str):
                            texts.append(output)
                    result_text = "\n\n".join(texts)
                
                if not result_text:
                    for key in ["content", "result", "response", "report", "text", "message"]:
                        if key in status and status[key]:
                            result_text = status[key]
                            break
                
                # Extract links - first check if API returned them directly
                extracted_links = {"extracted_links": []}
                
                # Check for groundingMetadata.sources from Gemini API
                grounding = status.get("groundingMetadata", {})
                sources = grounding.get("sources", []) or grounding.get("groundingSources", [])
                if sources:
                    add_log(f"Found {len(sources)} sources in groundingMetadata")
                    for src in sources:
                        if isinstance(src, dict):
                            url = src.get("uri") or src.get("url") or src.get("link", "")
                            title = src.get("title") or src.get("name", "")
                            if url:
                                extracted_links["extracted_links"].append({
                                    "url": url,
                                    "title": title,
                                    "relevance_score": 80
                                })
                
                # Also check for citations/references in the response
                citations = status.get("citations", []) or status.get("references", [])
                if citations:
                    add_log(f"Found {len(citations)} citations")
                    for cite in citations:
                        if isinstance(cite, dict):
                            url = cite.get("uri") or cite.get("url") or cite.get("link", "")
                            title = cite.get("title") or cite.get("name", "")
                            if url and url not in [l["url"] for l in extracted_links["extracted_links"]]:
                                extracted_links["extracted_links"].append({
                                    "url": url,
                                    "title": title,
                                    "relevance_score": 75
                                })
                
                # If no links from API metadata, extract from report text using Gemini
                if not extracted_links["extracted_links"] and result_text:
                    add_log(f"Extracting links from {len(result_text)} chars of report via Gemini")
                    try:
                        gemini_links = extract_links_with_gemini(result_text, api_key)
                        extracted_links = gemini_links
                        link_count = len(extracted_links.get("extracted_links", []))
                        add_log(f"Extracted {link_count} links via Gemini", "SUCCESS")
                    except Exception as e:
                        add_log(f"Link extraction failed: {e}", "ERROR")
                else:
                    add_log(f"Total links from API metadata: {len(extracted_links['extracted_links'])}", "SUCCESS")
                
                # Update database
                conn = get_db()
                cur = conn.cursor()
                now = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
                cur.execute("""
                    UPDATE deep_research_runs 
                    SET status = 'completed', result_text = ?, extracted_links = ?, 
                        logs = ?, completed_at = ?
                    WHERE id = ?
                """, (
                    result_text or "",
                    json.dumps(extracted_links),
                    "\n".join(logs),
                    now,
                    run_id
                ))
                
                # Get user_id for this deep research run
                cur.execute("SELECT user_id FROM deep_research_runs WHERE id = ?", (run_id,))
                dr_row = cur.fetchone()
                user_id = dr_row["user_id"] if dr_row else None
                
                # Auto-create crawl jobs for extracted links (HTML only - PDFs use Surya pipeline)
                links_list = extracted_links.get("extracted_links", [])
                if links_list and user_id:
                    html_count = 0
                    pdf_count = 0
                    for link_item in links_list:
                        link_url = link_item.get("url", "") if isinstance(link_item, dict) else str(link_item)
                        link_title = link_item.get("title", "") if isinstance(link_item, dict) else ""
                        if not link_url:
                            continue
                        
                        # Skip PDF links - they should use Surya server-side processing
                        url_lower = link_url.lower()
                        if url_lower.endswith('.pdf') or '/pdf/' in url_lower or 'pdf?' in url_lower:
                            pdf_count += 1
                            continue
                        
                        job_id = str(uuid.uuid4())
                        cur.execute("""
                            INSERT INTO crawl_jobs (id, deep_research_id, run_id, user_id, url, title, status, created_at)
                            VALUES (?, ?, ?, ?, ?, ?, 'PENDING', ?)
                        """, (job_id, run_id, run_id, user_id, link_url, link_title, now))
                        html_count += 1
                    
                    add_log(f"Created {html_count} HTML crawl jobs (skipped {pdf_count} PDFs for Surya pipeline)", "SUCCESS")
                
                conn.commit()
                conn.close()
                return
            
            elif state == "FAILED":
                error_msg = status.get("error", "Unknown error")
                add_log(f"Research failed: {error_msg}", "ERROR")
                
                conn = get_db()
                cur = conn.cursor()
                cur.execute("""
                    UPDATE deep_research_runs 
                    SET status = 'failed', error = ?, logs = ?, completed_at = ?
                    WHERE id = ?
                """, (error_msg, "\n".join(logs), datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"), run_id))
                conn.commit()
                conn.close()
                return
            
            # Update logs periodically
            conn = get_db()
            cur = conn.cursor()
            cur.execute("UPDATE deep_research_runs SET logs = ? WHERE id = ?", ("\n".join(logs), run_id))
            conn.commit()
            conn.close()
            
        except Exception as e:
            add_log(f"Poll error: {e}", "ERROR")
        
        time.sleep(poll_interval)
    
    # Timeout
    add_log("Timeout reached", "WARNING")
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        UPDATE deep_research_runs 
        SET status = 'timeout', error = 'Research did not complete within timeout', 
            logs = ?, completed_at = ?
        WHERE id = ?
    """, ("\n".join(logs), datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"), run_id))
    conn.commit()
    conn.close()


@app.route("/deep-research", methods=["GET"])
@optional_auth
def list_deep_research_runs():
    """List all Deep Research runs."""
    page = int(request.args.get("page", 1))
    page_size = int(request.args.get("pageSize", 10))
    
    conn = get_db()
    cur = conn.cursor()
    
    # Get user's runs if authenticated, otherwise all runs
    user_id = g.current_user["id"] if g.current_user else None
    
    if user_id:
        cur.execute("SELECT COUNT(*) FROM deep_research_runs WHERE user_id = ?", (user_id,))
    else:
        cur.execute("SELECT COUNT(*) FROM deep_research_runs")
    total = cur.fetchone()[0]
    
    offset = (page - 1) * page_size
    if user_id:
        cur.execute("""
            SELECT id, name, status, interaction_id, query, created_at, started_at, completed_at, error
            FROM deep_research_runs WHERE user_id = ?
            ORDER BY created_at DESC LIMIT ? OFFSET ?
        """, (user_id, page_size, offset))
    else:
        cur.execute("""
            SELECT id, name, status, interaction_id, query, created_at, started_at, completed_at, error
            FROM deep_research_runs
            ORDER BY created_at DESC LIMIT ? OFFSET ?
        """, (page_size, offset))
    
    rows = cur.fetchall()
    conn.close()
    
    items = []
    for row in rows:
        items.append({
            "id": row["id"],
            "name": row["name"],
            "status": row["status"],
            "interactionId": row["interaction_id"],
            "query": row["query"][:200] + "..." if len(row["query"] or "") > 200 else row["query"],
            "createdAt": row["created_at"],
            "startedAt": row["started_at"],
            "completedAt": row["completed_at"],
            "error": row["error"]
        })
    
    return jsonify({
        "items": items,
        "total": total,
        "page": page,
        "pageSize": page_size
    })


@app.route("/deep-research", methods=["POST"])
@require_auth
def create_deep_research_run():
    """Create and start a new Deep Research run.
    
    Request body:
    {
        "name": "Research name",
        "query": "Research query/prompt",
        "searchConfig": {
            "target_source_count": 50,
            "allowed_sources": ["scholar.google.com", "pubmed.ncbi.nlm.nih.gov"],
            "exclude_sources": []
        }
    }
    """
    user_id = g.current_user["id"]
    
    data = request.json
    if not data:
        return jsonify({"error": "Request body required"}), 400
    
    query = data.get("query", "").strip()
    if not query:
        return jsonify({"error": "Query is required"}), 400
    
    name = data.get("name", "Deep Research Run").strip()
    search_config = data.get("searchConfig", {})
    
    # Get API key - check user config first, then global, then env
    api_key = get_gemini_api_key(user_id)
    if not api_key:
        return jsonify({"error": "GEMINI_API_KEY not configured. Set it in Config > API Keys."}), 400
    
    # Build system context wrapper
    target_count = search_config.get("target_source_count", "a comprehensive list")
    allowed_sources = ", ".join(search_config.get("allowed_sources", []))
    exclude_sources = ", ".join(search_config.get("exclude_sources", []))
    
    system_context = f"""
    SYSTEM CONTEXT:
    You are acting as a specialized Scientific Research Data Extraction Platform.
    
    PLATFORM CAPABILITIES:
    - You accept arbitrary custom search logic defined by the user.
    - You must execute the search strictly adhering to the user's logic.
    - You represent a deterministic, rule-based crawler agent.
    
    SEARCH CONFIGURATION (STRICT CONSTRAINTS):
    - **Target Source Count**: You must attempt to find and list at least {target_count} distinct relevant papers/sources.
    - **Allowed/Preferred Sources**: {allowed_sources if allowed_sources else "Any reputable scientific source"}.
    - **Exclude Sources**: {exclude_sources if exclude_sources else "None"}.

    KEY REQUIREMENT:
    - Your final report MUST NECESSARILY contain a comprehensive list of all LINKS (URLs) 
      that satisfy the filtering and selection criteria defined in the user's prompt.
    - Do not just summarize; explicitly list the sources found.
    
    USER'S CUSTOM SEARCH LOGIC:
    {query}
    """
    
    run_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    user_id = g.current_user["id"] if g.current_user else None
    
    try:
        # Start Deep Research
        result = deep_research_start(system_context, api_key)
        interaction_id = result.get("id", "")
        
        if not interaction_id:
            return jsonify({"error": "Failed to start research - no interaction ID received"}), 500
        
        # Save to database
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO deep_research_runs 
            (id, name, status, interaction_id, query, search_config, created_at, user_id)
            VALUES (?, ?, 'pending', ?, ?, ?, ?, ?)
        """, (run_id, name, interaction_id, query, json.dumps(search_config), now, user_id))
        conn.commit()
        conn.close()
        
        # Start background polling thread
        poll_thread = threading.Thread(
            target=poll_deep_research,
            args=(run_id, interaction_id, api_key),
            daemon=True
        )
        poll_thread.start()
        
        log_message(f"Deep Research started: {name}", "INFO")
        
        return jsonify({
            "id": run_id,
            "name": name,
            "status": "pending",
            "interactionId": interaction_id,
            "createdAt": now
        }), 201
        
    except Exception as e:
        log_message(f"Deep Research creation failed: {e}", "ERROR")
        return jsonify({"error": f"Failed to start research: {str(e)}"}), 500


@app.route("/deep-research/<run_id>", methods=["GET"])
@optional_auth
def get_deep_research_run(run_id):
    """Get details of a Deep Research run."""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM deep_research_runs WHERE id = ?", (run_id,))
    row = cur.fetchone()
    conn.close()
    
    if not row:
        return jsonify({"error": "Deep Research run not found"}), 404
    
    result = {
        "id": row["id"],
        "name": row["name"],
        "status": row["status"],
        "interactionId": row["interaction_id"],
        "query": row["query"],
        "searchConfig": json.loads(row["search_config"]) if row["search_config"] else {},
        "resultText": row["result_text"],
        "extractedLinks": json.loads(row["extracted_links"]) if row["extracted_links"] else {},
        "logs": row["logs"],
        "createdAt": row["created_at"],
        "startedAt": row["started_at"],
        "completedAt": row["completed_at"],
        "error": row["error"]
    }
    
    return jsonify(result)


@app.route("/deep-research/<run_id>/links", methods=["GET"])
@optional_auth
def get_deep_research_links(run_id):
    """Get extracted links from a Deep Research run."""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT extracted_links, status FROM deep_research_runs WHERE id = ?", (run_id,))
    row = cur.fetchone()
    
    if not row:
        conn.close()
        return jsonify({"error": "Deep Research run not found"}), 404
    
    if row["status"] != "completed":
        conn.close()
        return jsonify({"error": f"Research not completed. Status: {row['status']}"}), 400
    
    links_data = json.loads(row["extracted_links"]) if row["extracted_links"] else {}
    
    # Handle multiple possible structures
    links_list = []
    if isinstance(links_data, dict):
        links_list = links_data.get("extracted_links", []) or links_data.get("extractedLinks", []) or links_data.get("links", [])
    elif isinstance(links_data, list):
        links_list = links_data
    
    # If no links found but crawl jobs exist, reconstruct from crawl_jobs table
    if not links_list:
        cur.execute("""
            SELECT url, title FROM crawl_jobs 
            WHERE deep_research_id = ? 
            ORDER BY created_at ASC
        """, (run_id,))
        job_rows = cur.fetchall()
        if job_rows:
            links_list = [{"url": r["url"], "title": r["title"] or "", "relevance_score": 70} for r in job_rows]
    
    conn.close()
    return jsonify({"extractedLinks": links_list})


@app.route("/deep-research/<run_id>/report", methods=["GET"])
@optional_auth
def get_deep_research_report(run_id):
    """Get the full research report text."""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT result_text, status, name FROM deep_research_runs WHERE id = ?", (run_id,))
    row = cur.fetchone()
    conn.close()
    
    if not row:
        return jsonify({"error": "Deep Research run not found"}), 404
    
    if row["status"] != "completed":
        return jsonify({"error": f"Research not completed. Status: {row['status']}"}), 400
    
    return jsonify({
        "name": row["name"],
        "report": row["result_text"]
    })


@app.route("/deep-research/<run_id>/logs", methods=["GET"])
@optional_auth
def get_deep_research_logs(run_id):
    """Get logs for a Deep Research run."""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT logs, status FROM deep_research_runs WHERE id = ?", (run_id,))
    row = cur.fetchone()
    conn.close()
    
    if not row:
        return jsonify({"error": "Deep Research run not found"}), 404
    
    return jsonify({
        "status": row["status"],
        "logs": row["logs"] or ""
    })


@app.route("/deep-research/<run_id>", methods=["DELETE"])
@optional_auth
def delete_deep_research_run(run_id):
    """Delete a Deep Research run."""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM deep_research_runs WHERE id = ?", (run_id,))
    deleted = cur.rowcount
    conn.commit()
    conn.close()
    
    if deleted == 0:
        return jsonify({"error": "Deep Research run not found"}), 404
    
    log_message(f"Deep Research run deleted: {run_id}", "INFO")
    return "", 204


# ============================================================================
# Crawl Jobs API - Chrome Extension Integration
# ============================================================================

CLAIM_EXPIRY_SECONDS = 600  # 10 minutes

@app.route("/crawl/jobs", methods=["GET"])
@require_auth
def list_crawl_jobs():
    """List pending crawl jobs for the authenticated user.
    
    Query params:
    - limit: max jobs to return (default 10)
    - mode: 'peek' (just list) or 'claim' (auto-claim returned jobs)
    - maxClaimAgeSec: reset claims older than this (default 300)
    - includeScripts: if '1', include domain scripts in response
    - scriptsEtag: etag for scripts cache
    - since: ISO timestamp for incremental script sync
    """
    user_id = g.current_user["id"]
    limit = min(int(request.args.get("limit", 10)), 50)
    mode = request.args.get("mode", "peek")
    max_claim_age = int(request.args.get("maxClaimAgeSec", 300))
    include_scripts = request.args.get("includeScripts") == "1"
    deep_research_id = request.args.get("deepResearchId")
    status_filter = request.args.get("status")
    
    conn = get_db()
    cur = conn.cursor()
    now = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    
    # Reset expired claims back to PENDING (only when not filtering by deepResearchId)
    if not deep_research_id:
        expiry_threshold = (datetime.now(timezone.utc) - timedelta(seconds=max_claim_age)).isoformat().replace("+00:00", "Z")
        cur.execute("""
            UPDATE crawl_jobs 
            SET status = 'PENDING', claimed_at = NULL, claim_expires_at = NULL, attempts = attempts + 1
            WHERE user_id = ? AND status = 'CLAIMED' AND claimed_at < ?
        """, (user_id, expiry_threshold))
        conn.commit()
    
    # Build query based on filters
    query = """
        SELECT id, deep_research_id, run_id, url, title, status, attempts, created_at, completed_at, error
        FROM crawl_jobs 
        WHERE user_id = ?
    """
    params = [user_id]
    
    if deep_research_id:
        query += " AND deep_research_id = ?"
        params.append(deep_research_id)
    
    if status_filter:
        query += " AND status = ?"
        params.append(status_filter)
    elif not deep_research_id:
        query += " AND status = 'PENDING'"
    
    query += " ORDER BY created_at ASC LIMIT ?"
    params.append(limit)
    
    cur.execute(query, params)
    rows = cur.fetchall()
    
    jobs = []
    for row in rows:
        job = {
            "id": row["id"],
            "jobId": row["id"],
            "deepResearchId": row["deep_research_id"],
            "runId": row["run_id"],
            "run_id": row["run_id"],  # Also include snake_case for extension compatibility
            "url": row["url"],
            "title": row["title"],
            "status": row["status"],
            "attempts": row["attempts"],
            "createdAt": row["created_at"],
            "completedAt": row["completed_at"],
            "error": row["error"]
        }
        jobs.append(job)
    
    # Auto-claim if mode=claim
    if mode == "claim" and jobs:
        job_ids = [j["jobId"] for j in jobs]
        claim_expires = (datetime.now(timezone.utc) + timedelta(seconds=CLAIM_EXPIRY_SECONDS)).isoformat().replace("+00:00", "Z")
        placeholders = ",".join(["?" for _ in job_ids])
        cur.execute(f"""
            UPDATE crawl_jobs 
            SET status = 'CLAIMED', claimed_at = ?, claim_expires_at = ?
            WHERE id IN ({placeholders})
        """, [now, claim_expires] + job_ids)

        # Update corresponding sources status
        try:
            cur.execute(f"UPDATE sources SET status = 'PROCESSING', updated_at = ? WHERE id IN ({placeholders})", [now] + job_ids)
        except Exception:
            pass
        conn.commit()
        for job in jobs:
            job["status"] = "CLAIMED"
    
    # Get domain scripts if requested
    scripts = []
    scripts_etag = None
    if include_scripts:
        since = request.args.get("since", "0000-01-01T00:00:00Z")
        cur.execute("""
            SELECT id, domain, script, condition, wait_before_ms, wait_after_ms, created_at, updated_at
            FROM domain_scripts
            WHERE (user_id IS NULL OR user_id = ?) AND updated_at > ?
            ORDER BY domain
        """, (user_id, since))
        script_rows = cur.fetchall()
        for sr in script_rows:
            scripts.append({
                "domain": sr["domain"],
                "script": sr["script"],
                "condition": sr["condition"],
                "waitBeforeMs": sr["wait_before_ms"],
                "waitAfterMs": sr["wait_after_ms"],
                "hash": sr["id"],
                "createdAt": sr["created_at"],
                "updatedAt": sr["updated_at"]
            })
        # Simple etag based on count and latest update
        if scripts:
            scripts_etag = f"{len(scripts)}-{scripts[-1]['updatedAt']}"
    
    conn.close()
    
    response = {"jobs": jobs}
    if include_scripts:
        response["scripts"] = scripts
        if scripts_etag:
            response["scriptsEtag"] = scripts_etag
    
    return jsonify(response)


@app.route("/crawl/claim", methods=["POST"])
@require_auth
def claim_crawl_job():
    """Explicitly claim a specific job."""
    user_id = g.current_user["id"]
    data = request.json or {}
    job_id = data.get("jobId")
    
    if not job_id:
        return jsonify({"error": "jobId required"}), 400
    
    conn = get_db()
    cur = conn.cursor()
    now = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    claim_expires = (datetime.now(timezone.utc) + timedelta(seconds=CLAIM_EXPIRY_SECONDS)).isoformat().replace("+00:00", "Z")
    
    # Only claim if PENDING and belongs to user
    cur.execute("""
        UPDATE crawl_jobs 
        SET status = 'CLAIMED', claimed_at = ?, claim_expires_at = ?
        WHERE id = ? AND user_id = ? AND status = 'PENDING'
    """, (now, claim_expires, job_id, user_id))

    try:
        cur.execute("UPDATE sources SET status = 'PROCESSING', updated_at = ? WHERE id = ?", (now, job_id))
    except Exception:
        pass
    
    if cur.rowcount == 0:
        # Check if already claimed or doesn't exist
        cur.execute("SELECT status FROM crawl_jobs WHERE id = ? AND user_id = ?", (job_id, user_id))
        row = cur.fetchone()
        conn.close()
        if not row:
            return jsonify({"error": "Job not found", "status": "NOT_FOUND"}), 404
        return jsonify({"status": row["status"], "message": "Job not in PENDING state"})
    
    conn.commit()
    conn.close()
    
    return jsonify({"status": "CLAIMED", "jobId": job_id})


@app.route("/crawl/jobs/<job_id>/status", methods=["GET"])
@require_auth
def get_crawl_job_status(job_id):
    """Get status of a specific crawl job."""
    user_id = g.current_user["id"]
    
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, status, attempts, error, claimed_at, completed_at
        FROM crawl_jobs WHERE id = ? AND user_id = ?
    """, (job_id, user_id))
    row = cur.fetchone()
    conn.close()
    
    if not row:
        return jsonify({"error": "Job not found"}), 404
    
    return jsonify({
        "jobId": row["id"],
        "status": row["status"],
        "attempts": row["attempts"],
        "error": row["error"],
        "claimedAt": row["claimed_at"],
        "completedAt": row["completed_at"]
    })


@app.route("/crawl/jobs/<job_id>/reset", methods=["POST"])
@require_auth
def reset_crawl_job(job_id):
    """Reset a failed or stuck job back to PENDING."""
    user_id = g.current_user["id"]
    
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        UPDATE crawl_jobs 
        SET status = 'PENDING', claimed_at = NULL, claim_expires_at = NULL, error = NULL
        WHERE id = ? AND user_id = ? AND status IN ('CLAIMED', 'FAILED')
    """, (job_id, user_id))

    try:
        now = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
        cur.execute("UPDATE sources SET status = 'PENDING', error = NULL, updated_at = ? WHERE id = ?", (now, job_id))
    except Exception:
        pass
    
    if cur.rowcount == 0:
        conn.close()
        return jsonify({"error": "Job not found or not resettable"}), 404
    
    conn.commit()
    conn.close()
    
    return jsonify({"status": "PENDING", "jobId": job_id})


@app.route("/crawl/jobs/<job_id>/fail", methods=["POST"])
@require_auth
def fail_crawl_job(job_id):
    """Mark a job as FAILED with an error message."""
    user_id = g.current_user["id"]
    data = request.json or {}
    error = data.get("error", "Unknown error")
    
    conn = get_db()
    cur = conn.cursor()
    now = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    
    cur.execute("""
        UPDATE crawl_jobs 
        SET status = 'FAILED', error = ?, completed_at = ?
        WHERE id = ? AND user_id = ?
    """, (error[:500], now, job_id, user_id))
    
    # Also update the corresponding source
    try:
        cur.execute("UPDATE sources SET status = 'FAILED', error = ?, updated_at = ? WHERE id = ?", (error[:500], now, job_id))
    except Exception:
        pass
    
    if cur.rowcount == 0:
        conn.close()
        return jsonify({"error": "Job not found"}), 404
    
    conn.commit()
    conn.close()
    
    log_message(f"Crawl job {job_id} marked as FAILED: {error[:100]}", "WARN")
    
    return jsonify({"status": "FAILED", "jobId": job_id})


@app.route("/crawl/jobs/reset-all", methods=["POST"])
@require_auth
def reset_all_crawl_jobs():
    """Reset all CLAIMED/FAILED jobs back to PENDING, optionally filtered by deepResearchId."""
    user_id = g.current_user["id"]
    data = request.json or {}
    deep_research_id = data.get("deepResearchId")
    
    conn = get_db()
    cur = conn.cursor()
    
    if deep_research_id:
        cur.execute("""
            UPDATE crawl_jobs 
            SET status = 'PENDING', claimed_at = NULL, claim_expires_at = NULL, error = NULL
            WHERE user_id = ? AND deep_research_id = ? AND status IN ('CLAIMED', 'FAILED')
        """, (user_id, deep_research_id))
    else:
        cur.execute("""
            UPDATE crawl_jobs 
            SET status = 'PENDING', claimed_at = NULL, claim_expires_at = NULL, error = NULL
            WHERE user_id = ? AND status IN ('CLAIMED', 'FAILED')
        """, (user_id,))
    
    reset_count = cur.rowcount

    try:
        now = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
        if deep_research_id:
            cur.execute("SELECT id FROM crawl_jobs WHERE user_id = ? AND deep_research_id = ?", (user_id, deep_research_id))
            ids = [r["id"] for r in cur.fetchall()]
        else:
            cur.execute("SELECT id FROM crawl_jobs WHERE user_id = ?", (user_id,))
            ids = [r["id"] for r in cur.fetchall()]
        if ids:
            placeholders = ",".join(["?" for _ in ids])
            cur.execute(f"UPDATE sources SET status = 'PENDING', error = NULL, updated_at = ? WHERE id IN ({placeholders})", [now] + ids)
    except Exception:
        pass
    conn.commit()
    conn.close()
    
    return jsonify({"status": "ok", "resetCount": reset_count})


@app.route("/runs/<run_id>/skip-crawling", methods=["POST"])
@require_auth
def skip_crawling(run_id):
    """Skip remaining crawl jobs and proceed with extraction.
    
    Marks all PENDING/CLAIMED/PDF_PENDING jobs as SKIPPED and updates
    corresponding sources. Sets run status to 'waiting' so extraction can start.
    """
    user_id = g.current_user["id"]
    
    conn = get_db()
    cur = conn.cursor()
    now = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    
    # Verify run belongs to user
    cur.execute("SELECT id, status FROM runs WHERE id = ? AND user_id = ?", (run_id, user_id))
    run = cur.fetchone()
    if not run:
        conn.close()
        return jsonify({"error": "Run not found"}), 404
    
    # Mark all pending crawl jobs as SKIPPED
    cur.execute("""
        UPDATE crawl_jobs 
        SET status = 'SKIPPED', completed_at = ?, error = 'Skipped by user'
        WHERE run_id = ? AND user_id = ? AND status IN ('PENDING', 'CLAIMED', 'PDF_PENDING')
    """, (now, run_id, user_id))
    skipped_jobs = cur.rowcount
    
    # Update corresponding sources to SKIPPED
    cur.execute("""
        UPDATE sources 
        SET status = 'SKIPPED', error = 'Skipped by user', updated_at = ?
        WHERE run_id = ? AND status IN ('PENDING', 'PDF_PENDING')
    """, (now, run_id))
    skipped_sources = cur.rowcount
    
    # Update run status to 'waiting' (ready for extraction)
    cur.execute("UPDATE runs SET status = 'waiting' WHERE id = ?", (run_id,))
    
    conn.commit()
    conn.close()
    
    log_message(f"Crawling skipped: {skipped_jobs} jobs, {skipped_sources} sources marked as SKIPPED", "INFO", run_id)
    
    return jsonify({
        "status": "ok",
        "skippedJobs": skipped_jobs,
        "skippedSources": skipped_sources,
        "runStatus": "waiting"
    })


@app.route("/crawl/queue/stats", methods=["GET"])
@require_auth
def get_queue_stats():
    """Get crawl job queue statistics for the user."""
    user_id = g.current_user["id"]
    
    conn = get_db()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT 
            COUNT(*) as total,
            SUM(CASE WHEN status = 'PENDING' THEN 1 ELSE 0 END) as pending,
            SUM(CASE WHEN status = 'CLAIMED' THEN 1 ELSE 0 END) as claimed,
            SUM(CASE WHEN status = 'DONE' THEN 1 ELSE 0 END) as done,
            SUM(CASE WHEN status = 'FAILED' THEN 1 ELSE 0 END) as failed,
            SUM(CASE WHEN status = 'SKIPPED' THEN 1 ELSE 0 END) as skipped,
            SUM(CASE WHEN status = 'PDF_PENDING' THEN 1 ELSE 0 END) as pdf_pending
        FROM crawl_jobs 
        WHERE user_id = ?
    """, (user_id,))
    
    row = cur.fetchone()
    conn.close()
    
    return jsonify({
        "total": row["total"] or 0,
        "pending": row["pending"] or 0,
        "claimed": row["claimed"] or 0,
        "done": row["done"] or 0,
        "failed": row["failed"] or 0,
        "skipped": row["skipped"] or 0,
        "pdfPending": row["pdf_pending"] or 0
    })


@app.route("/crawl/queue/clear", methods=["POST"])
@require_auth
def clear_queue():
    """Clear crawl jobs from the queue.
    
    Body options:
    - { "status": "PENDING" } - clear only jobs with specific status
    - { "clearAll": true } - clear all jobs for user
    """
    user_id = g.current_user["id"]
    data = request.json or {}
    status_filter = data.get("status")
    clear_all = data.get("clearAll", False)
    
    conn = get_db()
    cur = conn.cursor()
    now = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    
    if clear_all:
        # Delete all jobs for user
        cur.execute("SELECT COUNT(*) FROM crawl_jobs WHERE user_id = ?", (user_id,))
        total = cur.fetchone()[0]
        cur.execute("DELETE FROM crawl_jobs WHERE user_id = ?", (user_id,))
        cleared = cur.rowcount
    elif status_filter:
        # Delete jobs with specific status
        cur.execute("SELECT COUNT(*) FROM crawl_jobs WHERE user_id = ? AND status = ?", (user_id, status_filter))
        total = cur.fetchone()[0]
        cur.execute("DELETE FROM crawl_jobs WHERE user_id = ? AND status = ?", (user_id, status_filter))
        cleared = cur.rowcount
    else:
        conn.close()
        return jsonify({"error": "Specify 'status' or 'clearAll'"}), 400
    
    # Get remaining count
    cur.execute("SELECT COUNT(*) FROM crawl_jobs WHERE user_id = ?", (user_id,))
    remaining = cur.fetchone()[0]
    
    conn.commit()
    conn.close()
    
    return jsonify({
        "cleared": cleared,
        "remaining": remaining
    })


@app.route("/crawl/jobs/fix-run-ids", methods=["POST"])
@require_auth
def fix_crawl_job_run_ids():
    """Fix crawl jobs that have null run_id by copying from deep_research_id."""
    user_id = g.current_user["id"]
    
    conn = get_db()
    cur = conn.cursor()
    
    cur.execute("""
        UPDATE crawl_jobs 
        SET run_id = deep_research_id
        WHERE user_id = ? AND run_id IS NULL AND deep_research_id IS NOT NULL
    """, (user_id,))
    
    fixed_count = cur.rowcount
    conn.commit()
    conn.close()
    
    return jsonify({"status": "ok", "fixedCount": fixed_count})


@app.route("/crawl/jobs/purge-pdfs", methods=["POST"])
@require_auth
def purge_pdf_crawl_jobs():
    """Delete all PDF crawl jobs - PDFs should use Surya server-side processing."""
    user_id = g.current_user["id"]
    deep_research_id = request.json.get("deepResearchId") if request.json else None
    
    conn = get_db()
    cur = conn.cursor()
    
    # Find and delete PDF jobs
    if deep_research_id:
        cur.execute("""
            DELETE FROM crawl_jobs 
            WHERE user_id = ? AND deep_research_id = ? AND (
                LOWER(url) LIKE '%.pdf' OR 
                LOWER(url) LIKE '%/pdf/%' OR 
                LOWER(url) LIKE '%pdf?%'
            )
        """, (user_id, deep_research_id))
    else:
        cur.execute("""
            DELETE FROM crawl_jobs 
            WHERE user_id = ? AND (
                LOWER(url) LIKE '%.pdf' OR 
                LOWER(url) LIKE '%/pdf/%' OR 
                LOWER(url) LIKE '%pdf?%'
            )
        """, (user_id,))
    
    deleted_count = cur.rowcount
    conn.commit()
    conn.close()
    
    return jsonify({"status": "ok", "deletedCount": deleted_count})


def process_pdf_jobs_for_run(run_id: str, user_id: str):
    """Background task to download and process all PDF_PENDING jobs for a run.
    
    Downloads PDFs directly from URLs, processes via Surya for text extraction,
    and creates source entries. This runs in a background thread.
    """
    import requests
    from urllib.parse import urlparse
    
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # Get all PDF_PENDING jobs for this run
        cur.execute("""
            SELECT id, url, title FROM crawl_jobs 
            WHERE run_id = ? AND user_id = ? AND status = 'PDF_PENDING'
        """, (run_id, user_id))
        
        pdf_jobs = cur.fetchall()
        
        if not pdf_jobs:
            log_message("No PDF jobs to process", "INFO", run_id)
            return
        
        log_message(f"Processing {len(pdf_jobs)} PDF job(s)", "INFO", run_id)
        
        for job in pdf_jobs:
            job_id = job["id"]
            url = job["url"]
            title = job["title"] or "PDF Document"
            
            try:
                log_message(f"Downloading PDF: {url[:100]}...", "INFO", run_id)
                
                # Download PDF
                response = requests.get(url, timeout=60, stream=True, headers={
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
                })
                response.raise_for_status()
                
                # Save PDF to run uploads folder
                pdf_id = str(uuid.uuid4())
                pdf_filename = f"{pdf_id}.pdf"
                run_upload_dir = os.path.join(UPLOAD_FOLDER, run_id)
                pdfs_dir = os.path.join(run_upload_dir, "pdfs")
                os.makedirs(pdfs_dir, exist_ok=True)
                pdf_path = os.path.join(pdfs_dir, pdf_filename)
                
                with open(pdf_path, 'wb') as f:
                    for chunk in response.iter_content(chunk_size=8192):
                        f.write(chunk)
                
                pdf_size = os.path.getsize(pdf_path)
                now = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
                
                # Register file (use file_type='pdf' so download works via /files/<id>/download)
                cur.execute("""
                    INSERT INTO files (id, filename, original_name, mime_type, size_bytes, file_type, run_id, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (pdf_id, pdf_filename, title + ".pdf", "application/pdf", pdf_size, "pdf", run_id, now))
                
                # Update job status
                cur.execute("""
                    UPDATE crawl_jobs 
                    SET status = 'DONE', pdf_path = ?, completed_at = ?, error = NULL
                    WHERE id = ?
                """, (pdf_path, now, job_id))
                
                conn.commit()
                
                log_message(f"PDF downloaded: {pdf_size} bytes", "SUCCESS", run_id)
                
                # Process PDF via Surya to extract text
                source_id = None
                try:
                    from pdf_converter import convert_pdf_to_text
                    
                    log_message(f"Processing PDF via Surya: {pdf_path}", "INFO", run_id)
                    pdf_text = convert_pdf_to_text(pdf_path, use_cache=True)
                    
                    if pdf_text and len(pdf_text.strip()) > 100:
                        source_id = str(uuid.uuid4())
                        domain = ""
                        try:
                            domain = urlparse(url).netloc
                        except:
                            pass
                        
                        # Wrap text in basic HTML structure
                        html_content = f"""<!DOCTYPE html>
<html>
<head><title>{title}</title></head>
<body>
<source>
<h1>{title}</h1>
<div class="pdf-content">
{pdf_text}
</div>
</source>
</body>
</html>"""
                        
                        # Update existing source row created at input time
                        cur.execute(
                            """
                            UPDATE sources
                            SET domain = ?, title = ?, html_content = ?, pdf_file_id = ?, source_type = 'pdf', content_type = 'pdf', status = 'READY', error = NULL, updated_at = ?
                            WHERE id = ?
                            """,
                            (domain, title, html_content, pdf_id, now, job_id),
                        )
                        conn.commit()
                        
                        log_message(f"PDF processed via Surya: {len(pdf_text)} chars extracted", "SUCCESS", run_id)
                    else:
                        log_message(f"PDF processing yielded insufficient content", "WARN", run_id)
                except Exception as e:
                    log_message(f"PDF Surya processing failed: {str(e)}", "ERROR", run_id)
                
            except Exception as e:
                error_msg = str(e)
                log_message(f"PDF download failed for {url[:100]}: {error_msg}", "ERROR", run_id)
                
                # Mark job as failed
                now = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
                cur.execute("""
                    UPDATE crawl_jobs 
                    SET status = 'FAILED', completed_at = ?, error = ?
                    WHERE id = ?
                """, (now, error_msg[:500], job_id))

                try:
                    cur.execute("UPDATE sources SET status = 'FAILED', error = ?, updated_at = ? WHERE id = ?", (error_msg[:500], now, job_id))
                except Exception:
                    pass
                conn.commit()
        
        # Check if all jobs complete (HTML + PDF)
        check_and_update_run_crawl_status(conn, run_id, user_id)
        
        log_message(f"PDF processing complete for run", "SUCCESS", run_id)
        
    except Exception as e:
        log_message(f"PDF processing thread error: {str(e)}", "ERROR", run_id)
    finally:
        if conn:
            conn.close()


def check_and_update_run_crawl_status(conn, run_id, user_id):
    """Check if all crawl jobs for a run are complete and update run status.
    
    When all HTML crawl jobs are DONE, updates run status from 'crawling' to 'waiting'.
    This signals that the run is ready for extraction to start.
    """
    if not run_id:
        return
    
    cur = conn.cursor()
    
    # Get run info
    cur.execute("SELECT status, source_type FROM runs WHERE id = ?", (run_id,))
    run = cur.fetchone()
    
    if not run or run["status"] != "crawling":
        return
    
    # Count all job statuses - PDFs are now processed server-side
    cur.execute("""
        SELECT 
            COUNT(*) as total,
            SUM(CASE WHEN status IN ('PENDING', 'CLAIMED') THEN 1 ELSE 0 END) as html_pending,
            SUM(CASE WHEN status = 'PDF_PENDING' THEN 1 ELSE 0 END) as pdf_pending,
            SUM(CASE WHEN status = 'DONE' THEN 1 ELSE 0 END) as done,
            SUM(CASE WHEN status = 'FAILED' THEN 1 ELSE 0 END) as failed
        FROM crawl_jobs 
        WHERE run_id = ? AND user_id = ?
    """, (run_id, user_id))
    
    counts = cur.fetchone()
    total = counts["total"] or 0
    html_pending = counts["html_pending"] or 0
    pdf_pending = counts["pdf_pending"] or 0
    done = counts["done"] or 0
    failed = counts["failed"] or 0
    
    # All jobs complete when no HTML pending AND no PDF pending
    # (PDFs are processed server-side, so PDF_PENDING means still downloading)
    if html_pending == 0 and pdf_pending == 0 and total > 0:
        # Update run status to 'waiting' (ready for extraction)
        cur.execute("UPDATE runs SET status = 'waiting' WHERE id = ?", (run_id,))
        conn.commit()
        log_message(f"All crawl jobs complete ({done} done, {failed} failed). Run ready for extraction.", "SUCCESS", run_id)


@app.route("/crawl/result", methods=["POST"])
@require_auth
def submit_crawl_result():
    """Receive crawled HTML from extension."""
    user_id = g.current_user["id"]
    data = request.json or {}
    job_id = data.get("jobId")
    html = data.get("html", "")
    
    if not job_id:
        return jsonify({"error": "jobId required"}), 400
    
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        now = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
        
        # Verify job belongs to user and is claimed
        cur.execute("SELECT id, url, deep_research_id, run_id FROM crawl_jobs WHERE id = ? AND user_id = ?", (job_id, user_id))
        job = cur.fetchone()
        
        if not job:
            return jsonify({"error": "Job not found"}), 404
        
        # Update job with result
        cur.execute("""
            UPDATE crawl_jobs 
            SET status = 'DONE', html = ?, completed_at = ?, error = NULL
            WHERE id = ?
        """, (html, now, job_id))
        
        domain = ""
        try:
            from urllib.parse import urlparse
            domain = urlparse(job["url"]).netloc
        except:
            pass

        # Update existing source row created at input time
        cur.execute(
            """
            UPDATE sources
            SET domain = ?, html_content = ?, source_type = 'link', content_type = 'html', status = 'READY', error = NULL, updated_at = ?
            WHERE id = ?
            """,
            (domain, html, now, job_id),
        )
        
        conn.commit()
        
        log_message(f"Crawl result received for job {job_id}, {len(html)} bytes", "INFO")
        
        # Check if all crawl jobs for this run are complete
        check_and_update_run_crawl_status(conn, job["run_id"], user_id)
        
        return jsonify({"status": "DONE", "jobId": job_id, "sourceId": job_id})
    except Exception as e:
        log_message(f"Error in submit_crawl_result: {str(e)}", "ERROR")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()


@app.route("/crawl/result/pdf", methods=["POST"])
@require_auth
def submit_crawl_result_pdf():
    """Receive PDF URL from extension - server downloads and processes the PDF.
    
    When extension detects a PDF (by content-type or URL), it submits the URL here
    instead of trying to extract HTML. Server downloads the PDF and processes it.
    """
    user_id = g.current_user["id"]
    data = request.json or {}
    job_id = data.get("jobId")
    pdf_url = data.get("url") or data.get("pdfUrl")
    
    if not job_id:
        return jsonify({"error": "jobId required"}), 400
    if not pdf_url:
        return jsonify({"error": "url required"}), 400
    
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        now = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
        
        # Verify job belongs to user
        cur.execute("SELECT id, url, run_id, title FROM crawl_jobs WHERE id = ? AND user_id = ?", (job_id, user_id))
        job = cur.fetchone()
        
        if not job:
            return jsonify({"error": "Job not found"}), 404
        
        run_id = job["run_id"]
        title = job["title"] or "PDF Document"
        
        # Mark job as PDF_PENDING for server-side download
        cur.execute("""
            UPDATE crawl_jobs 
            SET status = 'PDF_PENDING', url = ?
            WHERE id = ?
        """, (pdf_url, job_id))
        
        # Update source to PDF type
        cur.execute("""
            UPDATE sources
            SET source_type = 'pdf', status = 'PDF_PENDING', url = ?, updated_at = ?
            WHERE id = ?
        """, (pdf_url, now, job_id))
        
        conn.commit()
        
        log_message(f"PDF URL received for job {job_id}: {pdf_url[:100]}...", "INFO", run_id)
        
        # Start background download for this single PDF
        def download_single_pdf():
            import requests as req
            conn2 = None
            try:
                conn2 = get_db()
                cur2 = conn2.cursor()
                
                log_message(f"Downloading PDF: {pdf_url[:100]}...", "INFO", run_id)
                
                response = req.get(pdf_url, timeout=120, stream=True, headers={
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
                })
                response.raise_for_status()
                
                # Save PDF
                pdf_id = str(uuid.uuid4())
                pdf_filename = f"{pdf_id}.pdf"
                run_upload_dir = os.path.join(UPLOAD_FOLDER, run_id)
                pdfs_dir = os.path.join(run_upload_dir, "pdfs")
                os.makedirs(pdfs_dir, exist_ok=True)
                pdf_path = os.path.join(pdfs_dir, pdf_filename)
                
                with open(pdf_path, 'wb') as f:
                    for chunk in response.iter_content(chunk_size=8192):
                        f.write(chunk)
                
                pdf_size = os.path.getsize(pdf_path)
                dl_now = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
                
                # Register file
                cur2.execute("""
                    INSERT INTO files (id, filename, original_name, mime_type, size_bytes, file_type, run_id, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (pdf_id, pdf_filename, title + ".pdf", "application/pdf", pdf_size, "pdf", run_id, dl_now))
                
                # Update job status
                cur2.execute("""
                    UPDATE crawl_jobs 
                    SET status = 'DONE', pdf_path = ?, completed_at = ?, error = NULL
                    WHERE id = ?
                """, (pdf_path, dl_now, job_id))
                
                # Update source with PDF file reference
                cur2.execute("""
                    UPDATE sources
                    SET pdf_file_id = ?, status = 'READY', updated_at = ?
                    WHERE id = ?
                """, (pdf_id, dl_now, job_id))
                
                conn2.commit()
                log_message(f"PDF downloaded: {pdf_size} bytes", "SUCCESS", run_id)
                
                # Check if all jobs complete
                check_and_update_run_crawl_status(conn2, run_id, user_id)
                
            except Exception as e:
                log_message(f"PDF download failed: {str(e)}", "ERROR", run_id)
                if conn2:
                    cur2 = conn2.cursor()
                    cur2.execute("UPDATE crawl_jobs SET status = 'FAILED', error = ? WHERE id = ?", (str(e), job_id))
                    cur2.execute("UPDATE sources SET status = 'FAILED', error = ? WHERE id = ?", (str(e), job_id))
                    conn2.commit()
            finally:
                if conn2:
                    conn2.close()
        
        import threading
        threading.Thread(target=download_single_pdf, daemon=True).start()
        
        return jsonify({"status": "PDF_PENDING", "jobId": job_id, "message": "PDF download started"})
    except Exception as e:
        log_message(f"Error in submit_crawl_result_pdf: {str(e)}", "ERROR")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()


@app.route("/crawl/result/pdf-binary", methods=["POST"])
@require_auth
def submit_crawl_result_pdf_binary():
    """Receive PDF binary content directly from extension.
    
    Extension fetches the PDF binary and uploads it here, avoiding browser downloads.
    Expects multipart form data with:
    - jobId: the crawl job ID
    - pdfFile: the PDF binary file
    - url: original PDF URL (optional, for logging)
    """
    user_id = g.current_user["id"]
    job_id = request.form.get("jobId")
    pdf_url = request.form.get("url", "")
    
    if not job_id:
        return jsonify({"error": "jobId required"}), 400
    
    if "pdfFile" not in request.files:
        return jsonify({"error": "pdfFile required"}), 400
    
    pdf_file = request.files["pdfFile"]
    
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        now = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
        
        # Verify job belongs to user
        cur.execute("SELECT id, url, run_id, title FROM crawl_jobs WHERE id = ? AND user_id = ?", (job_id, user_id))
        job = cur.fetchone()
        
        if not job:
            return jsonify({"error": "Job not found"}), 404
        
        run_id = job["run_id"]
        title = job["title"] or "PDF Document"
        original_url = pdf_url or job["url"]
        
        log_message(f"Receiving PDF binary for job {job_id} ({len(pdf_file.read())} bytes)", "INFO", run_id)
        pdf_file.seek(0)  # Reset after reading length
        
        # Save PDF
        pdf_id = str(uuid.uuid4())
        pdf_filename = f"{pdf_id}.pdf"
        run_upload_dir = os.path.join(UPLOAD_FOLDER, run_id)
        pdfs_dir = os.path.join(run_upload_dir, "pdfs")
        os.makedirs(pdfs_dir, exist_ok=True)
        pdf_path = os.path.join(pdfs_dir, pdf_filename)
        
        pdf_file.save(pdf_path)
        pdf_size = os.path.getsize(pdf_path)
        
        # Register file
        cur.execute("""
            INSERT INTO files (id, filename, original_name, mime_type, size_bytes, file_type, run_id, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (pdf_id, pdf_filename, title + ".pdf", "application/pdf", pdf_size, "pdf", run_id, now))
        
        # Update job status to DONE
        cur.execute("""
            UPDATE crawl_jobs 
            SET status = 'DONE', pdf_path = ?, completed_at = ?, error = NULL
            WHERE id = ?
        """, (pdf_path, now, job_id))
        
        # Update source with PDF file reference
        cur.execute("""
            UPDATE sources
            SET pdf_file_id = ?, source_type = 'pdf', status = 'READY', url = ?, updated_at = ?
            WHERE id = ?
        """, (pdf_id, original_url, now, job_id))
        
        conn.commit()
        log_message(f"PDF received and saved: {pdf_size} bytes from extension", "SUCCESS", run_id)
        
        # Check if all jobs complete
        check_and_update_run_crawl_status(conn, run_id, user_id)
        
        return jsonify({"status": "DONE", "jobId": job_id, "fileId": pdf_id, "size": pdf_size})
    except Exception as e:
        log_message(f"Error in submit_crawl_result_pdf_binary: {str(e)}", "ERROR")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()


@app.route("/runs/<run_id>/logs/append", methods=["POST"])
@optional_auth
def append_run_logs(run_id):
    """Append logs from extension to a run's log file."""
    data = request.json or {}
    entries = data.get("entries", [])
    
    if not entries:
        return jsonify({"status": "ok", "appended": 0})
    
    conn = get_db()
    cur = conn.cursor()
    now = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    
    for entry in entries:
        level = entry.get("level", "INFO")
        message = entry.get("message", "")
        source = entry.get("source", "extension")
        context = json.dumps(entry.get("context")) if entry.get("context") else None
        
        cur.execute("""
            INSERT INTO logs (created_at, level, message, run_id, source, context)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (now, level, f"[{source}] {message}", run_id, source, context))
    
    conn.commit()
    conn.close()
    
    return jsonify({"status": "ok", "appended": len(entries)})


@app.route("/crawl/scripts", methods=["GET"])
@require_auth
def list_domain_scripts():
    """List all domain scripts for the user."""
    user_id = g.current_user["id"]
    
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, domain, script, condition, wait_before_ms, wait_after_ms, created_at, updated_at
        FROM domain_scripts
        WHERE user_id IS NULL OR user_id = ?
        ORDER BY domain
    """, (user_id,))
    rows = cur.fetchall()
    conn.close()
    
    scripts = []
    for row in rows:
        scripts.append({
            "id": row["id"],
            "domain": row["domain"],
            "script": row["script"],
            "condition": row["condition"],
            "waitBeforeMs": row["wait_before_ms"],
            "waitAfterMs": row["wait_after_ms"],
            "createdAt": row["created_at"],
            "updatedAt": row["updated_at"]
        })
    
    return jsonify(scripts)


@app.route("/crawl/scripts", methods=["POST"])
@require_auth
def upsert_domain_script():
    """Create or update a domain script."""
    user_id = g.current_user["id"]
    data = request.json or {}
    
    domain = data.get("domain", "").lower().strip()
    if not domain:
        return jsonify({"error": "domain required"}), 400
    
    script = data.get("script", "")
    condition = data.get("condition", "")
    wait_before = int(data.get("waitBeforeMs", 0))
    wait_after = int(data.get("waitAfterMs", 0))
    
    conn = get_db()
    cur = conn.cursor()
    now = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    
    # Check if exists
    cur.execute("SELECT id FROM domain_scripts WHERE domain = ? AND (user_id = ? OR user_id IS NULL)", (domain, user_id))
    existing = cur.fetchone()
    
    if existing:
        cur.execute("""
            UPDATE domain_scripts 
            SET script = ?, condition = ?, wait_before_ms = ?, wait_after_ms = ?, updated_at = ?
            WHERE id = ?
        """, (script, condition, wait_before, wait_after, now, existing["id"]))
        script_id = existing["id"]
    else:
        script_id = str(uuid.uuid4())
        cur.execute("""
            INSERT INTO domain_scripts (id, domain, user_id, script, condition, wait_before_ms, wait_after_ms, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (script_id, domain, user_id, script, condition, wait_before, wait_after, now, now))
    
    conn.commit()
    conn.close()
    
    return jsonify({
        "id": script_id,
        "domain": domain,
        "script": script,
        "condition": condition,
        "waitBeforeMs": wait_before,
        "waitAfterMs": wait_after
    })


@app.route("/crawl/scripts/<script_id>", methods=["DELETE"])
@require_auth
def delete_domain_script(script_id):
    """Delete a domain script."""
    user_id = g.current_user["id"]
    
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM domain_scripts WHERE id = ? AND user_id = ?", (script_id, user_id))
    deleted = cur.rowcount
    conn.commit()
    conn.close()
    
    if deleted == 0:
        return jsonify({"error": "Script not found"}), 404
    
    return "", 204


# ============================================================================
# Health Check
# ============================================================================

@app.route("/health", methods=["GET"])
def health():
    """Health check endpoint."""
    return jsonify({"status": "ok", "timestamp": datetime.now(timezone.utc).isoformat()})

# ============================================================================
# Main
# ============================================================================

if __name__ == "__main__":
    log_message("Server starting...", "INFO")
    # Disable reloader to prevent server restart when files are uploaded to uploads/ folder
    # This caused CONNECTION_RESET errors for browser file uploads
    app.run(host="0.0.0.0", port=5007, debug=True, threaded=True, use_reloader=False)

