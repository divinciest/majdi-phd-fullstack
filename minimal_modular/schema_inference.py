"""Schema inference from Excel files with caching."""
import hashlib
import json
import re
from typing import List
from openpyxl import load_workbook
from cache_utils import get_schema_cache, set_schema_cache
from llm_client import call_openai


def infer_schema_from_excel(excel_path: str, instructions: str = "", use_cache: bool = True) -> dict:
    """Read Excel headers and create schema object.
    
    Args:
        excel_path: Path to Excel file with headers in first row
        use_cache: Whether to use cached results (default: True)
        
    Returns:
        Schema dict with 'title' and 'fields' array
    """
    wb = load_workbook(excel_path)
    ws = wb.active

    engine_context = (
        "You extract structured experimental concrete research data. "
        "Your schema must support concrete mix design and chloride migration testing context (e.g., NT BUILD 492). "
        "Use stable, single-line canonical field names, keep units when present, and do not invent fields."
    )

    context_hash = hashlib.sha256((engine_context + "\n" + (instructions or "")).encode("utf-8")).hexdigest()

    # Check cache first (only valid if context matches)
    if use_cache:
        cached = get_schema_cache(excel_path)
        if cached is not None and cached.get("schemaContextHash") == context_hash:
            return cached
    
    def _collapse_ws(s: str) -> str:
        return re.sub(r"\s+", " ", (s or "").strip())

    def _deterministic_canonicalize(headers: List[str]) -> List[str]:
        out = []
        used = set()
        for h in headers:
            c = _collapse_ws(h)
            if not c:
                c = "Column"
            base = c
            i = 2
            while c in used:
                c = f"{base} ({i})"
                i += 1
            used.add(c)
            out.append(c)
        return out

    def _parse_json_from_llm_text(content: str):
        content = (content or "").strip()
        if "```json" in content:
            content = content.split("```json", 1)[1].split("```", 1)[0].strip()
        elif "```" in content:
            content = content.split("```", 1)[1].split("```", 1)[0].strip()
        return json.loads(content)

    def _build_schema_with_llm(headers: List[str], title: str) -> dict:
        system_prompt = (
            "You are a scientific schema engineering assistant for concrete research extraction. "
            "Your job is to: (1) standardize Excel column headers into canonical field names, and "
            "(2) write a concise field description for each field to guide downstream extraction. "
            "Return ONLY valid JSON."
        )
        user_prompt = (
            "ENGINE CONTEXT:\n"
            + engine_context
            + "\n\nUSER INSTRUCTIONS (may contain additional constraints/preferences):\n"
            + (instructions or "")
            + "\n\n"
            "Given these Excel headers, produce a JSON schema with field descriptions.\n\n"
            "INPUT HEADERS (order is authoritative):\n"
            + json.dumps(headers, ensure_ascii=False)
            + "\n\nOUTPUT RULES:\n"
            "- Return a JSON object with keys: title, fields, fieldMapping\n"
            "- fields must be an array of objects, same length and order as input\n"
            "- Each field object must have: name, type, description\n"
            "- type must be 'string' for all fields\n"
            "- name must be a canonicalized version of the header; must be unique\n"
            "- description must be short and specific (1 sentence) and mention units if present\n"
            "- fieldMapping must map canonical name -> original header (exact text)\n"
            "- Do not invent or drop fields\n"
        )

        resp = call_openai(system_prompt=system_prompt, user_prompt=user_prompt, use_cache=True)
        obj = _parse_json_from_llm_text(resp["choices"][0]["message"]["content"])

        if not isinstance(obj, dict):
            raise ValueError("Invalid schema output")
        fields = obj.get("fields")
        mapping = obj.get("fieldMapping")
        if not isinstance(fields, list) or len(fields) != len(headers):
            raise ValueError("Invalid fields")
        if not isinstance(mapping, dict) or len(mapping) != len(headers):
            raise ValueError("Invalid fieldMapping")

        used = set()
        normalized_fields = []
        for i, f in enumerate(fields):
            if not isinstance(f, dict):
                raise ValueError("Invalid field entry")
            name = _collapse_ws(str(f.get("name", "")))
            ftype = str(f.get("type", "string") or "string").strip().lower()
            desc = _collapse_ws(str(f.get("description", "")))
            if not name:
                raise ValueError("Empty canonical field name")
            if name in used:
                raise ValueError("Duplicate canonical field name")
            if ftype != "string":
                raise ValueError("Invalid field type")
            if not desc:
                raise ValueError("Empty field description")
            used.add(name)
            normalized_fields.append({"name": name, "type": "string", "description": desc})

        normalized_mapping = {}
        for i, canonical in enumerate([f["name"] for f in normalized_fields]):
            raw = mapping.get(canonical)
            if raw is None:
                raise ValueError("fieldMapping missing canonical key")
            raw = str(raw)
            if raw != headers[i]:
                raise ValueError("fieldMapping order mismatch")
            normalized_mapping[canonical] = raw

        return {
            "title": obj.get("title") or title,
            "fields": normalized_fields,
            "canonicalized": True,
            "fieldMapping": normalized_mapping,
            "schemaVersion": "canon_desc_v1",
            "schemaContextHash": context_hash,
        }

    # Extract headers from first row
    raw_headers = []
    for cell in ws[1]:
        val = str(cell.value or "").strip()
        if val:
            raw_headers.append(val)

    try:
        schema = _build_schema_with_llm(raw_headers, ws.title)
    except Exception:
        canonical_headers = _deterministic_canonicalize(raw_headers)
        fields = [
            {
                "name": canonical_headers[i],
                "type": "string",
                "description": f"Extract '{canonical_headers[i]}' exactly as reported in the paper; use Missing if not found.",
            }
            for i in range(len(canonical_headers))
        ]
        schema = {
            "title": ws.title,
            "fields": fields,
            "canonicalized": True,
            "fieldMapping": {canonical_headers[i]: raw_headers[i] for i in range(len(raw_headers))},
            "schemaVersion": "canon_desc_v1",
            "schemaContextHash": context_hash,
        }
    
    # Cache the result
    if use_cache:
        set_schema_cache(excel_path, schema)
     
    return schema
